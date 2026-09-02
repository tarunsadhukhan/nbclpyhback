"""
Seed worker_rate_mst from the mill's WORKER RATE MUSTER.xlsx (one-off loader).

ECODE is matched to hrms_ed_official_details.emp_code (active=1) to resolve
eb_id; rows whose ECODE has no employee are reported and skipped. Re-runnable:
a worker that already has an active rate row is skipped.

    python dbqueries/migrations/seed_worker_rate_mst_from_excel.py <xlsx> [tenant_db]
"""

import os
import sys

import openpyxl
import pymysql
from dotenv import load_dotenv

FLAGS = ("DA_ALL", "HRA", "HRD", "QUARTER", "PF", "ESI", "PTAX")


def yn(v) -> str:
    return "Y" if str(v or "").strip().upper() == "Y" else "N"


def num(v):
    return None if v in (None, "") else float(v)


def main(xlsx: str, db: str) -> None:
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", "..", "env", "database.env"))
    conn = pymysql.connect(
        host=os.environ["DATABASE_HOST"], port=int(os.environ["DATABASE_PORT"]),
        user=os.environ["DATABASE_USER"], password=os.environ["DATABASE_PASSWORD"],
        database=db, cursorclass=pymysql.cursors.DictCursor,
    )
    cur = conn.cursor()
    cur.execute("SELECT emp_code, eb_id FROM hrms_ed_official_details WHERE active = 1 AND emp_code IS NOT NULL")
    eb_by_code = {str(r["emp_code"]).strip(): r["eb_id"] for r in cur.fetchall()}
    cur.execute("SELECT eb_id FROM worker_rate_mst WHERE is_active = 1")
    existing = {r["eb_id"] for r in cur.fetchall()}

    ws = openpyxl.load_workbook(xlsx, data_only=True)["Sheet1"]
    hdr = [str(c.value).strip() for c in ws[1]]
    col = {h: i for i, h in enumerate(hdr)}

    inserted, skipped_dup, unmatched = 0, 0, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[col["ECODE"]] or "").strip()
        if not code:
            continue
        eb_id = eb_by_code.get(code)
        if eb_id is None:
            unmatched.append((code, row[col["ENAME"]]))
            continue
        if eb_id in existing:
            skipped_dup += 1
            continue
        cur.execute(
            "INSERT INTO worker_rate_mst (eb_id, fbasic, fbasic_hr, da_all, da_rate, hra, hrd, quarter, pf, esi, ptax, is_active) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)",
            (
                eb_id, num(row[col["FBASIC"]]), num(row[col["FBASIC_HR"]]),
                yn(row[col["DA_ALL"]]), num(row[col["DA_RATE"]]),
                *[yn(row[col[f]]) for f in FLAGS[1:]],
            ),
        )
        existing.add(eb_id)
        inserted += 1

    conn.commit()
    conn.close()
    print(f"inserted={inserted} skipped_existing={skipped_dup} unmatched={len(unmatched)}")
    for code, name in unmatched:
        print(f"  UNMATCHED ECODE {code} ({name})")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "nbcl")
