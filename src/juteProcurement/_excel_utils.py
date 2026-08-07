"""
Shared helpers for jute module excel-download endpoints.

Extracted from src/juteProcurement/mr.py so that the same parse/build pattern
can be reused across all jute index-page download endpoints.
"""

from datetime import datetime, date
from io import BytesIO
from typing import Iterable, Mapping, Sequence, Tuple, Optional

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook


MAX_DOWNLOAD_ROWS = 50_000

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def parse_iso_date(value: Optional[str], field_name: str) -> Optional[date]:
    """Parse a YYYY-MM-DD query string into a date, or raise HTTPException(400)."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid {field_name}; expected YYYY-MM-DD",
        )


def _coerce_value(value):
    """Convert date/datetime to ISO strings; pass everything else through."""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return str(value)
    return value


def build_excel_response(
    rows: Iterable[Mapping],
    columns: Sequence[Tuple[str, str]],
    sheet_title: str,
    filename_prefix: str,
    from_date_str: Optional[str] = None,
    to_date_str: Optional[str] = None,
) -> StreamingResponse:
    """
    Build an .xlsx StreamingResponse from row dicts.

    Args:
        rows: iterable of mapping rows (each like dict(r._mapping))
        columns: list of (header_label, dict_key) tuples — defines order + headers
        sheet_title: workbook sheet name
        filename_prefix: download filename prefix (e.g. "jute_po")
        from_date_str / to_date_str: optional original ISO strings used to build
            the filename suffix; both empty => "_all"
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col_idx, (header, _) in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_, key) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_coerce_value(row.get(key)))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    xlsx_bytes = bio.getvalue()

    suffix_parts = []
    if from_date_str:
        suffix_parts.append(from_date_str)
    if to_date_str:
        suffix_parts.append(to_date_str)
    suffix = "_".join(suffix_parts) if suffix_parts else "all"
    filename = f"{filename_prefix}_{suffix}.xlsx"

    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def enforce_row_cap(total: int) -> None:
    """Raise HTTPException(400) if total exceeds MAX_DOWNLOAD_ROWS."""
    if total > MAX_DOWNLOAD_ROWS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Result set too large — max {MAX_DOWNLOAD_ROWS} rows, "
                f"got {total}. Narrow the date range or search term."
            ),
        )
