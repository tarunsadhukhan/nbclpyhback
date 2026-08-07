"""
Jute Procurement Report endpoints.
Provides endpoints for jute stock and related reports.
"""

import logging
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from src.config.db import get_tenant_db
from src.authorization.utils import get_current_user_with_refresh
from src.juteProcurement.reportQueries import (
    get_jute_stock_report_query,
    get_jute_qty_wise_report_query,
    get_jute_txn_summary_query,
    get_jute_period_wise_query,
    get_jute_mr_in_stock_query,
    get_jute_mr_wise_query,
    get_jute_godown_wise_query,
    get_jute_with_value_query,
    get_jute_percent_claims_query,
    get_jute_mukham_moisture_query,
    get_batch_cost_report_query,
    get_mr_list_query,
    get_mr_list_count_query,
    get_jute_tally_download_query,
    get_jute_tally_download_items_per_mr_query,
    get_jute_tally_check_list_query,
)
from datetime import datetime, date
from openpyxl import Workbook

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/stock")
def get_jute_stock_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Get daily jute stock position report for a given branch and date.

    Returns opening stock, receipt, issue, closing stock, and MTD receipt/issue
    grouped by item group and item.

    Query params:
    - branch_id: Branch ID (required)
    - date: Report date in YYYY-MM-DD format (required)
    """
    try:
        q_branch_id = request.query_params.get("branch_id")
        q_date = request.query_params.get("date")

        if not q_branch_id:
            raise HTTPException(status_code=400, detail="branch_id is required")
        if not q_date:
            raise HTTPException(status_code=400, detail="date is required")

        try:
            branch_id = int(q_branch_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid branch_id")

        query = get_jute_stock_report_query()
        rows = db.execute(query, {
            "branch_id": branch_id,
            "report_date": q_date,
        }).fetchall()

        data = [dict(r._mapping) for r in rows]

        return {"data": data}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute stock report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute stock report: {str(e)}",
        )


@router.get("/batch-cost")
def get_batch_cost_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Get batch cost report: yarn quality-wise planned vs actual jute issue.

    Compares planned issue (from batch daily assignments + batch plan percentages)
    against actual issue for a given branch and date. Returns planned weight,
    actual weight, average rate, issue value, and variance per quality.

    Query params:
    - branch_id: Branch ID (required)
    - date: Report date in YYYY-MM-DD format (required)
    """
    try:
        q_branch_id = request.query_params.get("branch_id")
        q_date = request.query_params.get("date")

        if not q_branch_id:
            raise HTTPException(status_code=400, detail="branch_id is required")
        if not q_date:
            raise HTTPException(status_code=400, detail="date is required")

        try:
            branch_id = int(q_branch_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid branch_id")

        query = get_batch_cost_report_query()
        rows = db.execute(query, {
            "branch_id": branch_id,
            "report_date": q_date,
        }).fetchall()

        data = [dict(r._mapping) for r in rows]

        return {"data": data}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching batch cost report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching batch cost report: {str(e)}",
        )


def _parse_iso_date(value: str, field_name: str) -> str:
    """Validate a YYYY-MM-DD date string. Raises HTTPException(400) if invalid."""
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid {field_name} format, expected YYYY-MM-DD",
        )
    return value


@router.get("/mr-list")
def get_mr_list_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Jute MR List report (header-only).

    Lists approved/finalised Jute MR headers within a date range.

    Query params:
    - co_id: Company ID (required, int)
    - branch_id: Branch ID (optional, int)
    - date_from: Start date YYYY-MM-DD (required)
    - date_to: End date YYYY-MM-DD (required)
    - search: Optional search term (party name, vehicle_no, challan_no,
      formatted MR number)
    - page: Page number (default 1)
    - limit: Page size (default 50, max 10000)

    Returns:
        {"data": [row, ...], "total": <int>}
    """
    try:
        q_co_id = request.query_params.get("co_id")
        q_branch_id = request.query_params.get("branch_id")
        q_date_from = request.query_params.get("date_from")
        q_date_to = request.query_params.get("date_to")
        q_search = request.query_params.get("search")
        q_page = request.query_params.get("page")
        q_limit = request.query_params.get("limit")

        if not q_co_id:
            raise HTTPException(status_code=400, detail="co_id is required")
        if not q_date_from:
            raise HTTPException(status_code=400, detail="date_from is required")
        if not q_date_to:
            raise HTTPException(status_code=400, detail="date_to is required")

        try:
            co_id = int(q_co_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid co_id")

        branch_id = None
        if q_branch_id not in (None, ""):
            try:
                branch_id = int(q_branch_id)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Invalid branch_id")

        date_from = _parse_iso_date(q_date_from, "date_from")
        date_to = _parse_iso_date(q_date_to, "date_to")

        try:
            page = int(q_page) if q_page not in (None, "") else 1
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid page")
        try:
            limit = int(q_limit) if q_limit not in (None, "") else 50
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid limit")

        page = max(page, 1)
        limit = max(min(limit, 10000), 1)
        offset = (page - 1) * limit

        search_like = None
        if q_search and q_search.strip():
            search_like = f"%{q_search.strip()}%"

        base_params = {
            "co_id": co_id,
            "branch_id": branch_id,
            "date_from": date_from,
            "date_to": date_to,
            "search_like": search_like,
        }

        list_params = {
            **base_params,
            "limit": limit,
            "offset": offset,
        }

        list_query = get_mr_list_query()
        rows = db.execute(list_query, list_params).fetchall()

        data = []
        for row in rows:
            mapped = dict(row._mapping)
            # Convert date-like fields to ISO string for JSON serialization
            for date_field in ("jute_mr_date", "challan_date"):
                value = mapped.get(date_field)
                if value is not None and hasattr(value, "isoformat"):
                    mapped[date_field] = value.isoformat()
            data.append(mapped)

        count_query = get_mr_list_count_query()
        count_result = db.execute(count_query, base_params).scalar()
        total = int(count_result) if count_result is not None else 0

        return {"data": data, "total": total}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute MR list report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute MR list report: {str(e)}",
        )


@router.get("/qty-wise")
async def get_jute_qty_wise_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Jute Quantity Wise report (#2) — opening/receipt/issue/closing per jute
    quality over a date range, in weight (kg) and quantity. Also backs the
    Inventory Snapshot (#14): the page passes date_from = FY start, date_to = as-of.

    Query params:
    - branch_id: Branch ID (required)
    - date_from: Start date YYYY-MM-DD (required)
    - date_to: End date YYYY-MM-DD (required)
    """
    try:
        q_branch_id = request.query_params.get("branch_id")
        q_date_from = request.query_params.get("date_from")
        q_date_to = request.query_params.get("date_to")

        if not q_branch_id:
            raise HTTPException(status_code=400, detail="branch_id is required")
        if not q_date_from:
            raise HTTPException(status_code=400, detail="date_from is required")
        if not q_date_to:
            raise HTTPException(status_code=400, detail="date_to is required")

        try:
            branch_id = int(q_branch_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid branch_id")

        date_from = _parse_iso_date(q_date_from, "date_from")
        date_to = _parse_iso_date(q_date_to, "date_to")

        query = get_jute_qty_wise_report_query()
        rows = db.execute(query, {
            "branch_id": branch_id,
            "date_from": date_from,
            "date_to": date_to,
        }).fetchall()

        data = [dict(r._mapping) for r in rows]
        return {"data": data}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute quantity wise report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute quantity wise report: {str(e)}",
        )


@router.get("/txn-summary")
async def get_jute_txn_summary_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Jute Issue & Receipt Summary (#3) — transaction-level rows over a date range.

    Query params:
    - branch_id (required, int)
    - date_from, date_to (required, YYYY-MM-DD)
    - type: 'R' | 'I' | 'ALL' (default 'ALL')
    - search: optional MR-number filter
    """
    try:
        q_branch_id = request.query_params.get("branch_id")
        q_date_from = request.query_params.get("date_from")
        q_date_to = request.query_params.get("date_to")
        q_type = (request.query_params.get("type") or "ALL").upper()
        q_search = request.query_params.get("search")

        if not q_branch_id:
            raise HTTPException(status_code=400, detail="branch_id is required")
        if not q_date_from:
            raise HTTPException(status_code=400, detail="date_from is required")
        if not q_date_to:
            raise HTTPException(status_code=400, detail="date_to is required")
        if q_type not in ("R", "I", "ALL"):
            raise HTTPException(status_code=400, detail="type must be R, I or ALL")

        try:
            branch_id = int(q_branch_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid branch_id")

        date_from = _parse_iso_date(q_date_from, "date_from")
        date_to = _parse_iso_date(q_date_to, "date_to")
        search_like = f"%{q_search.strip()}%" if q_search and q_search.strip() else None

        query = get_jute_txn_summary_query(q_type)
        rows = db.execute(query, {
            "branch_id": branch_id,
            "date_from": date_from,
            "date_to": date_to,
            "search_like": search_like,
        }).fetchall()

        data = []
        for row in rows:
            mapped = dict(row._mapping)
            value = mapped.get("txn_date")
            if value is not None and hasattr(value, "isoformat"):
                mapped["txn_date"] = value.isoformat()
            data.append(mapped)
        return {"data": data}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute txn summary report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute txn summary report: {str(e)}",
        )


@router.get("/period-wise")
async def get_jute_period_wise_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Jute Month Wise (#10) / Day Wise (#11) — receipt vs issue weight & qty
    aggregated per period.

    Query params:
    - branch_id (required, int)
    - date_from, date_to (required, YYYY-MM-DD)
    - period: 'month' (default) | 'day'
    """
    try:
        q_branch_id = request.query_params.get("branch_id")
        q_date_from = request.query_params.get("date_from")
        q_date_to = request.query_params.get("date_to")
        q_period = (request.query_params.get("period") or "month").lower()

        if not q_branch_id:
            raise HTTPException(status_code=400, detail="branch_id is required")
        if not q_date_from:
            raise HTTPException(status_code=400, detail="date_from is required")
        if not q_date_to:
            raise HTTPException(status_code=400, detail="date_to is required")
        if q_period not in ("month", "day"):
            raise HTTPException(status_code=400, detail="period must be month or day")

        try:
            branch_id = int(q_branch_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid branch_id")

        date_from = _parse_iso_date(q_date_from, "date_from")
        date_to = _parse_iso_date(q_date_to, "date_to")

        query = get_jute_period_wise_query(q_period)
        rows = db.execute(query, {
            "branch_id": branch_id,
            "date_from": date_from,
            "date_to": date_to,
        }).fetchall()

        data = [dict(r._mapping) for r in rows]
        return {"data": data}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute period wise report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute period wise report: {str(e)}",
        )


def _require_branch_id(request: Request) -> int:
    """Validate and return the required branch_id query param."""
    q = request.query_params.get("branch_id")
    if not q:
        raise HTTPException(status_code=400, detail="branch_id is required")
    try:
        return int(q)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid branch_id")


def _serialize_dates(rows):
    """Map SQLAlchemy rows to dicts, converting any date-like values to ISO."""
    out = []
    for row in rows:
        m = dict(row._mapping)
        for k, v in list(m.items()):
            if v is not None and hasattr(v, "isoformat"):
                m[k] = v.isoformat()
        out.append(m)
    return out


def _branch_daterange_params(request: Request):
    """Validate branch_id + date_from + date_to and return them as a params dict."""
    branch_id = _require_branch_id(request)
    q_from = request.query_params.get("date_from")
    q_to = request.query_params.get("date_to")
    if not q_from:
        raise HTTPException(status_code=400, detail="date_from is required")
    if not q_to:
        raise HTTPException(status_code=400, detail="date_to is required")
    return {
        "branch_id": branch_id,
        "date_from": _parse_iso_date(q_from, "date_from"),
        "date_to": _parse_iso_date(q_to, "date_to"),
    }


@router.get("/percent-claims")
async def get_jute_percent_claims_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Jute Percent Claims (#7) — per-supplier pass/claim. Params: branch_id, date_from, date_to."""
    try:
        params = _branch_daterange_params(request)
        rows = db.execute(get_jute_percent_claims_query(), params).fetchall()
        return {"data": [dict(r._mapping) for r in rows]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute percent claims report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute percent claims report: {str(e)}",
        )


@router.get("/mukham-moisture")
async def get_jute_mukham_moisture_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Mukham Moisture (#9) — supplied vs allowed moisture per supplier/mukam. Params: branch_id, date_from, date_to."""
    try:
        params = _branch_daterange_params(request)
        rows = db.execute(get_jute_mukham_moisture_query(), params).fetchall()
        return {"data": [dict(r._mapping) for r in rows]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute mukham moisture report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute mukham moisture report: {str(e)}",
        )


@router.get("/with-value")
async def get_jute_with_value_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Jute with Value (#1) — opening/receipt/issue/closing per quality with
    receipt & issue value and average issue rate, over a date range.

    Query params: branch_id (req), date_from (req), date_to (req).
    """
    try:
        branch_id = _require_branch_id(request)
        q_date_from = request.query_params.get("date_from")
        q_date_to = request.query_params.get("date_to")
        if not q_date_from:
            raise HTTPException(status_code=400, detail="date_from is required")
        if not q_date_to:
            raise HTTPException(status_code=400, detail="date_to is required")
        date_from = _parse_iso_date(q_date_from, "date_from")
        date_to = _parse_iso_date(q_date_to, "date_to")

        rows = db.execute(get_jute_with_value_query(), {
            "branch_id": branch_id,
            "date_from": date_from,
            "date_to": date_to,
        }).fetchall()
        return {"data": [dict(r._mapping) for r in rows]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute with value report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute with value report: {str(e)}",
        )


@router.get("/mr-in-stock")
async def get_jute_mr_in_stock_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Jute MR in Stock (#4) — per MR-line outstanding stock. Param: branch_id."""
    try:
        branch_id = _require_branch_id(request)
        rows = db.execute(
            get_jute_mr_in_stock_query(), {"branch_id": branch_id}
        ).fetchall()
        return {"data": _serialize_dates(rows)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute MR in stock report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute MR in stock report: {str(e)}",
        )


@router.get("/mr-wise")
async def get_jute_mr_wise_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """MR Wise (#6) — per MR-line received/issued/balance. Param: branch_id."""
    try:
        branch_id = _require_branch_id(request)
        rows = db.execute(
            get_jute_mr_wise_query(), {"branch_id": branch_id}
        ).fetchall()
        return {"data": _serialize_dates(rows)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute MR wise report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute MR wise report: {str(e)}",
        )


@router.get("/godown-wise")
async def get_jute_godown_wise_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Godown Wise Stock (#5) — balance grouped by warehouse + quality. Param: branch_id."""
    try:
        branch_id = _require_branch_id(request)
        rows = db.execute(
            get_jute_godown_wise_query(), {"branch_id": branch_id}
        ).fetchall()
        return {"data": [dict(r._mapping) for r in rows]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute godown wise report: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute godown wise report: {str(e)}",
        )


def _compute_fy_start(today: date) -> date:
    """
    Compute the financial-year start date (April 1) for a given date.

    If today.month >= 4, FY started this year on April 1.
    Otherwise (Jan-Mar), FY started the prior year on April 1.
    """
    if today.month >= 4:
        return date(today.year, 4, 1)
    return date(today.year - 1, 4, 1)


TALLY_MAX_ROWS = 10_000

# Fixed 73-column xlsx layout for the Purchase sheet.
# Each entry: (header_text, sql_alias_or_None).
#
# Duplicate headers (Excel allows them; Python dicts don't) must be addressed
# by *column index*, so we keep the SQL-alias mapping here too for the rare
# columns where we do source data from SQL.
TALLY_PURCHASE_HEADERS = [
    "Vch No.",                        # 1
    "Vch Type",                       # 2
    "Date",                           # 3
    "Supplier Inv No",                # 4
    "Supplier Inv Date",              # 5
    "Receipt Note No",                # 6
    "Receipt Note Date",              # 7
    "Order No",                       # 8
    "Order Date",                     # 9
    "Party Name",                     # 10
    "Registration Type",              # 11
    "GSTIN No",                       # 12
    "Country",                        # 13
    "State",                          # 14
    "Pincode",                        # 15
    "Address 1",                      # 16
    "Address 2",                      # 17
    "Address 3",                      # 18
    "Purchase Ledger",                # 19
    "Purchase Ledger Description",    # 20
    "Item Name",                      # 21
    "Item Description",               # 22
    "UNITS.",                         # 23
    "Tracking No",                    # 24
    "Order No",                       # 25 (dup of 8)
    "Order Due Date",                 # 26
    "Godown",                         # 27
    "Batch",                          # 28
    "Qty",                            # 29
    "Rate",                           # 30
    "Amt",                            # 31
    "Amount1",                        # 32
    "Discount Ledger",                # 33
    "Discount Ledger Description",    # 34
    "Amount1",                        # 35 (dup of 32)
    "Additional Ledger",              # 36
    "Additional Ledger Description",  # 37
    "Amount",                         # 38
    "INPUT IGST",                     # 39
    "INPUT CGST",                     # 40
    "INPUT SGST",                     # 41
    "Roundoff amt",                   # 42
    "Total",                          # 43
    "Ref: No.",                       # 44
    "Due on",                         # 45
    "e Way Bill No.",                 # 46
    "e Way Bill Date",                # 47
    "Sub Type",                       # 48
    "Doc Type",                       # 49
    "Status Of eway Bill",            # 50
    "Mode",                           # 51
    "Distance (In KM)",               # 52
    "Transporter Name",               # 53
    "Vehicle No.",                    # 54
    "Doc/Loading/RR/AirWay No.",      # 55
    "Date",                           # 56 (dup of 3)
    "Transporter ID",                 # 57
    "Consignor:",                     # 58
    "Address 1",                      # 59 (dup of 16)
    "Address 2",                      # 60 (dup of 17)
    "Pin Code",                       # 61
    "Place",                          # 62
    "State",                          # 63 (dup of 14)
    "GSTIN/UIN",                      # 64
    "Consignee",                      # 65
    "Address 1",                      # 66 (dup of 16)
    "Address 2",                      # 67 (dup of 17)
    "To Place (Destination )",        # 68
    "Pin",                            # 69
    "State",                          # 70 (dup of 14)
    "GSTIN/UIN",                      # 71 (dup of 64)
    "Narration",                      # 72
    "TALLYIMPORTSTATUS",              # 73
]

CHECK_LIST_HEADERS = [
    "company_id",
    "mrdate",
    "mr_print_no",
    "invoice_no",
    "invoice_date",
    "supp_name",
    "supptally",
    "jute_quality",
    "qualitytally",
    "godowsntally",
    "claimtally",
    "purtally",
]

# SQL alias per column index (1-based) for the Purchase sheet.
# None means the column has no SQL source (left blank).
TALLY_SQL_ALIAS_BY_COL = {
    1: "Vch No.",
    2: "Vch Type",
    3: "Date",
    4: "Supplier Inv No",
    5: "Supplier Inv Date",
    6: "Receipt Note No",
    7: "Receipt Note Date",
    8: "Order No",
    9: "Order Date",
    10: "Party Name",
    19: "Purchase Ledger",
    21: "Item Name",
    22: "Item Description",
    27: "Godown",
    28: "Batch",
    29: "Qty",
    30: "Rate",
    31: "Amt",
    44: "Ref: No.",
    45: "Due on",
    72: "Narration",
}

# Columns that are filled on every row within the block (header + all item
# rows + TDS row + padding). These are the "shared" columns.
SHARED_COLS = (1, 2, 3, 4, 5, 6, 7, 10, 27, 28, 44, 45)

# Columns that are filled *only* on the first item row in a block.
FIRST_ITEM_ONLY_COLS = (8, 9, 19, 72)

# Columns that are filled on each item row (including the first).
PER_ITEM_COLS = (21, 22, 29, 30, 31)


def _build_tally_workbook(tally_rows, check_rows, co_id):
    """
    Build an xlsx workbook with Purchase + Check List sheets from raw SQL rows.

    Args:
        tally_rows: list of SQLAlchemy row mappings from
            get_jute_tally_download_query (one per MR line-item).
        check_rows: list of SQLAlchemy row mappings from
            get_jute_tally_check_list_query.
        co_id: numeric company id (int) for the Check List `company_id` col.

    Returns:
        openpyxl.Workbook — caller saves via wb.save(bio).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase"

    # Header row
    for c, h in enumerate(TALLY_PURCHASE_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)

    # Group SQL rows by Vch No. (preserves SQL order within each group).
    # dict preserves insertion order (Python 3.7+), so group order matches
    # SQL order of first occurrence.
    groups = {}
    group_order = []
    for r in tally_rows:
        m = dict(r._mapping)
        vch = m.get("Vch No.")
        if vch not in groups:
            groups[vch] = []
            group_order.append(vch)
        groups[vch].append(m)

    next_row = 2
    for vch in group_order:
        items = groups[vch]
        n = len(items)
        block_size = max(6, n + 2)
        first = items[0]

        # Shared values (common to every row in this block).
        shared = {col: first.get(TALLY_SQL_ALIAS_BY_COL[col]) for col in SHARED_COLS}

        # Item rows (0..n-1).
        for i, item in enumerate(items):
            row_idx = next_row + i
            # Shared columns
            for col, val in shared.items():
                ws.cell(row=row_idx, column=col, value=val)
            # Per-item columns (all item rows)
            for col in PER_ITEM_COLS:
                alias = TALLY_SQL_ALIAS_BY_COL[col]
                ws.cell(row=row_idx, column=col, value=item.get(alias))
            # First-item-only columns
            if i == 0:
                # Col 19 must be literal 'Purchase of Raw Jute' per spec;
                # the SQL alias already produces this value but we pin it
                # explicitly so future SQL drift doesn't break the template.
                ws.cell(row=row_idx, column=19, value="Purchase of Raw Jute")
                ws.cell(row=row_idx, column=8, value=first.get("Order No"))
                ws.cell(row=row_idx, column=9, value=first.get("Order Date"))
                ws.cell(row=row_idx, column=72, value=first.get("Narration"))

        # TDS row (row n within the block, if TDS Deducted > 0).
        tds_deducted = first.get("TDS Deducted") or 0
        try:
            tds_deducted_val = float(tds_deducted)
        except (TypeError, ValueError):
            tds_deducted_val = 0.0

        if tds_deducted_val > 0:
            tds_row_idx = next_row + n
            for col, val in shared.items():
                ws.cell(row=tds_row_idx, column=col, value=val)
            ws.cell(
                row=tds_row_idx,
                column=36,
                value="TDS ON PURCHASE OF GOODS (194Q)",
            )
            # Negative TDS amount goes in col 38.
            ws.cell(row=tds_row_idx, column=38, value=-int(round(tds_deducted_val)))

        # Padding rows — everything from (n+1) up to block_size-1 (0-based
        # block index). If TDS is 0, the n-th row becomes a padding row.
        pad_start = n if tds_deducted_val <= 0 else n + 1
        for p in range(pad_start, block_size):
            pad_row_idx = next_row + p
            for col, val in shared.items():
                ws.cell(row=pad_row_idx, column=col, value=val)

        next_row += block_size

    # --- Check List sheet ---
    ws2 = wb.create_sheet("Check List")
    for c, h in enumerate(CHECK_LIST_HEADERS, start=1):
        ws2.cell(row=1, column=c, value=h)

    for i, r in enumerate(check_rows, start=2):
        m = dict(r._mapping)
        # company_id — force the user-provided co_id (int) for consistency.
        ws2.cell(row=i, column=1, value=co_id)
        ws2.cell(row=i, column=2, value=m.get("mrdate"))
        ws2.cell(row=i, column=3, value=m.get("mr_print_no"))
        ws2.cell(row=i, column=4, value=m.get("invoice_no"))
        ws2.cell(row=i, column=5, value=m.get("invoice_date"))
        ws2.cell(row=i, column=6, value=m.get("supp_name"))
        ws2.cell(row=i, column=7, value=m.get("supptally"))
        ws2.cell(row=i, column=8, value=m.get("jute_quality"))
        ws2.cell(row=i, column=9, value=m.get("qualitytally"))
        ws2.cell(row=i, column=10, value=m.get("godowsntally"))
        ws2.cell(row=i, column=11, value=m.get("claimtally"))
        ws2.cell(row=i, column=12, value=m.get("purtally"))

    return wb


@router.get("/tally-download")
def get_jute_tally_download(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """
    Jute Purchase Tally download — returns a multi-sheet xlsx blob
    (Purchase + Check List).

    The Purchase sheet uses a fixed 6-row block per MR:
      - N item rows (one per MR line)
      - 1 TDS row if 194Q TDS > 0 (else a padding row)
      - padding rows to fill the block to max(6, N+2) rows total

    Computes per-supplier cumulative naramt from the start of the current
    financial year and applies 194Q TDS at 0.1% once cumulative crosses
    TDS_CAP_INR (INR 50 lakh).

    Query params:
      - co_id (required, int): Company ID
      - branch_id (optional, int): Branch filter
      - date_from (required, YYYY-MM-DD): Inclusive start of MR date range
      - date_to (required, YYYY-MM-DD): Inclusive end of MR date range

    Returns:
        StreamingResponse carrying the xlsx bytes
        (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet).

    Raises:
        400 if parameters are missing/invalid, or if the emitted xlsx would
            exceed TALLY_MAX_ROWS (10,000) total Purchase-sheet rows.
    """
    try:
        q_co_id = request.query_params.get("co_id")
        q_branch_id = request.query_params.get("branch_id")
        q_date_from = request.query_params.get("date_from")
        q_date_to = request.query_params.get("date_to")

        if not q_co_id:
            raise HTTPException(status_code=400, detail="co_id is required")
        if not q_date_from:
            raise HTTPException(status_code=400, detail="date_from is required")
        if not q_date_to:
            raise HTTPException(status_code=400, detail="date_to is required")

        try:
            co_id = int(q_co_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid co_id")

        branch_id = None
        if q_branch_id not in (None, ""):
            try:
                branch_id = int(q_branch_id)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Invalid branch_id")

        date_from = _parse_iso_date(q_date_from, "date_from")
        date_to = _parse_iso_date(q_date_to, "date_to")

        # Compute FY start dynamically (April 1 of the current FY).
        fy_start = _compute_fy_start(date.today()).isoformat()

        base_params = {
            "co_id": co_id,
            "branch_id": branch_id,
            "date_from": date_from,
            "date_to": date_to,
        }

        # Enforce the xlsx-row cap by counting items-per-MR first and computing
        # the total emitted rows (sum of max(6, N+2)). Cheaper than running the
        # full tally query for oversized ranges.
        items_per_mr_query = get_jute_tally_download_items_per_mr_query()
        mr_item_rows = db.execute(items_per_mr_query, base_params).fetchall()
        total_xlsx_rows = 0
        for mr_row in mr_item_rows:
            mapping = dict(mr_row._mapping)
            n = int(mapping.get("item_count") or 0)
            total_xlsx_rows += max(6, n + 2)

        if total_xlsx_rows > TALLY_MAX_ROWS:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Date range too wide — max {TALLY_MAX_ROWS} rows, "
                    f"got {total_xlsx_rows}"
                ),
            )

        # Main Tally data + Check List queries.
        list_query = get_jute_tally_download_query()
        tally_rows = db.execute(
            list_query,
            {**base_params, "fy_start": fy_start},
        ).fetchall()

        check_list_query = get_jute_tally_check_list_query()
        check_rows = db.execute(check_list_query, base_params).fetchall()

        wb = _build_tally_workbook(tally_rows, check_rows, co_id)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        xlsx_bytes = bio.getvalue()

        filename = f"jute_purchase_tally_{date_from}_{date_to}.xlsx"
        return StreamingResponse(
            iter([xlsx_bytes]),
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching jute tally download: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching jute tally download: {str(e)}",
        )
