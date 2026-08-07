"""
Export helpers for procurement reports.

Builds xlsx (openpyxl) and pdf (reportlab) BytesIO streams from list-of-dict
report rows. Used by all procurement report download endpoints.
"""

from datetime import datetime
from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    LongTable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def _stringify(value) -> str:
    """Render a cell value as a string. None becomes empty string."""
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return str(value)


def build_xlsx(
    rows: list[dict],
    columns: Sequence[tuple[str, str]],
    sheet_name: str = "Report",
) -> BytesIO:
    """Build an xlsx workbook in-memory.

    Args:
        rows: list of dicts (one per row).
        columns: list of (header_label, dict_key) tuples - defines column order.
        sheet_name: worksheet title.

    Returns:
        BytesIO positioned at 0, ready to stream.
    """
    wb = Workbook()
    ws = wb.active
    # Sheet name length capped at 31 chars by Excel.
    ws.title = (sheet_name or "Report")[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center")

    # Write header row
    for col_idx, (label, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_label, key) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_stringify(row.get(key)))

    # Freeze first row
    ws.freeze_panes = "A2"

    # Auto-width approximation: max(len(str(value))) per column, capped.
    for col_idx, (label, key) in enumerate(columns, start=1):
        width = len(str(label))
        for row in rows:
            val = _stringify(row.get(key))
            if len(val) > width:
                width = len(val)
        # Add a little padding; cap at 60 to avoid runaway widths.
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(
    rows: list[dict],
    columns: Sequence[tuple[str, str]],
    title: str,
) -> BytesIO:
    """Build a landscape A4 PDF in-memory.

    Args:
        rows: list of dicts (one per row).
        columns: list of (header_label, dict_key) tuples - defines column order.
        title: report title rendered at the top of the document.

    Returns:
        BytesIO positioned at 0, ready to stream.
    """
    buf = BytesIO()
    page_size = landscape(A4)

    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=title or "Report",
    )

    styles = getSampleStyleSheet()
    title_para = Paragraph(f"<b>{title or 'Report'}</b>", styles["Title"])
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta_para = Paragraph(f"Generated: {ts}", styles["Normal"])

    # Build the table data: header row + body rows.
    header_row = [label for label, _key in columns]
    body_rows = [
        [_stringify(row.get(key)) for _label, key in columns]
        for row in rows
    ]
    table_data = [header_row] + body_rows

    # Use LongTable when many rows, so it paginates across pages cleanly.
    TableCls = LongTable if len(body_rows) > 30 else Table
    table = TableCls(table_data, repeatRows=1)

    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]
    )
    table.setStyle(style)

    story = [title_para, meta_para, Spacer(1, 6), table]
    doc.build(story)
    buf.seek(0)
    return buf
