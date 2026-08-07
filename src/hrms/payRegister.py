"""HRMS Pay Register endpoints — list, create, update, setup, salary data, process payroll, export, payslips."""
import logging
import math
import re
from collections import OrderedDict
from decimal import ROUND_HALF_EVEN, ROUND_HALF_UP, ROUND_CEILING, ROUND_FLOOR, Decimal
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from src.config.db import get_tenant_db
from src.authorization.utils import get_current_user_with_refresh
from src.models.hrms import PayPeriod, PayEmployeePayroll, PayEmployeePayperiod
from .query import (
    get_pay_register_list,
    get_pay_register_list_count,
    get_pay_register_by_id,
    check_duplicate_pay_register,
    get_pay_register_salary,
    get_payscheme_mapped_employees,
    get_pay_scheme_details_by_id,
    get_pay_scheme_details_by_scheme_id,
    get_payslip_print_components,
    get_employee_pf_map,
    get_employee_esi_map,
    get_company_name,
    get_pay_schemes_for_company,
    get_all_pay_components_for_company,
    get_employee_pay_structure,
    get_custom_component_values,
    delete_existing_payroll_for_period,
    delete_existing_payperiod_entries,
)
from src.common.utils import parse_json_body

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

logger = logging.getLogger(__name__)

router = APIRouter()


# ─── Pay Register List ──────────────────────────────────────────────

@router.get("/pay_register_list")
def pay_register_list(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 20))
        search_raw = request.query_params.get("search")
        search = f"%{search_raw}%" if search_raw else None
        from_date = request.query_params.get("from_date") or None
        to_date = request.query_params.get("to_date") or None
        status_raw = request.query_params.get("status")
        status_id = int(status_raw) if status_raw else None

        params = {
            "co_id": int(co_id),
            "search": search,
            "from_date": from_date,
            "to_date": to_date,
            "status_id": status_id,
            "page_size": page_size,
            "offset": (page - 1) * page_size,
        }

        rows = db.execute(get_pay_register_list(), params).fetchall()
        data = [dict(r._mapping) for r in rows]

        count_row = db.execute(get_pay_register_list_count(), {
            "co_id": int(co_id),
            "search": search,
            "from_date": from_date,
            "to_date": to_date,
            "status_id": status_id,
        }).fetchone()
        total = count_row._mapping["total"] if count_row else 0

        return {"data": data, "total": total, "page": page, "page_size": page_size}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Pay Register By ID ────────────────────────────────────────────

@router.get("/pay_register_by_id/{period_id}")
def pay_register_by_id(
    period_id: int,
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        row = db.execute(
            get_pay_register_by_id(),
            {"id": period_id, "co_id": int(co_id)},
        ).fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Pay register not found")

        detail = dict(row._mapping)

        # Approval button check: if status is pending approval (20) or open (1),
        # the current user might be allowed to approve.
        user_id = token_data.get("user_id", 0)
        approve_button = False
        status = detail.get("status_id")
        if status in (1, 20):
            approve_button = True

        detail["approveButton"] = approve_button

        return {"data": detail}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Pay Register Create Setup ─────────────────────────────────────

@router.get("/pay_register_create_setup")
def pay_register_create_setup(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        from sqlalchemy import text

        # Pay schemes — sourced from pay_scheme_master, scoped by company.
        # (pay_scheme_master.payscheme_id is the id referenced by pay_scheme_details
        #  / pay_period, so these match what the register calc expects.)
        schemes = db.execute(
            get_pay_schemes_for_company(), {"co_id": int(co_id)},
        ).fetchall()
        print(f"Fetched {(schemes)} pay schemes for company {co_id}")

        # Branches
        branches = db.execute(
            text("""
                SELECT branch_id, branch_name
                FROM branch_mst
                WHERE co_id = :co_id AND active = 1
                ORDER BY branch_name
            """),
            {"co_id": int(co_id)},
        ).fetchall()

        return {
            "data": {
                "pay_schemes": [
                    {"label": r._mapping["name"], "value": str(r._mapping["id"])}
                    for r in schemes
                ],
                "branches": [
                    {"label": r._mapping["branch_name"], "value": str(r._mapping["branch_id"])}
                    for r in branches
                ],
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Pay Register Create ───────────────────────────────────────────

@router.post("/pay_register_create")
def pay_register_create(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        body = parse_json_body(request)
        user_id = token_data.get("user_id", 0)

        from_date = body.get("from_date")
        to_date = body.get("to_date")
        payscheme_id = body.get("payscheme_id")
        branch_id = body.get("branch_id")

        if not from_date or not to_date or not payscheme_id:
            raise HTTPException(
                status_code=400,
                detail="from_date, to_date, and payscheme_id are required",
            )

        # Duplicate check (skip cancelled/rejected/closed statuses)
        dup = db.execute(
            check_duplicate_pay_register(),
            {
                "from_date": from_date,
                "to_date": to_date,
                "payscheme_id": int(payscheme_id),
                "co_id": int(co_id),
            },
        ).fetchone()
        if dup and dup._mapping["cnt"] > 0:
            raise HTTPException(status_code=409, detail="Pay register already exists for this period and scheme")

        # Supersede a prior PROCESSED (status 28) register for the same
        # period + branch + pay scheme by marking it REJECTED (status 4) before
        # creating the replacement, so only the new run stays active. Status 28
        # is excluded from the duplicate check above, so it never raises the 409
        # — we reject the stale processed row here instead. This runs in the same
        # transaction as the insert below, so it commits atomically with it.
        branch_val = int(branch_id) if branch_id else None
        db.query(PayPeriod).filter(
            PayPeriod.from_date == from_date,
            PayPeriod.to_date == to_date,
            PayPeriod.payscheme_id == int(payscheme_id),
            PayPeriod.branch_id == branch_val,
            PayPeriod.co_id == int(co_id),
            PayPeriod.status_id == 28,  # PROCESSED
        ).update(
            {PayPeriod.status_id: 4, PayPeriod.updated_by: user_id},  # REJECTED
            synchronize_session=False,
        )

        period = PayPeriod(
            from_date=from_date,
            to_date=to_date,
            payscheme_id=int(payscheme_id),
            branch_id=int(branch_id) if branch_id else None,
            co_id=int(co_id),
            status_id=1,  # Open
            updated_by=user_id,
        )
        db.add(period)
        db.flush()
        db.commit()

        # Generate payroll lines as part of creation (mirrors sls-payroll-service:
        # create + process is one operation). The period is already committed, so it
        # survives even if processing fails — we surface the error instead of 500ing.
        period_id = period.id
        processing: dict | None = None
        processing_error: str | None = None
        try:
            processing = _run_payroll_processing(
                db, int(co_id), period, int(payscheme_id),
                int(branch_id) if branch_id else None, user_id,
            )
            db.commit()
        except HTTPException as pe:
            db.rollback()
            processing_error = pe.detail if isinstance(pe.detail, str) else "Payroll processing failed"
        except Exception as e:
            db.rollback()
            logger.exception("Payroll processing during create failed")
            processing_error = str(e)

        resp: dict = {"id": period_id, "message": "Pay register created successfully"}
        if processing is not None:
            resp["processing"] = processing
        if processing_error is not None:
            resp["processing_error"] = processing_error
        return {"data": resp}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ─── Pay Register Update (status / approval) ───────────────────────

@router.put("/pay_register_update/{period_id}")
def pay_register_update(
    period_id: int,
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        body = parse_json_body(request)
        user_id = token_data.get("user_id", 0)

        period = db.query(PayPeriod).filter(
            PayPeriod.id == period_id,
            PayPeriod.co_id == int(co_id),
        ).first()
        if not period:
            raise HTTPException(status_code=404, detail="Pay register not found")

        # Allow updating status and basic fields
        for field in ("from_date", "to_date", "payscheme_id", "branch_id", "status_id"):
            if field in body:
                setattr(period, field, body[field])
        period.updated_by = user_id

        db.commit()
        return {"data": {"id": period_id, "message": "Pay register updated successfully"}}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ─── Monthly Salary Payment Data ───────────────────────────────────

@router.get("/pay_register_salary")
def pay_register_salary(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Returns per-employee salary rows with columns driven by
    tbl_payslip_print_component: Emp Code + Name anchors, then the configured
    columns in order (meta PF/UAN/ESI as text, pay components as numbers). Falls
    back to all pay components (with Department) when no display config exists.
    Also returns a `totals` grand-total row (each numeric column summed)."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        pay_period_id_raw = request.query_params.get("pay_period_id")
        if not pay_period_id_raw:
            raise HTTPException(status_code=400, detail="pay_period_id is required")
        pay_period_id = int(pay_period_id_raw)

        # Scheme / branch override the pay period's own values when the caller
        # passes them (the view page normally only knows the period id).
        pay_scheme_id_raw = request.query_params.get("pay_scheme_id")
        branch_id_raw = request.query_params.get("branch_id")

        ctx = _load_register_context(
            db, int(co_id), pay_period_id,
            pay_scheme_id_override=int(pay_scheme_id_raw) if pay_scheme_id_raw else None,
            branch_id_override=int(branch_id_raw) if branch_id_raw else None,
        )

        render_cols = _build_register_columns(ctx["components"], ctx["used_config"])

        # Frontend DataGrid column defs (header always renders, even with no rows).
        columns = []
        for col in render_cols:
            col_def = {"field": col["key"], "headerName": col["label"], "minWidth": 120}
            if col["kind"] == "num":
                col_def["type"] = "number"
            columns.append(col_def)

        data = []
        for emp in ctx["employees"]:
            row = {"id": emp["eb_id"], "employee_id": emp["eb_id"]}
            for col in render_cols:
                row[col["key"]] = _register_cell_value(col, emp)
            data.append(row)

        # Vertical grand-total row (summed numeric columns), only when there is
        # data to total. The frontend pins it as the last row.
        totals = None
        if ctx["employees"]:
            totals = _grand_total_row(render_cols, ctx["employees"])
            totals["id"] = "__totals__"

        return {"data": data, "columns": columns, "totals": totals}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Pay Register Export (.xlsx) & Pay Slip Print (PDF) ────────────
#
# Both outputs share the same component-selection logic: prefer the columns
# configured in tbl_payslip_print_component for this payscheme/company/branch;
# fall back to every component of the pay scheme when none are configured. They
# also share the same employee values (the pivoted pay_register_salary data) and
# header metadata (company / scheme / period).


# ─── Meta (employee-info) columns ──────────────────────────────────
#
# Negative sentinel component_ids configured in tbl_payslip_print_component.
# They render as TEXT (identity / statutory values), never as amounts, and are
# excluded from the grand total. Keep in sync with the frontend
# META_COMPONENT_OPTIONS (PayRegisterDispSet/types/payRegisterDispSetTypes.ts).
#   -1 Emp Code / -2 Name / -3 Department  → come from the salary pivot
#   -4 PF No / -5 UAN No                    → hrms_ed_pf
#   -6 ESI No                               → hrms_ed_esi
META_FIELDS: dict[int, str] = {
    -1: "emp_code",
    -2: "emp_name",
    -3: "department_name",
    -4: "pf_no",
    -5: "pf_uan_no",
    -6: "esi_no",
}
# meta_keys sourced straight off the employee salary row (no extra query)
_META_FROM_EMPLOYEE = {"emp_code", "emp_name", "department_name"}
_META_PF_KEYS = {"pf_no", "pf_uan_no"}
_META_ESI_KEYS = {"esi_no"}
# Emp Code (-1) and Name (-2) are always rendered as leading anchor columns, so
# they are skipped when walking the configured rows to avoid duplication.
_ANCHOR_META_IDS = {-1, -2}
_GRAND_TOTAL_LABEL = "GRAND TOTAL"


def _meta_value(emp: dict, meta_key: str | None) -> str:
    """Text value of a meta column for one employee.

    emp_code / emp_name / department_name live on the employee row; pf_no /
    pf_uan_no / esi_no live in emp['meta'] (populated by _load_register_context).
    """
    if not meta_key:
        return ""
    if meta_key in _META_FROM_EMPLOYEE:
        return str(emp.get(meta_key) or "")
    return str((emp.get("meta") or {}).get(meta_key) or "")


def _resolve_print_components(
    db: Session,
    co_id: int,
    payscheme_id: int,
    branch_id: int | None,
) -> tuple[list[dict], bool]:
    """Ordered component config for grid + export + payslip.

    Returns (components, used_config). Each entry:
      {component_id, label, group, total_print, payslip_print, is_meta, meta_key}.
    Uses tbl_payslip_print_component when rows exist for this
    payscheme/company/branch (used_config=True); otherwise falls back to all
    pay-scheme components (used_config=False, no meta).
    """
    rows = []
    if branch_id is not None:
        rows = db.execute(
            get_payslip_print_components(),
            {
                "payscheme_id": payscheme_id,
                "company_id": co_id,
                "branch_id": branch_id,
            },
        ).fetchall()

    if rows:
        components: list[dict] = []
        for r in rows:
            m = r._mapping
            cid = m["component_id"]
            if cid is None:
                continue
            cid = int(cid)
            is_meta = cid < 0
            meta_key = META_FIELDS.get(cid)
            # Unknown negative sentinel — not a real component, skip it.
            if is_meta and meta_key is None:
                continue
            components.append({
                "component_id": cid,
                "label": (m["desc_print"] or "").strip() or f"Component {cid}",
                "group": (m["fixed_var_cols"] or "").strip(),
                "total_print": bool(m["total_print"]),
                "payslip_print": bool(m["payslip_print"]),
                "is_meta": is_meta,
                "meta_key": meta_key,
            })
        if components:
            return components, True

    # Fallback — every component of the pay scheme, all shown, no special total.
    fallback = db.execute(
        get_pay_scheme_details_by_scheme_id(),
        {"payscheme_id": payscheme_id},
    ).fetchall()
    components = []
    seen: set[int] = set()
    for r in fallback:
        m = r._mapping
        cid = m["component_id"]
        if cid is None or int(cid) in seen:
            continue
        seen.add(int(cid))
        components.append({
            "component_id": int(cid),
            "label": (m["component_name"] or m["component_code"] or f"Component {cid}"),
            "group": "",
            "total_print": False,
            "payslip_print": True,
            "is_meta": False,
            "meta_key": None,
        })
    return components, False


def _build_register_columns(components: list[dict], used_config: bool) -> list[dict]:
    """Ordered render columns for the grid + Excel export.

    Each column: {key, label, kind: 'text'|'num', ...}. Emp Code + Name are
    always the leading anchor columns. In configured mode the remaining columns
    follow payslip_order (meta render as text, components as numbers). In
    fallback mode: Emp Code, Name, Department, then every component. There is no
    per-row Total column — the only total is the vertical grand-total row.
    """
    cols: list[dict] = [
        {"key": "emp_code", "label": "Emp Code", "kind": "text"},
        {"key": "emp_name", "label": "Name", "kind": "text"},
    ]

    if not used_config:
        cols.append({"key": "department_name", "label": "Department", "kind": "text"})
        for c in components:
            cid = c["component_id"]
            cols.append({
                "key": f"c{cid}", "label": c["label"],
                "kind": "num", "component_id": cid,
            })
        return cols

    for c in components:
        cid = c["component_id"]
        if cid in _ANCHOR_META_IDS:
            continue  # Emp Code / Name already shown as anchors
        if c["is_meta"]:
            cols.append({
                "key": f"m{abs(cid)}", "label": c["label"],
                "kind": "text", "meta_key": c["meta_key"],
            })
        else:
            cols.append({
                "key": f"c{cid}", "label": c["label"],
                "kind": "num", "component_id": cid,
            })

    return cols


def _register_cell_value(col: dict, emp: dict):
    """Resolve one cell's value for an employee from a render column descriptor."""
    if "meta_key" in col:
        return _meta_value(emp, col["meta_key"])
    if "component_id" in col:
        return float(emp["values"].get(col["component_id"], 0) or 0)
    # leading anchor text column (emp_code / emp_name / department_name)
    return str(emp.get(col["key"]) or "")


def _grand_total_row(render_cols: list[dict], employees: list[dict]) -> dict:
    """Vertical grand-total row: each numeric column summed across all
    employees; text columns blank except a 'GRAND TOTAL' label under Name."""
    row: dict = {}
    for col in render_cols:
        if col["kind"] == "num":
            row[col["key"]] = round(
                sum(float(_register_cell_value(col, e) or 0) for e in employees), 2
            )
        else:
            row[col["key"]] = ""
    if "emp_name" in row:
        row["emp_name"] = _GRAND_TOTAL_LABEL
    elif render_cols:
        row[render_cols[0]["key"]] = _GRAND_TOTAL_LABEL
    return row


def _attach_meta_values(db: Session, components: list[dict], employees: list[dict]) -> None:
    """Populate emp['meta'] with PF / UAN / ESI text for the employees, but only
    when at least one such meta column is configured. Sourced by eb_id from
    hrms_ed_pf (pf_no, pf_uan_no) and hrms_ed_esi (esi_no)."""
    needs_pf = any(c["meta_key"] in _META_PF_KEYS for c in components)
    needs_esi = any(c["meta_key"] in _META_ESI_KEYS for c in components)
    if not (needs_pf or needs_esi) or not employees:
        return

    eb_ids_csv = ",".join(str(emp["eb_id"]) for emp in employees)
    by_eb: dict[int, dict] = {emp["eb_id"]: emp for emp in employees}

    if needs_pf:
        for r in db.execute(get_employee_pf_map(), {"eb_ids": eb_ids_csv}).fetchall():
            m = r._mapping
            emp = by_eb.get(m["eb_id"])
            if emp is not None:
                emp["meta"]["pf_no"] = m["pf_no"]
                emp["meta"]["pf_uan_no"] = m["pf_uan_no"]

    if needs_esi:
        for r in db.execute(get_employee_esi_map(), {"eb_ids": eb_ids_csv}).fetchall():
            m = r._mapping
            emp = by_eb.get(m["eb_id"])
            if emp is not None:
                emp["meta"]["esi_no"] = m["esi_no"]


def _load_register_context(
    db: Session,
    co_id: int,
    pay_period_id: int,
    pay_scheme_id_override: int | None = None,
    branch_id_override: int | None = None,
) -> dict:
    """Common data for the grid + both export endpoints: header metadata, the
    resolved component config, and the per-employee pivoted salary values
    (with meta PF/UAN/ESI text).

    Returns: {
        company_name, scheme_name, period_label, used_config,
        components: [ {component_id, label, group, total_print, payslip_print,
                       is_meta, meta_key} ],
        employees: [ {eb_id, emp_code, emp_name, department_name,
                      values: {cid: amount}, meta: {pf_no,...}} ],
    }
    """
    period = db.query(PayPeriod).filter(
        PayPeriod.id == pay_period_id,
        PayPeriod.co_id == co_id,
    ).first()
    if not period:
        raise HTTPException(status_code=404, detail="Pay register not found")

    payscheme_id = pay_scheme_id_override or period.payscheme_id
    branch_id = branch_id_override if branch_id_override is not None else period.branch_id

    # Header metadata — scheme name + formatted period from the register row.
    detail_row = db.execute(
        get_pay_register_by_id(), {"id": pay_period_id, "co_id": co_id},
    ).fetchone()
    detail = dict(detail_row._mapping) if detail_row else {}
    scheme_name = detail.get("payscheme") or f"Pay Scheme {payscheme_id}"
    from_desc = detail.get("fromDateDesc") or ""
    to_desc = detail.get("toDateDesc") or ""
    period_label = f"Pay Register for {from_desc} - {to_desc}".strip(" -")

    company_row = db.execute(get_company_name(), {"co_id": co_id}).fetchone()
    company_name = (company_row._mapping["co_name"] if company_row else "") or "Company"

    components, used_config = _resolve_print_components(db, co_id, payscheme_id, branch_id)

    # Per-employee pivoted salary values, keyed by component_id.
    salary_rows = db.execute(
        get_pay_register_salary(),
        {
            "co_id": co_id,
            "pay_period_id": pay_period_id,
            "pay_scheme_id": payscheme_id,
            "branch_id": branch_id,
        },
    ).fetchall()

    employees: "OrderedDict[int, dict]" = OrderedDict()
    for r in salary_rows:
        m = r._mapping
        emp_id = m["employee_id"]
        if emp_id not in employees:
            employees[emp_id] = {
                "eb_id": emp_id,
                "emp_code": m["emp_code"],
                "emp_name": (m["emp_name"] or "").strip(),
                "department_name": m["department_name"] or "",
                "values": {},
                "meta": {},
            }
        employees[emp_id]["values"][int(m["component_id"])] = float(m["amount"] or 0)

    employee_list = list(employees.values())

    # Pull PF / UAN / ESI text for the meta columns (only when configured).
    _attach_meta_values(db, components, employee_list)

    return {
        "company_name": company_name,
        "scheme_name": scheme_name,
        "period_label": period_label,
        "used_config": used_config,
        "components": components,
        "employees": employee_list,
    }


@router.get("/pay_register_export")
def pay_register_export(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Export the pay register as an .xlsx: company/scheme/period headers, one row
    per employee with the configured component columns, and a grand-total row."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        pay_period_id_raw = request.query_params.get("pay_period_id")
        if not pay_period_id_raw:
            raise HTTPException(status_code=400, detail="pay_period_id is required")

        ctx = _load_register_context(db, int(co_id), int(pay_period_id_raw))

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        employees = ctx["employees"]
        render_cols = _build_register_columns(ctx["components"], ctx["used_config"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Pay Register"

        bold = Font(bold=True)
        total_span = max(len(render_cols), 1)

        # Styling to match the on-screen grid header (#3ea6da, white bold) with
        # thin cell borders around the whole table.
        header_fill = PatternFill("solid", fgColor="3EA6DA")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="E0E0E0")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Track the widest text per column for auto-fit (seeded with the header).
        col_widths = [len(str(col["label"])) for col in render_cols]

        def _note_width(ci: int, value) -> None:
            col_widths[ci] = max(col_widths[ci], len(str(value if value is not None else "")))

        # Header block (company / scheme / period), each on its own merged row.
        for idx, line in enumerate(
            (ctx["company_name"], ctx["scheme_name"], ctx["period_label"]), start=1,
        ):
            cell = ws.cell(row=idx, column=1, value=line)
            cell.font = bold
            if total_span > 1:
                ws.merge_cells(
                    start_row=idx, start_column=1, end_row=idx, end_column=total_span,
                )
            cell.alignment = Alignment(horizontal="left")

        header_row_idx = 5  # leave a blank row (4) between header block and table
        for col_idx, col in enumerate(render_cols, start=1):
            c = ws.cell(row=header_row_idx, column=col_idx, value=col["label"])
            c.font = header_font
            c.fill = header_fill
            c.border = cell_border
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Per-employee rows (meta columns as text, components as numbers).
        row_idx = header_row_idx + 1
        for emp in employees:
            for ci, col in enumerate(render_cols):
                value = _register_cell_value(col, emp)
                c = ws.cell(row=row_idx, column=1 + ci, value=value)
                c.border = cell_border
                if col["kind"] == "num":
                    c.alignment = Alignment(horizontal="right")
                _note_width(ci, value)
            row_idx += 1

        # Vertical grand-total row — each numeric column summed across employees.
        if employees:
            grand = _grand_total_row(render_cols, employees)
            for ci, col in enumerate(render_cols):
                value = grand.get(col["key"], "")
                c = ws.cell(row=row_idx, column=1 + ci, value=value)
                c.font = bold
                c.border = cell_border
                if col["kind"] == "num":
                    c.alignment = Alignment(horizontal="right")
                _note_width(ci, value)

        # Auto-fit column widths (header + cell content, with a little padding).
        for ci, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = min(max(width + 2, 10), 60)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"PayRegister_{pay_period_id_raw}.xlsx"
        return StreamingResponse(
            iter([bio.getvalue()]),
            media_type=XLSX_MEDIA_TYPE,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_register_export failed")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/pay_register_payslips")
def pay_register_payslips(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Generate a PDF with one payslip per employee (page break between).

    Each slip: company/scheme/period + employee header, the components flagged
    payslip_print (fallback: all), grouped by fixed_var_cols, then a Total line
    summing the total_print components (fallback: all shown components)."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        pay_period_id_raw = request.query_params.get("pay_period_id")
        if not pay_period_id_raw:
            raise HTTPException(status_code=400, detail="pay_period_id is required")

        ctx = _load_register_context(db, int(co_id), int(pay_period_id_raw))

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
        )

        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        h_style = styles["Heading2"]
        normal = styles["Normal"]

        # Amount table = real pay components only (meta columns are identity /
        # statutory text, shown in the employee header instead).
        slip_components = [
            c for c in ctx["components"] if c["payslip_print"] and not c["is_meta"]
        ]
        if not slip_components:
            slip_components = [c for c in ctx["components"] if not c["is_meta"]]
        any_total_flag = any(c["total_print"] for c in slip_components)

        # Configured statutory meta fields (PF No / UAN No / ESI No) for the
        # header — Emp Code / Name / Department are already in the fixed header.
        header_meta = [
            c for c in ctx["components"]
            if c["is_meta"] and c["meta_key"] not in _META_FROM_EMPLOYEE
        ]

        bio = BytesIO()
        doc = SimpleDocTemplate(
            bio, pagesize=A4,
            topMargin=15 * mm, bottomMargin=15 * mm,
            leftMargin=15 * mm, rightMargin=15 * mm,
            title=f"Pay Slips {pay_period_id_raw}",
        )

        story: list = []
        employees = ctx["employees"]

        if not employees:
            story.append(Paragraph(ctx["company_name"], title_style))
            story.append(Paragraph(ctx["scheme_name"], h_style))
            story.append(Paragraph(ctx["period_label"], normal))
            story.append(Spacer(1, 8 * mm))
            story.append(Paragraph("No salary data available for this pay register.", normal))
        else:
            for emp_idx, emp in enumerate(employees):
                story.append(Paragraph(ctx["company_name"], title_style))
                story.append(Paragraph(ctx["scheme_name"], h_style))
                story.append(Paragraph(ctx["period_label"], normal))
                story.append(Spacer(1, 4 * mm))
                story.append(Paragraph(
                    f"<b>Emp Code:</b> {emp['emp_code']}　"
                    f"<b>Name:</b> {emp['emp_name']}　"
                    f"<b>Department:</b> {emp['department_name']}",
                    normal,
                ))
                if header_meta:
                    meta_line = "　".join(
                        f"<b>{c['label']}:</b> {_meta_value(emp, c['meta_key']) or '-'}"
                        for c in header_meta
                    )
                    story.append(Paragraph(meta_line, normal))
                story.append(Spacer(1, 4 * mm))

                # Component lines, grouped by fixed_var_cols. A blank group key
                # means "no grouping" — those lines render in document order.
                table_data = [["Component", "Amount"]]
                running_total = 0.0
                current_group = None
                for comp in slip_components:
                    grp = comp["group"]
                    if grp and grp != current_group:
                        current_group = grp
                        table_data.append([grp, ""])
                    amount = float(emp["values"].get(comp["component_id"], 0) or 0)
                    table_data.append([comp["label"], f"{amount:,.2f}"])
                    if comp["total_print"] or not any_total_flag:
                        running_total += amount

                table_data.append(["TOTAL", f"{running_total:,.2f}"])

                tbl = Table(table_data, colWidths=[110 * mm, 50 * mm])
                tbl.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3ea6da")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]))
                story.append(tbl)

                if emp_idx < len(employees) - 1:
                    story.append(PageBreak())

        doc.build(story)
        bio.seek(0)

        filename = f"PaySlips_{pay_period_id_raw}.pdf"
        return StreamingResponse(
            iter([bio.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_register_payslips failed")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Payroll Processing (Calculate & Insert) ───────────────────────

# Status constants matching Java backend
STATUS_PROCESSED = "28"   # PROCESSED in status_mst (28); 32 is LOCKED
STATUS_OPEN = "1"
EMP_PAY_ROLL_SUCCESS = 1


def _round_value(value: float, digits: int | None, round_type: int | None) -> float:
    """Apply rounding per the pay_components ROUNDOF_TYPE dropdown.

    round_type: 0/None = None (no rounding), 1 = Round Up (ceil),
                2 = Round Down (floor), 3 = Round Nearest (half-even)
    """
    if digits is None:
        digits = 0
    if round_type is None:
        round_type = 0

    # 0 = None: return the value unchanged (no rounding applied).
    if round_type not in (1, 2, 3):
        return value

    d = Decimal(str(value))
    # Pre-round to digits+1 to avoid floating-point artefacts
    d = d.quantize(Decimal(10) ** -(digits + 1), rounding=ROUND_HALF_EVEN)

    if round_type == 1:
        d = d.quantize(Decimal(10) ** -digits, rounding=ROUND_CEILING)
    elif round_type == 2:
        d = d.quantize(Decimal(10) ** -digits, rounding=ROUND_FLOOR)
    elif round_type == 3:
        d = d.quantize(Decimal(10) ** -digits, rounding=ROUND_HALF_UP)

    return float(d)


_DIV_ZERO_RE = re.compile(r"(\([^()]*\)|[\d.]+)\s*/\s*0\.0")
_ALPHA_RE = re.compile(r"(math\.floor|Math\.floor)|[a-zA-Z_]+")


def _handle_division_by_zero(formula: str) -> str:
    """Replace division-by-zero patterns with 0.0 (mirrors Java handleDivisionByZero)."""
    expr = formula
    # Check if formula still has unresolved alphabetic references (skip math.floor)
    has_unresolved = False
    for m in _ALPHA_RE.finditer(expr):
        if m.group().lower() not in ("math.floor",):
            has_unresolved = True
            break

    if not has_unresolved:
        for old, new in [
            ("/(0.0+0.0)", "/0.0"),
            ("/(0.0-0.0)", "/0.0"),
            ("/(0.0*0.0)", "/0.0"),
            ("(0.0/0.0)", "0.0"),
            ("--", "+"),
        ]:
            expr = expr.replace(old, new)
        expr = _DIV_ZERO_RE.sub("0.0", expr)

    return expr


# ─── Tokenized formula evaluator ────────────────────────────────────
#
# Identifiers in formulas refer to components by CODE (pay_components.CODE)
# or by the synthetic form c<id> (e.g. c5). Display names are NEVER looked up.
# Substitution is token-level so prefix-overlapping codes such as "Basic" and
# "BasicRate" no longer collide.

_IDENT_RE = re.compile(r"[A-Za-z_][A-Za-z_0-9]*")
_CID_RE = re.compile(r"^c(\d+)$")

# Keywords that appear in legacy JS-style formulas (e.g.
# "var result=0; if(...) result=...; else result=...;"). They are not
# component codes and must be left untouched during substitution / extraction.
_JS_KEYWORDS = frozenset({"var", "if", "else", "elif", "result", "true", "false", "and", "or", "not"})


def _is_math_function_token(expr: str, match: re.Match) -> bool:
    """True if the matched 'math' identifier is the head of math.floor / math.ceil."""
    if match.group() != "math":
        return False
    tail = expr[match.end(): match.end() + 6]
    return tail.startswith(".floor") or tail.startswith(".ceil")


def _resolve_identifier(ident: str, code_to_id: dict[str, int]) -> int | None:
    """Resolve an identifier to a component_id via CODE map or c<id> form. Returns None if unknown."""
    if ident in code_to_id:
        return code_to_id[ident]
    m = _CID_RE.match(ident)
    if m:
        return int(m.group(1))
    return None


def _extract_formula_dependencies(
    formula: str, code_to_id: dict[str, int]
) -> tuple[set[int], list[str]]:
    """Find component_ids a formula depends on, plus any unknown identifiers."""
    deps: set[int] = set()
    unknowns: list[str] = []
    expr = formula.replace("Math.floor", "math.floor").replace("Math.ceil", "math.ceil")
    for m in _IDENT_RE.finditer(expr):
        if _is_math_function_token(expr, m):
            continue
        ident = m.group()
        if ident == "math" or ident in _JS_KEYWORDS:
            continue
        ref_id = _resolve_identifier(ident, code_to_id)
        if ref_id is None:
            unknowns.append(ident)
        else:
            deps.add(ref_id)
    return deps, unknowns


def _substitute_formula(
    formula: str, cooked: dict[int, float], code_to_id: dict[str, int]
) -> str:
    """Token-level substitution of cooked component values into formula."""
    expr = formula.replace("Math.floor", "math.floor").replace("Math.ceil", "math.ceil")

    def _replace(match: re.Match) -> str:
        if _is_math_function_token(expr, match):
            return match.group()
        ident = match.group()
        if ident == "math" or ident in _JS_KEYWORDS:
            return ident
        ref_id = _resolve_identifier(ident, code_to_id)
        if ref_id is not None and ref_id in cooked:
            return str(cooked[ref_id])
        return ident

    return _IDENT_RE.sub(_replace, expr)


def _replace_unknown_idents_with_zero(
    formula: str, code_to_id: dict[str, int]
) -> str:
    """Replace identifiers that don't resolve to any known component with '0'.

    Used after _substitute_formula so unresolved tokens (formulas referencing
    codes that don't exist in this company's pay_components) degrade to 0
    rather than blowing up _safe_eval. Matches the legacy Java behavior where
    dangling references quietly evaluated as 0. JS keywords (var/if/else/result)
    are preserved so the formula remains parseable.
    """
    expr = formula.replace("Math.floor", "math.floor").replace("Math.ceil", "math.ceil")

    def _replace(match: re.Match) -> str:
        if _is_math_function_token(expr, match):
            return match.group()
        ident = match.group()
        if ident == "math" or ident in _JS_KEYWORDS:
            return ident
        if _resolve_identifier(ident, code_to_id) is not None:
            return ident  # known component — leave for _safe_eval to flag if still uncooked
        return "0"

    return _IDENT_RE.sub(_replace, expr)


def _is_js_style_formula(expr: str) -> bool:
    """True when the formula uses legacy JS-style statement syntax (var / if / else)."""
    return bool(re.search(r"\b(var|if|else)\b", expr))


def _js_to_python(formula: str) -> str:
    """Translate a legacy JS-style if/else formula to executable Python source.

    Handles the constrained grammar actually used in pay_scheme formulas:
        var result=0;
        if (cond) result=expr;
        else if (cond) result=expr;
        else result=expr;

    Boolean operators &&, ||, and the JS idiom of using bitwise &/| as logical
    AND/OR (works in JS because comparisons return 1/0) are rewritten to
    Python's `and`/`or` so operator precedence stays correct relative to ==,
    <=, etc.
    """
    s = formula

    # Drop the `var` keyword — Python doesn't need it
    s = re.sub(r"\bvar\s+", "", s)
    # `else if` -> `elif`
    s = re.sub(r"\belse\s+if\b", "elif", s)
    # Logical operators: handle && / || before bare & / | so they don't double-replace
    s = s.replace("&&", " and ")
    s = s.replace("||", " or ")
    # Bare `&` and `|` used as logical operators in legacy formulas. The lookarounds
    # avoid touching ==, !=, <=, >= and the (already-handled) && / ||.
    s = re.sub(r"(?<![=!<>])&(?![=&])", " and ", s)
    s = re.sub(r"(?<![=!<>])\|(?![=|])", " or ", s)

    # Split on `;` and rebuild as Python statements
    parts = [p.strip() for p in s.split(";") if p.strip()]
    lines: list[str] = []
    for p in parts:
        parsed = _split_if_statement(p)
        if parsed:
            kw, cond, body = parsed
            lines.append(f"{kw} {cond}: {body}")
            continue
        m = re.match(r"^else\s+(.+)$", p)
        if m:
            lines.append(f"else: {m.group(1)}")
            continue
        lines.append(p)
    return "\n".join(lines)


def _split_if_statement(stmt: str) -> tuple[str, str, str] | None:
    """Parse 'if (cond) body' / 'elif (cond) body' into (keyword, cond, body).

    Uses paren-depth tracking so condition parens stay matched and body parens
    (e.g. '(11258*3.25)/100') aren't accidentally absorbed into the condition,
    which a greedy regex would do.
    """
    m = re.match(r"^(if|elif)\s*\(", stmt)
    if not m:
        return None
    kw = m.group(1)
    open_idx = m.end() - 1  # position of the opening '('
    depth = 0
    i = open_idx
    while i < len(stmt):
        c = stmt[i]
        if c == "(":
            depth += 1
        elif c == ")":
            depth -= 1
            if depth == 0:
                break
        i += 1
    if depth != 0:
        return None
    cond = stmt[open_idx: i + 1]
    body = stmt[i + 1:].strip()
    if not body:
        return None
    return kw, cond, body


def _safe_eval(formula: str) -> float:
    """Evaluate a fully-substituted numeric expression or JS-style if/else block.

    Plain expressions ("A*B/100") go through Python's eval. JS-style blocks
    ("var result=0; if(...) result=...; else result=...;") are translated to
    Python statements and exec'd to extract `result`. Any leftover alphabetic
    token (other than math.floor/ceil and the recognized JS keywords) raises
    ValueError.
    """
    expr = formula.replace("Math.floor", "math.floor").replace("Math.ceil", "math.ceil")

    if _is_js_style_formula(expr):
        py_source = _js_to_python(expr)

        # Verify nothing alphabetic remains besides math.floor/ceil and allowed keywords
        check = re.sub(r"\bmath\.(floor|ceil)\b", "", py_source)
        for kw in _JS_KEYWORDS:
            check = re.sub(r"\b" + kw + r"\b", "", check)
        if re.search(r"[A-Za-z_]", check):
            raise ValueError(f"Formula contains unresolved references: {formula}")

        try:
            local_ns: dict = {"math": math, "result": 0}
            exec(py_source, {"__builtins__": {}}, local_ns)  # noqa: S102
            return float(local_ns.get("result", 0) or 0)
        except ZeroDivisionError:
            return 0.0
        except Exception as e:
            raise ValueError(f"Failed to evaluate formula '{formula}': {e}")

    sanitized = re.sub(r"math\.(floor|ceil)", "", expr)
    if re.search(r"[a-zA-Z_]", sanitized):
        raise ValueError(f"Formula contains unresolved references: {formula}")
    try:
        result = eval(expr, {"__builtins__": {}}, {"math": math})  # noqa: S307
    except ZeroDivisionError:
        return 0.0
    except Exception as e:
        raise ValueError(f"Failed to evaluate formula '{formula}': {e}")
    return float(result)


def _evaluate_employee_formulas(
    component_formulas: dict[int, str],
    components_meta: dict[int, dict],
) -> tuple[dict[int, float], dict[int, str], dict[int, str]]:
    """Evaluate all component formulas in dependency order.

    Returns:
        (cooked, errors, warnings)
        - cooked: comp_id → resolved numeric value (computed whenever possible,
                  even when the formula references unknown identifiers — those
                  are treated as 0, matching legacy Java behavior).
        - errors: fatal failures (cyclic dependency, eval error, depends on
                  another failed component). Calculation cannot proceed for
                  these components.
        - warnings: non-fatal issues (unknown identifiers — the formula still
                    evaluates with the unknown tokens substituted to 0). UI
                    should surface these so the scheme can be cleaned up.
    """
    cooked: dict[int, float] = {}
    errors: dict[int, str] = {}
    warnings: dict[int, str] = {}

    # Build code → comp_id map (skip blanks; first one wins on collision)
    code_to_id: dict[str, int] = {}
    for cid, meta in components_meta.items():
        code = (meta.get("code") or "").strip()
        if code and code not in code_to_id:
            code_to_id[code] = cid

    # Pass 1: cook plain-numeric formulas; extract deps for the rest
    deps_map: dict[int, set[int]] = {}
    for comp_id, formula in component_formulas.items():
        try:
            cooked[comp_id] = float(formula)
            deps_map[comp_id] = set()
            continue
        except (ValueError, TypeError):
            pass

        deps, unknowns = _extract_formula_dependencies(formula, code_to_id)
        deps_map[comp_id] = deps

        if unknowns:
            unique_unknowns = sorted(set(unknowns))
            warnings[comp_id] = (
                f"unknown identifier(s) {unique_unknowns} in formula '{formula}' — "
                f"treated as 0 (not found in pay_components.CODE; use a valid CODE or c<id>)"
            )

    # Pass 2: topological resolution — only cook a component once every dep is cooked
    pending = {cid for cid in component_formulas if cid not in cooked}

    while pending:
        progressed = False
        for comp_id in list(pending):
            deps = deps_map.get(comp_id, set())

            # Self-reference is always a cycle
            if comp_id in deps:
                errors[comp_id] = (
                    f"cyclic dependency: component {comp_id} references itself"
                )
                pending.discard(comp_id)
                progressed = True
                continue

            # Bail out if any dep already failed
            failed_deps = deps & errors.keys()
            if failed_deps:
                errors[comp_id] = (
                    f"depends on failed component(s) {sorted(failed_deps)}"
                )
                pending.discard(comp_id)
                progressed = True
                continue

            # Wait if any dep is still uncooked
            if deps - cooked.keys():
                continue

            # All deps ready — substitute, replace any leftover unknown
            # identifiers with 0, then evaluate
            formula = component_formulas[comp_id]
            substituted = _substitute_formula(formula, cooked, code_to_id)
            substituted = _replace_unknown_idents_with_zero(substituted, code_to_id)
            substituted = _handle_division_by_zero(substituted)
            try:
                val = _safe_eval(substituted)
            except ValueError as e:
                errors[comp_id] = str(e)
                pending.discard(comp_id)
                progressed = True
                continue

            meta = components_meta.get(comp_id, {})
            val = _round_value(val, meta.get("roundof"), meta.get("roundof_type"))
            if math.isnan(val) or math.isinf(val):
                val = 0.0

            cooked[comp_id] = val
            pending.discard(comp_id)
            progressed = True

        if not progressed:
            # No comp could advance — remaining set is one or more cycles
            break

    # Anything left pending is part of a cycle
    for comp_id in pending:
        deps = deps_map.get(comp_id, set())
        cycle_partners = sorted(deps & pending)
        errors[comp_id] = (
            f"cyclic dependency: component {comp_id} depends on {cycle_partners} "
            f"which transitively depend back on it"
        )

    return cooked, errors, warnings


def _run_payroll_processing(
    db: Session,
    co_id: int,
    period: PayPeriod,
    pay_scheme_id: int,
    branch_id: int | None,
    user_id: int,
) -> dict:
        """Calculate payroll for all employees in a scheme/period and write
        pay_employee_payroll + pay_employee_payperiod rows. Mirrors Java
        PayProcessServiceImpl.processPayToPayRoll2(). Does NOT commit — the caller
        owns the transaction. Sets period.status_id to Processed on success.
        Raises HTTPException(400) when the scheme has no mapped employees or no
        formulas.
        """
        pay_period_id = period.id

        # ── 2. Fetch employees mapped to this pay scheme ────────────
        emp_rows = db.execute(
            get_payscheme_mapped_employees(),
            {"pay_scheme_id": pay_scheme_id, "branch_id": branch_id},
        ).fetchall()

        if not emp_rows:
            raise HTTPException(
                status_code=400,
                detail="No employees found mapped to this pay scheme",
            )

        employee_ids = [r._mapping["eb_id"] for r in emp_rows]

        # ── 3. Fetch pay scheme formulas ────────────────────────────
        formula_rows = db.execute(
            get_pay_scheme_details_by_id(),
            {"pay_scheme_id": pay_scheme_id},
        ).fetchall()

        if not formula_rows:
            raise HTTPException(
                status_code=400,
                detail="No pay scheme details/formulas found for this scheme",
            )

        # Map: component_id → formula string
        scheme_formulas: dict[int, str] = {}
        for r in formula_rows:
            m = r._mapping
            scheme_formulas[m["component_id"]] = m["formula"] or "0"

        # ── 4. Fetch all component metadata ─────────────────────────
        comp_rows = db.execute(
            get_all_pay_components_for_company(),
            {"co_id": int(co_id)},
        ).fetchall()

        # Map: component_id → {code, name, type, roundof, roundof_type, is_custom_component}
        components_meta: dict[int, dict] = {}
        for r in comp_rows:
            m = dict(r._mapping)
            components_meta[m["id"]] = m

        # ── 5b. Fetch per-employee base structure values ───────────
        # pay_employee_structure carries per-employee amounts that override the
        # scheme's default formula for input (type 0) components — e.g. one
        # employee's BASIC differs from the scheme default. Without this, every
        # employee falls back to the scheme formula's constant. (paycreate.md
        # Step 6 / Step 9b.)
        structure_rows = db.execute(
            get_employee_pay_structure(),
            {"pay_scheme_id": pay_scheme_id},
        ).fetchall()

        # Map: eb_id → {component_id → amount_string}
        emp_structure: dict[int, dict[int, str]] = {}
        for r in structure_rows:
            m = r._mapping
            emp_structure.setdefault(m["eb_id"], {})[m["component_id"]] = str(m["amount"] or "0")

        # ── 6. Fetch custom component values (from upload) ──────────
        # Variable per-period inputs come from pay_components_custom (the manual /
        # Excel grid) and take precedence over the base structure. Matched by
        # period id or the period's date range (saved rows may carry a NULL
        # PAY_PERIOD_ID).
        custom_rows = db.execute(
            get_custom_component_values(),
            {
                "pay_period_id": pay_period_id,
                "from_date": period.from_date,
                "to_date": period.to_date,
            },
        ).fetchall()

        # Map: eb_id → {component_id → value_string}
        emp_custom: dict[int, dict[int, str]] = {}
        for r in custom_rows:
            m = r._mapping
            emp_custom.setdefault(m["eb_id"], {})[m["component_id"]] = str(m["value"] or "0")

        # ── 7. Delete existing payroll & payperiod entries (re-process) ─
        db.execute(
            delete_existing_payroll_for_period(),
            {"pay_period_id": pay_period_id, "pay_scheme_id": pay_scheme_id},
        )
        db.execute(
            delete_existing_payperiod_entries(),
            {"pay_period_id": pay_period_id, "pay_scheme_id": pay_scheme_id},
        )
        db.flush()

        # ── 8. Process each employee ────────────────────────────────
        payroll_records: list[PayEmployeePayroll] = []
        payperiod_records: list[PayEmployeePayperiod] = []
        errors: list[dict] = []

        for emp_id in employee_ids:
            # Build per-employee formula map (copy from scheme)
            emp_formulas = dict(scheme_formulas)

            # Resolve each component's input, highest precedence first:
            #   1. IS_EXCEL_DOWNLOADABLE manual-entry columns (e.g. payable/calendar
            #      days, OT) take their value from the custom grid; 0 when absent.
            #   2. Any other component explicitly supplied via the custom grid.
            #   3. Per-employee base structure override for input (type 0)
            #      components — replaces the scheme formula's default constant.
            #   4. Otherwise the component keeps its pay scheme formula.
            struct_vals = emp_structure.get(emp_id, {})
            custom_vals = emp_custom.get(emp_id, {})
            for comp_id in list(emp_formulas):
                meta = components_meta.get(comp_id)
                if not meta:
                    continue
                if meta.get("is_excel_downloadable"):
                    emp_formulas[comp_id] = custom_vals.get(comp_id, "0")
                elif comp_id in custom_vals:
                    # any other component explicitly supplied via the custom grid
                    emp_formulas[comp_id] = custom_vals[comp_id]
                elif meta.get("type") in (0, "0") and comp_id in struct_vals:
                    # per-employee base structure value (paycreate.md Step 9b)
                    emp_formulas[comp_id] = struct_vals[comp_id]

            # Evaluate all formulas for this employee
            try:
                calculated, eval_errors, eval_warnings = _evaluate_employee_formulas(
                    emp_formulas, components_meta,
                )
            except Exception as e:
                errors.append({"employee_id": emp_id, "error": str(e)})
                continue

            # Surface per-component formula failures (don't drop the whole employee)
            for failed_cid, msg in eval_errors.items():
                meta = components_meta.get(failed_cid, {})
                errors.append({
                    "employee_id": emp_id,
                    "component_id": failed_cid,
                    "component_code": meta.get("code"),
                    "component_name": meta.get("name"),
                    "error": msg,
                })
                logger.warning(
                    "pay_register_process: emp=%s comp=%s formula error: %s",
                    emp_id, failed_cid, msg,
                )

            # Log soft warnings (unknown identifiers — value was still cooked as 0)
            for warn_cid, msg in eval_warnings.items():
                logger.info(
                    "pay_register_process: emp=%s comp=%s formula warning: %s",
                    emp_id, warn_cid, msg,
                )

            # Build PayEmployeePayroll records for each component
            basic_val = 0.0
            net_val = 0.0
            gross_val = 0.0

            for comp_id, amount in calculated.items():
                rec = PayEmployeePayroll(
                    eb_id=emp_id,
                    component_id=comp_id,
                    payperiod_id=pay_period_id,
                    payscheme_id=pay_scheme_id,
                    amount=amount,
                    status_id=EMP_PAY_ROLL_SUCCESS,
                    businessunit_id=branch_id,
                    source="PROCESSED",
                    updated_by=user_id,
                )
                payroll_records.append(rec)

                # Track summary values for pay_employee_payperiod
                meta = components_meta.get(comp_id, {})
                comp_name = (meta.get("name") or "").upper()
                if "BASIC" in comp_name:
                    basic_val = amount
                elif "NET" in comp_name:
                    net_val = amount
                elif "GROSS" in comp_name and "DEDUCTION" not in comp_name:
                    gross_val = amount

            # Build pay_employee_payperiod summary record
            pp_rec = PayEmployeePayperiod(
                employeeid=emp_id,
                pay_period_id=pay_period_id,
                pay_scheme_id=pay_scheme_id,
                basic=basic_val,
                net=net_val,
                gross=gross_val,
                status_id=EMP_PAY_ROLL_SUCCESS,
                updated_by=user_id,
            )
            payperiod_records.append(pp_rec)

        # ── 9. Bulk insert ──────────────────────────────────────────
        if payroll_records:
            db.add_all(payroll_records)
            db.add_all(payperiod_records)

        # ── 10. Mark the pay period Processed (caller commits) ──────
        period.status_id = int(STATUS_PROCESSED)
        period.updated_by = user_id

        return {
            "message": "Payroll processed successfully",
            "employees_processed": len(employee_ids) - len(errors),
            "total_employees": len(employee_ids),
            "records_created": len(payroll_records),
            "errors": errors if errors else None,
        }


@router.post("/pay_register_process")
def pay_register_process(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Re-process payroll for an existing pay period (list/view 'Process' action).

    Body: { "pay_period_id": int, "pay_scheme_id": int, "branch_id": int|null }
    """
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        body = parse_json_body(request)
        user_id = token_data.get("user_id", 0)
        pay_period_id = body.get("pay_period_id")
        pay_scheme_id = body.get("pay_scheme_id")
        branch_id = body.get("branch_id")

        if not pay_period_id or not pay_scheme_id:
            raise HTTPException(
                status_code=400,
                detail="pay_period_id and pay_scheme_id are required",
            )

        period = db.query(PayPeriod).filter(
            PayPeriod.id == int(pay_period_id),
            PayPeriod.co_id == int(co_id),
        ).first()
        if not period:
            raise HTTPException(status_code=404, detail="Pay period not found")

        summary = _run_payroll_processing(
            db, int(co_id), period, int(pay_scheme_id),
            int(branch_id) if branch_id else None, user_id,
        )
        db.commit()
        return {"data": summary}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.exception("Payroll processing failed")
        raise HTTPException(status_code=500, detail=str(e))
