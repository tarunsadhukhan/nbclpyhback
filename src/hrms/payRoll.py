"""HRMS Pay Roll endpoints — pay_components_custom upload, search, save, template."""
import logging
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import bindparam, text
from sqlalchemy.orm import Session

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import get_tenant_db
from src.models.hrms import PayComponentsCustom

from .query import (
    find_pay_period_by_dates,
    get_branches_for_company,
    get_categories_for_company,
    get_departments_for_company,
    get_excel_downloadable_component_codes,
    get_pay_roll_custom_components,
    get_pay_roll_employee_list,
    get_pay_roll_existing_pcc,
    get_pay_schemes_for_company,
    soft_delete_pcc_for_period,
)
from src.common.utils import parse_json_body

logger = logging.getLogger(__name__)
router = APIRouter()

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ─── Helpers ───────────────────────────────────────────────────────


def _require_co_id(request: Request) -> int:
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")
    try:
        return int(co_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="co_id must be an integer")


def _parse_date(value: str | None, field: str) -> str:
    if not value:
        raise HTTPException(status_code=400, detail=f"{field} is required")
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"{field} must be in YYYY-MM-DD format",
        )
    return value


def _resolve_pay_period_id(
    db: Session, branch_id: int, from_date: str, to_date: str,
) -> int | None:
    row = db.execute(
        find_pay_period_by_dates(),
        {"branch_id": branch_id, "from_date": from_date, "to_date": to_date},
    ).fetchone()
    return int(row._mapping["id"]) if row else None


def _load_setup_options(
    db: Session, co_id: int, branch_id: int | None = None,
) -> dict:
    schemes = db.execute(get_pay_schemes_for_company(), {"co_id": co_id}).fetchall()
    branches = db.execute(get_branches_for_company(), {"co_id": co_id}).fetchall()
    dept_cat_params = {"co_id": co_id, "branch_id": branch_id}
    depts = db.execute(get_departments_for_company(), dept_cat_params).fetchall()
    cats = db.execute(get_categories_for_company(), dept_cat_params).fetchall()

    def _to_options(rows):
        return [
            {"label": r._mapping["name"], "value": str(r._mapping["id"])}
            for r in rows
        ]

    return {
        "pay_schemes": _to_options(schemes),
        "branches": _to_options(branches),
        "departments": _to_options(depts),
        "categories": _to_options(cats),
    }


# ─── Setup (filter dropdowns) ──────────────────────────────────────


@router.get("/pay_roll_setup")
def pay_roll_setup(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        co_id = _require_co_id(request)
        branch_id_raw = request.query_params.get("branch_id")
        branch_id: int | None = None
        if branch_id_raw:
            try:
                branch_id = int(branch_id_raw)
            except ValueError:
                branch_id = None
        return {"data": _load_setup_options(db, co_id, branch_id)}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_roll_setup failed")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Search Attendance — load components + employees + existing values ─


@router.post("/pay_roll_data")
def pay_roll_data(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Returns the dynamic pay component columns and employees for the grid.

    Body: { from_date, to_date, branch_id, pay_scheme_id?, dept_id?, emp_category_id? }
    """
    try:
        co_id = _require_co_id(request)
        body = parse_json_body(request)

        from_date = _parse_date(body.get("from_date"), "from_date")
        to_date = _parse_date(body.get("to_date"), "to_date")
        branch_id = body.get("branch_id")
        if not branch_id:
            raise HTTPException(status_code=400, detail="branch_id is required")

        pay_scheme_id = body.get("pay_scheme_id")
        dept_id = body.get("dept_id")
        emp_category_id = body.get("emp_category_id")

        comp_rows = db.execute(
            get_pay_roll_custom_components(),
            {"pay_scheme_id": int(pay_scheme_id) if pay_scheme_id else None},
        ).fetchall()
        pay_components = [
            {
                "id": int(r._mapping["id"]),
                "code": r._mapping["code"],
                "name": r._mapping["name"],
            }
            for r in comp_rows
        ]

        emp_rows = db.execute(
            get_pay_roll_employee_list(),
            {
                "branch_id": int(branch_id),
                "pay_scheme_id": int(pay_scheme_id) if pay_scheme_id else None,
                "dept_id": int(dept_id) if dept_id else None,
                "emp_category_id": int(emp_category_id) if emp_category_id else None,
            },
        ).fetchall()

        pay_period_id = _resolve_pay_period_id(db, branch_id, from_date, to_date)

        existing_rows = db.execute(
            get_pay_roll_existing_pcc(),
            {
                "pay_period_id": pay_period_id,
                "from_date": from_date,
                "to_date": to_date,
            },
        ).fetchall()

        pcc_by_emp: dict[int, list[dict]] = {}
        for r in existing_rows:
            m = r._mapping
            emp_id = int(m["emp_id"])
            try:
                value = float(m["value"]) if m["value"] is not None else 0.0
            except (TypeError, ValueError):
                value = 0.0
            pcc_by_emp.setdefault(emp_id, []).append({
                "component_id": int(m["component_id"]),
                "value": value,
            })

        emp_pay_scheme_list = []
        for r in emp_rows:
            m = r._mapping
            eid = int(m["emp_id"])
            emp_pay_scheme_list.append({
                "emp_id": eid,
                "emp_code": m["emp_code"],
                "emp_name": (m["emp_name"] or "").strip(),
                "pay_scheme_name": m["pay_scheme_name"],
                "pcc_list": pcc_by_emp.get(eid, []),
            })

        return {
            "data": {
                "pay_components": pay_components,
                "emp_pay_scheme_list": emp_pay_scheme_list,
                "pay_period_id": pay_period_id,
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_roll_data failed")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Save — upsert pay_components_custom rows ──────────────────────


@router.post("/pay_roll_save")
def pay_roll_save(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Save / replace pay_components_custom for the given period and employees.

    Body: {
        from_date, to_date, pay_period_id?,
        employees: [{ eb_id, pay_comp_values: { "<component_id>": <number> } }]
    }
    """
    try:
        co_id = _require_co_id(request)
        body = parse_json_body(request)
        user_id = token_data.get("user_id", 0)

        from_date = _parse_date(body.get("from_date"), "from_date")
        to_date = _parse_date(body.get("to_date"), "to_date")
        pay_period_id = body.get("pay_period_id")
        employees = body.get("employees") or []
        if not isinstance(employees, list) or not employees:
            raise HTTPException(status_code=400, detail="employees list is required")

        if pay_period_id is None:
            pay_period_id = _resolve_pay_period_id(db, co_id, from_date, to_date)

        emp_ids = [int(emp["eb_id"]) for emp in employees if emp.get("eb_id")]

        if emp_ids:
            db.execute(
                soft_delete_pcc_for_period(),
                {
                    "updated_by": user_id,
                    "pay_period_id": int(pay_period_id) if pay_period_id else None,
                    "from_date": from_date,
                    "to_date": to_date,
                    "emp_ids": emp_ids,
                },
            )

        new_rows: list[PayComponentsCustom] = []
        for emp in employees:
            eb_id = int(emp["eb_id"])
            values = emp.get("pay_comp_values") or {}
            if not isinstance(values, dict):
                continue
            for component_id_raw, val in values.items():
                try:
                    component_id = int(component_id_raw)
                except (TypeError, ValueError):
                    continue
                try:
                    val_num = float(val) if val is not None else 0.0
                except (TypeError, ValueError):
                    val_num = 0.0
                new_rows.append(PayComponentsCustom(
                    component_id=component_id,
                    value=str(val_num),
                    employeeid=eb_id,
                    status_id=1,
                    pay_period_id=int(pay_period_id) if pay_period_id else None,
                    from_date=from_date,
                    to_date=to_date,
                    updated_by=user_id,
                ))

        if new_rows:
            db.add_all(new_rows)

        db.commit()
        return {
            "data": {
                "message": "Pay roll data saved successfully",
                "records_saved": len(new_rows),
                "pay_period_id": pay_period_id,
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.exception("pay_roll_save failed")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Fetch Attendance / Generic / Cumulative (helpers) ─────────────
#
# These mirror the legacy "Fetch Attendance / Paygeneric / Cumulative" buttons.
# In the legacy app each one calls a separate service that recomputes the
# underlying attendance / generic / cumulative tables for the period. The
# current backend doesn't own those computations yet, so the endpoints accept
# the request and return a success acknowledgement so the UI flow stays intact.


def _fetch_helper_response(label: str, from_date: str, to_date: str) -> dict:
    return {
        "data": {
            "message": f"{label} fetch acknowledged",
            "from_date": from_date,
            "to_date": to_date,
            "records_updated": 0,
        }
    }


@router.get("/pay_roll_fetch_attendance")
def pay_roll_fetch_attendance(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        _require_co_id(request)
        from_date = _parse_date(request.query_params.get("from_date"), "from_date")
        to_date = _parse_date(request.query_params.get("to_date"), "to_date")
        return _fetch_helper_response("Attendance", from_date, to_date)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_roll_fetch_attendance failed")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/pay_roll_fetch_generic")
def pay_roll_fetch_generic(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        _require_co_id(request)
        from_date = _parse_date(request.query_params.get("from_date"), "from_date")
        to_date = _parse_date(request.query_params.get("to_date"), "to_date")
        return _fetch_helper_response("Pay Generic", from_date, to_date)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_roll_fetch_generic failed")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/pay_roll_fetch_cumulative")
def pay_roll_fetch_cumulative(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    try:
        _require_co_id(request)
        from_date = _parse_date(request.query_params.get("from_date"), "from_date")
        to_date = _parse_date(request.query_params.get("to_date"), "to_date")
        return _fetch_helper_response("Cumulative", from_date, to_date)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_roll_fetch_cumulative failed")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Download Excel Template ───────────────────────────────────────


@router.post("/pay_roll_download_template")
def pay_roll_download_template(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Build a blank Excel template with employee rows + custom-component columns.

    Body: { branch_id, pay_scheme_id?, dept_id?, emp_category_id? }
    """
    try:
        co_id = _require_co_id(request)
        body = parse_json_body(request)

        branch_id = body.get("branch_id")
        if not branch_id:
            raise HTTPException(status_code=400, detail="branch_id is required")

        pay_scheme_id = body.get("pay_scheme_id")
        dept_id = body.get("dept_id")
        emp_category_id = body.get("emp_category_id")

        comp_rows = db.execute(
            get_pay_roll_custom_components(),
            {"pay_scheme_id": int(pay_scheme_id) if pay_scheme_id else None},
        ).fetchall()
        components = [
            (int(r._mapping["id"]), r._mapping["code"], r._mapping["name"])
            for r in comp_rows
        ]

        emp_rows = db.execute(
            get_pay_roll_employee_list(),
            {
                "branch_id": int(branch_id),
                "pay_scheme_id": int(pay_scheme_id) if pay_scheme_id else None,
                "dept_id": int(dept_id) if dept_id else None,
                "emp_category_id": int(emp_category_id) if emp_category_id else None,
            },
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "PayRoll"

        headers = ["Employee Code", "Employee Name", "PayScheme Name"]
        for _cid, code, _name in components:
            headers.append(code)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, r in enumerate(emp_rows, start=2):
            m = r._mapping
            ws.cell(row=row_idx, column=1, value=m["emp_code"])
            ws.cell(row=row_idx, column=2, value=(m["emp_name"] or "").strip())
            ws.cell(row=row_idx, column=3, value=m["pay_scheme_name"])
            for col_offset, _ in enumerate(components, start=4):
                ws.cell(row=row_idx, column=col_offset, value=0)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        return StreamingResponse(
            iter([bio.getvalue()]),
            media_type=XLSX_MEDIA_TYPE,
            headers={
                "Content-Disposition": 'attachment; filename="PayRoll.xlsx"',
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("pay_roll_download_template failed")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Upload Excel ──────────────────────────────────────────────────


@router.post("/pay_roll_upload")
def pay_roll_upload(
    request: Request,
    file: UploadFile = File(...),
    from_date: str = Form(...),
    to_date: str = Form(...),
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Upload an Excel file with employee × pay-component values.

    Expected columns: 'Employee Code', 'Employee Name', 'PayScheme Name',
    followed by one column per custom-component CODE.
    """
    try:
        co_id = _require_co_id(request)
        user_id = token_data.get("user_id", 0)
        from_date = _parse_date(from_date, "from_date")
        to_date = _parse_date(to_date, "to_date")

        if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="File must be an Excel (.xlsx/.xls)")

        content = file.file.read()
        try:
            wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        header_row = [str(c).strip() if c is not None else "" for c in rows[0]]
        if len(header_row) < 4:
            raise HTTPException(
                status_code=400,
                detail="Header must include Employee Code, Employee Name, PayScheme Name "
                       "and at least one component column",
            )

        # Component CODEs are globally unique, so a single code -> id map resolves
        # whatever columns were downloaded for the selected scheme — the uploaded
        # file's PayScheme column is not needed (the template leaves it blank).
        component_rows = db.execute(
            get_excel_downloadable_component_codes(), {},
        ).fetchall()
        code_to_component_id: dict[str, int] = {
            r._mapping["code"]: int(r._mapping["id"]) for r in component_rows
        }

        component_headers = [
            (idx, str(h).strip() if h is not None else "")
            for idx, h in enumerate(header_row[3:], start=3)
        ]

        emp_code_to_id: dict[str, int] = {}
        for data_row in rows[1:]:
            if not data_row or all(v is None for v in data_row):
                continue
            emp_code = str(data_row[0]).strip() if data_row[0] is not None else ""
            if emp_code and emp_code not in emp_code_to_id:
                emp_code_to_id[emp_code] = 0

        if emp_code_to_id:
            lookup_stmt = text("""
                SELECT eb_id, emp_code
                FROM hrms_ed_official_details
                WHERE active = 1 AND emp_code IN :codes
            """).bindparams(bindparam("codes", expanding=True))
            lookup = db.execute(
                lookup_stmt,
                {"codes": list(emp_code_to_id.keys())},
            ).fetchall()
            for r in lookup:
                emp_code_to_id[r._mapping["emp_code"]] = int(r._mapping["eb_id"])

        pay_period_id = _resolve_pay_period_id(db, co_id, from_date, to_date)

        affected_emp_ids: list[int] = []
        new_rows: list[PayComponentsCustom] = []
        skipped: list[str] = []

        for data_row in rows[1:]:
            if not data_row or all(v is None for v in data_row):
                continue
            emp_code = str(data_row[0]).strip() if data_row[0] is not None else ""
            if not emp_code:
                continue
            eb_id = emp_code_to_id.get(emp_code, 0)
            if not eb_id:
                skipped.append(emp_code)
                continue
            affected_emp_ids.append(eb_id)
            for col_idx, header in component_headers:
                component_id = code_to_component_id.get(header)
                if component_id is None:
                    continue
                raw = data_row[col_idx] if col_idx < len(data_row) else None
                try:
                    val_num = float(raw) if raw is not None and raw != "" else 0.0
                except (TypeError, ValueError):
                    val_num = 0.0
                new_rows.append(PayComponentsCustom(
                    component_id=component_id,
                    value=str(val_num),
                    employeeid=eb_id,
                    status_id=1,
                    pay_period_id=int(pay_period_id) if pay_period_id else None,
                    from_date=from_date,
                    to_date=to_date,
                    updated_by=user_id,
                ))

        if affected_emp_ids:
            db.execute(
                soft_delete_pcc_for_period(),
                {
                    "updated_by": user_id,
                    "pay_period_id": int(pay_period_id) if pay_period_id else None,
                    "from_date": from_date,
                    "to_date": to_date,
                    "emp_ids": list(set(affected_emp_ids)),
                },
            )

        if new_rows:
            db.add_all(new_rows)

        db.commit()

        return {
            "data": {
                "message": "Pay roll Excel uploaded successfully",
                "records_saved": len(new_rows),
                "skipped_employee_codes": skipped,
                "pay_period_id": pay_period_id,
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.exception("pay_roll_upload failed")
        raise HTTPException(status_code=500, detail=str(e))
