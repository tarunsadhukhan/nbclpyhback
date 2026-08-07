"""
HRMS Bio Attendance — employee↔device mapping + bio data processing.

Ported from the SJM deployment (sjmvowerp3be bioEmpMstLink.py / bioAttUpdation.py),
trimmed to this repo's scope:

* Tab 1 — Employee Bio Link: CRUD on `tbl_master_bio_link_mst` (match_type='E'),
  linking an employee (master_id = eb_id) to a biometric device id (bio_dev_id).
  Emp code / name are always joined from hrms_ed_official_details.

* Tab 2 — Bio Data Process: processes raw punches in `bio_attendance_table` into
  `bio_attendance_basic` (eSSL-style daily row per employee/day, via the pure
  essl_resolver) and `bio_attendance_process` (spell rows projected from basic).
  Does NOT write to `daily_attendance` — that final-process step is deliberately
  out of scope here.

All bio tables are created lazily (CREATE TABLE IF NOT EXISTS) — no migration
needed. Endpoints are registered under /api/hrmsMasters.
"""

import os
import re
import threading
from datetime import date, datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.sql import text

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import get_tenant_db
from src.hrms.essl_resolver import Punch, segment_sessions, resolve_session

router = APIRouter()

SPELL_HOURS = 8.0

# ─── Lazy table creation ────────────────────────────────────────────

# Matches the production schema exactly: master_id is the eb_id, match_type is
# 'E' for employee links, bio_dev_id is the device-side enrolled user id. The
# employee's emp_code/name are always JOINed from hrms_ed_official_details —
# never mirrored here (older tenants lack the mirror columns).
_CREATE_LINK_MST_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS tbl_master_bio_link_mst (
        tbl_mst_bio_link_id INT NOT NULL AUTO_INCREMENT,
        master_id           INT NULL,
        match_type          VARCHAR(10) NULL,
        bio_dev_id          INT NULL,
        PRIMARY KEY (tbl_mst_bio_link_id)
    )
    """
)

# Raw punches — written by an external device feed and/or manual entry.
_CREATE_BIO_ATT_TABLE_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS bio_attendance_table (
        bio_att_id       INT NOT NULL AUTO_INCREMENT,
        bio_att_log_id   BIGINT NULL,
        emp_code         VARCHAR(50) NULL,
        emp_anme         VARCHAR(120) NULL,
        bio_id           INT NULL,
        log_date         DATETIME NULL,
        company_name     VARCHAR(120) NULL,
        department       VARCHAR(120) NULL,
        designation      VARCHAR(120) NULL,
        employement_type VARCHAR(60) NULL,
        device_direction VARCHAR(10) NULL,
        device_name      VARCHAR(120) NULL,
        eb_id            INT NULL,
        dept_id          INT NULL,
        desig_id         INT NULL,
        device_id        INT NULL,
        manual_auto      TINYINT NULL,
        PRIMARY KEY (bio_att_id),
        KEY idx_bio_att_log_date (log_date),
        KEY idx_bio_att_emp_code (emp_code)
    )
    """
)

_CREATE_BIO_BASIC_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS bio_attendance_basic (
        id                  INT AUTO_INCREMENT PRIMARY KEY,
        eb_id               INT       NOT NULL,
        emp_code            VARCHAR(32) NULL,
        bio_id              INT       NULL,
        dept_id             INT       NULL,
        desig_id            INT       NULL,
        tran_date           DATE      NOT NULL,
        shift               VARCHAR(8) NOT NULL,
        sched_in_time       TIME      NULL,
        sched_out_time      TIME      NULL,
        actual_in           DATETIME  NULL,
        actual_out          DATETIME  NULL,
        work_dur_minutes    INT       NOT NULL DEFAULT 0,
        ot_minutes          INT       NOT NULL DEFAULT 0,
        break_minutes       INT       NOT NULL DEFAULT 0,
        total_dur_minutes   INT       NOT NULL DEFAULT 0,
        late_by_minutes     INT       NOT NULL DEFAULT 0,
        early_going_minutes INT       NOT NULL DEFAULT 0,
        status              VARCHAR(64) NOT NULL,
        punch_records       TEXT      NULL,
        created_at          DATETIME  NOT NULL DEFAULT CURRENT_TIMESTAMP,
        UNIQUE KEY uniq_eb_date (eb_id, tran_date)
    )
    """
)

_CREATE_BIO_PROCESS_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS bio_attendance_process (
        bio_att_process_id INT AUTO_INCREMENT PRIMARY KEY,
        eb_id              INT NOT NULL,
        bio_id             INT NULL,
        dept_id            INT NULL,
        desig_id           INT NULL,
        attendance_date    DATE NOT NULL,
        spell_name         VARCHAR(8) NULL,
        attendance_type    VARCHAR(4) NULL,
        attendance_source  VARCHAR(8) NULL,
        check_in           DATETIME NULL,
        check_out          DATETIME NULL,
        Time_duration      DECIMAL(5,2) NULL,
        Working_hours      DECIMAL(5,2) NULL,
        Ot_hours           DECIMAL(5,2) NULL,
        spell_start_time   TIME NULL,
        spell_end_time     TIME NULL,
        spell_hours        DECIMAL(4,1) NULL,
        processed          TINYINT NOT NULL DEFAULT 0,
        created_at         DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
        KEY idx_bio_process_date (attendance_date)
    )
    """
)

def ensure_bio_tables(db: Session) -> None:
    """Create every bio-attendance table that does not exist yet."""
    for stmt in (
        _CREATE_LINK_MST_SQL,
        _CREATE_BIO_ATT_TABLE_SQL,
        _CREATE_BIO_BASIC_SQL,
        _CREATE_BIO_PROCESS_SQL,
    ):
        db.execute(stmt)
    db.commit()


# ─── Employee Bio Link (tab 1) helpers ──────────────────────────────


def _parse_branch_ids(raw: str | None) -> list[int]:
    if not raw:
        return []
    out: list[int] = []
    for tok in str(raw).split(","):
        tok = tok.strip()
        if not tok:
            continue
        try:
            out.append(int(tok))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid branch_id: {tok}")
    return out


def _to_int(value, field: str):
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"Invalid integer for {field}")


def _validate_payload(db: Session, body: dict) -> dict:
    eb_id = _to_int(body.get("master_id"), "master_id")
    if not eb_id:
        raise HTTPException(status_code=400, detail="Employee (master_id) is required")
    bio_dev_id = _to_int(body.get("bio_dev_id"), "bio_dev_id")
    if bio_dev_id is None:
        raise HTTPException(status_code=400, detail="Bio device id is required")

    emp = db.execute(
        text("""
            SELECT o.emp_code
            FROM hrms_ed_official_details o
            WHERE o.eb_id = :eb_id AND COALESCE(o.active, 1) = 1
            LIMIT 1
        """),
        {"eb_id": eb_id},
    ).fetchone()
    if not emp:
        raise HTTPException(status_code=400, detail="Employee not found")

    return {
        "master_id": eb_id,
        "bio_dev_id": bio_dev_id,
    }


def _check_duplicates(db: Session, data: dict, exclude_id: int | None = None) -> None:
    exclude_sql = "AND tbl_mst_bio_link_id <> :exclude_id" if exclude_id else ""
    params: dict = {"master_id": data["master_id"], "bio_dev_id": data["bio_dev_id"]}
    if exclude_id:
        params["exclude_id"] = exclude_id

    dup_emp = db.execute(
        text(f"""
            SELECT COUNT(*) AS cnt FROM tbl_master_bio_link_mst
            WHERE match_type = 'E' AND master_id = :master_id {exclude_sql}
        """),
        params,
    ).fetchone()
    if dup_emp and dup_emp.cnt > 0:
        raise HTTPException(status_code=400, detail="This employee already has a bio link")

    dup_bio = db.execute(
        text(f"""
            SELECT COUNT(*) AS cnt FROM tbl_master_bio_link_mst
            WHERE match_type = 'E' AND bio_dev_id = :bio_dev_id {exclude_sql}
        """),
        params,
    ).fetchone()
    if dup_bio and dup_bio.cnt > 0:
        raise HTTPException(
            status_code=400,
            detail="This bio device id is already linked to another employee",
        )


def _list_query(branch_filter_sql: str = ""):
    return text(f"""
        SELECT
            l.tbl_mst_bio_link_id,
            l.master_id,
            l.bio_dev_id,
            o.emp_code,
            CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) AS employee_name,
            o.branch_id,
            b.branch_name
        FROM tbl_master_bio_link_mst l
        LEFT JOIN hrms_ed_personal_details p ON p.eb_id = l.master_id
        LEFT JOIN hrms_ed_official_details o ON o.eb_id = l.master_id AND COALESCE(o.active, 1) = 1
        LEFT JOIN branch_mst b ON b.branch_id = o.branch_id
        WHERE l.match_type = 'E'
          {branch_filter_sql}
          AND (:search IS NULL
               OR o.emp_code LIKE :search
               OR CAST(l.bio_dev_id AS CHAR) LIKE :search
               OR CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) LIKE :search)
        ORDER BY l.tbl_mst_bio_link_id DESC
    """)


# ─── Employee Bio Link (tab 1) endpoints ────────────────────────────


@router.get("/bio_emp_link_setup")
async def bio_emp_link_setup(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Setup endpoint returning active employees (filtered by branch)."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        ensure_bio_tables(db)
        branch_ids = _parse_branch_ids(request.query_params.get("branch_id"))

        branch_filter_sql = ""
        params: dict = {}
        if branch_ids:
            placeholders = ",".join(f":b{i}" for i in range(len(branch_ids)))
            branch_filter_sql = f"AND o.branch_id IN ({placeholders})"
            params = {f"b{i}": bid for i, bid in enumerate(branch_ids)}

        employees = db.execute(
            text(f"""
                SELECT
                    p.eb_id,
                    o.emp_code,
                    CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) AS full_name,
                    l.tbl_mst_bio_link_id AS link_id,
                    l.bio_dev_id
                FROM hrms_ed_personal_details p
                JOIN hrms_ed_official_details o ON o.eb_id = p.eb_id AND COALESCE(o.active, 1) = 1
                LEFT JOIN tbl_master_bio_link_mst l ON l.master_id = p.eb_id AND l.match_type = 'E'
                WHERE COALESCE(p.active, 1) = 1
                  {branch_filter_sql}
                ORDER BY o.emp_code
            """),
            params,
        ).fetchall()

        return {"employees": [dict(r._mapping) for r in employees]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/get_bio_emp_link_table")
async def get_bio_emp_link_table(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated list of employee bio links (match_type = 'E')."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        ensure_bio_tables(db)
        search = request.query_params.get("search")
        search_param = f"%{search}%" if search else None

        page = int(request.query_params.get("page", 1))
        limit = int(request.query_params.get("limit", 10))

        branch_ids = _parse_branch_ids(request.query_params.get("branch_id"))

        params: dict = {"search": search_param}
        branch_filter_sql = ""
        if branch_ids:
            placeholders = ",".join(f":b{i}" for i in range(len(branch_ids)))
            branch_filter_sql = f"AND o.branch_id IN ({placeholders})"
            for i, bid in enumerate(branch_ids):
                params[f"b{i}"] = bid

        result = db.execute(_list_query(branch_filter_sql), params).fetchall()

        all_data = [dict(r._mapping) for r in result]
        total = len(all_data)
        start = (page - 1) * limit
        return {
            "data": all_data[start : start + limit],
            "total": total,
            "page": page,
            "limit": limit,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/get_bio_emp_link_by_id/{record_id}")
async def get_bio_emp_link_by_id(
    record_id: int,
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Single employee bio link row by id."""
    try:
        row = db.execute(
            text("""
                SELECT l.tbl_mst_bio_link_id, l.master_id, l.bio_dev_id,
                       o.emp_code,
                       CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) AS employee_name
                FROM tbl_master_bio_link_mst l
                LEFT JOIN hrms_ed_personal_details p ON p.eb_id = l.master_id
                LEFT JOIN hrms_ed_official_details o ON o.eb_id = l.master_id AND COALESCE(o.active, 1) = 1
                WHERE l.tbl_mst_bio_link_id = :id AND l.match_type = 'E'
                LIMIT 1
            """),
            {"id": record_id},
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Record not found")
        return {"data": dict(row._mapping)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/bio_emp_link_create")
async def bio_emp_link_create(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Create a new employee bio link (match_type = 'E')."""
    try:
        body = await request.json()
        co_id = body.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        ensure_bio_tables(db)
        data = _validate_payload(db, body)
        _check_duplicates(db, data)

        result = db.execute(
            text("""
                INSERT INTO tbl_master_bio_link_mst
                    (master_id, match_type, bio_dev_id)
                VALUES
                    (:master_id, 'E', :bio_dev_id)
            """),
            data,
        )
        db.commit()
        return {
            "message": "Bio link created successfully",
            "tbl_mst_bio_link_id": result.lastrowid,
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/bio_emp_link_edit/{record_id}")
async def bio_emp_link_edit(
    record_id: int,
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Update an existing employee bio link."""
    try:
        body = await request.json()
        existing = db.execute(
            text("""
                SELECT tbl_mst_bio_link_id FROM tbl_master_bio_link_mst
                WHERE tbl_mst_bio_link_id = :id AND match_type = 'E'
            """),
            {"id": record_id},
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Record not found")

        data = _validate_payload(db, body)
        _check_duplicates(db, data, exclude_id=record_id)

        params = dict(data)
        params["id"] = record_id
        db.execute(
            text("""
                UPDATE tbl_master_bio_link_mst SET
                    master_id   = :master_id,
                    bio_dev_id  = :bio_dev_id
                WHERE tbl_mst_bio_link_id = :id
            """),
            params,
        )
        db.commit()
        return {"message": "Bio link updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/bio_emp_link_delete/{record_id}")
async def bio_emp_link_delete(
    record_id: int,
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Delete an employee bio link (hard delete — table has no active flag)."""
    try:
        row = db.execute(
            text("""
                SELECT tbl_mst_bio_link_id FROM tbl_master_bio_link_mst
                WHERE tbl_mst_bio_link_id = :id AND match_type = 'E'
            """),
            {"id": record_id},
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Record not found")

        db.execute(
            text("DELETE FROM tbl_master_bio_link_mst WHERE tbl_mst_bio_link_id = :id"),
            {"id": record_id},
        )
        db.commit()
        return {"message": "Bio link deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ─── Punch source: <BIO_PUNCH_LOGS_DB>.punch_logs ───────────────────
# The biometric devices push their logs into the `punch_logs` table of a
# separate schema on the SAME MySQL server (default: sls). `user_id` there is
# the bio device id (bio_dev_id in tbl_master_bio_link_mst). We copy new rows
# into bio_attendance_table before processing; the device gives no reliable
# in/out direction, and the essl resolver is positional anyway.


def _punch_logs_db() -> str:
    db_name = os.environ.get("BIO_PUNCH_LOGS_DB", "sls").strip()
    if not re.fullmatch(r"[A-Za-z0-9_]+", db_name):
        raise ValueError(f"Invalid BIO_PUNCH_LOGS_DB: {db_name!r}")
    return db_name


def sync_punch_logs(db: Session) -> dict:
    """Copy new rows from <punch_logs_db>.punch_logs into bio_attendance_table.

    user_id -> emp_code + bio_id (device-side id), punch_time -> log_date,
    device_sn -> device_name. Idempotent: dedupes on (emp_code, log_date), so
    it needs no high-water mark on the source (whose `id` has no PK on some
    deployments). Non-fatal: if the source table/schema is missing, returns the
    error instead of raising so processing can continue on existing punches.
    """
    try:
        src = _punch_logs_db()
        res = db.execute(text(
            f"""
            INSERT INTO bio_attendance_table
                (emp_code, bio_id, log_date, device_direction, device_name, manual_auto)
            SELECT p.user_id,
                   CASE WHEN p.user_id REGEXP '^[0-9]+$'
                        THEN CAST(p.user_id AS UNSIGNED) ELSE NULL END,
                   p.punch_time,
                   '',
                   p.device_sn,
                   0
            FROM `{src}`.punch_logs p
            WHERE p.punch_time IS NOT NULL
              AND NOT EXISTS (
                  SELECT 1 FROM bio_attendance_table b
                  -- explicit COLLATE: punch_logs (unicode_ci) vs tenant table
                  -- (0900_ai_ci) otherwise throw an illegal-mix-of-collations error
                  WHERE b.emp_code = p.user_id COLLATE utf8mb4_0900_ai_ci
                    AND b.log_date = p.punch_time
              )
            """
        ))
        db.commit()
        return {"source": f"{src}.punch_logs", "inserted": int(res.rowcount or 0)}
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        return {"source": "punch_logs", "inserted": 0, "error": str(e)}


# ─── Bio data process (tab 2) — punch preparation ───────────────────

# Back-fill emp_code from the linked employee's official record, matching the
# device bio id (bio_dev_id) against the punch's bio_id.
_EMPCODE_BACKFILL_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN tbl_master_bio_link_mst m
        ON m.match_type = 'E'
       AND m.bio_dev_id = b.bio_id
    JOIN hrms_ed_official_details o
        ON o.eb_id = m.master_id AND COALESCE(o.active, 1) = 1
       SET b.emp_code = o.emp_code
    """
)

# Fill a blank bio_att_log_id with a generated local sequence (base 500000).
_LOGID_INIT_SQL = text("SET @ba_seq := :base - 1")
_LOGID_FILL_SQL = text(
    """
    UPDATE bio_attendance_table
       SET bio_att_log_id = (@ba_seq := @ba_seq + 1)
     WHERE bio_att_log_id IS NULL OR bio_att_log_id = 0
     ORDER BY bio_att_id
    """
)

# Set device_id from the punch direction: 'in' -> 22, anything else -> 14.
_DEVICE_ID_SQL = text(
    """
    UPDATE bio_attendance_table
       SET device_id = CASE
           WHEN LOWER(TRIM(device_direction)) = 'in' THEN 22
           ELSE 14
       END
    """
)

# Rows still missing eb_id, matched to the link master by the device-side id
# (bio_id, an INT — collation-proof) or by the official emp code (after the
# emp_code back-fill above).
_UNRESOLVED_SQL = text(
    """
    SELECT DISTINCT b.emp_code, CAST(m.master_id AS SIGNED) AS eb_id
    FROM bio_attendance_table b
    JOIN tbl_master_bio_link_mst m
        ON m.match_type = 'E'
    LEFT JOIN hrms_ed_official_details o
        ON o.eb_id = m.master_id AND COALESCE(o.active, 1) = 1
    WHERE b.eb_id IS NULL
      AND b.emp_code IS NOT NULL AND b.emp_code <> ''
      AND (b.bio_id = m.bio_dev_id OR b.emp_code = o.emp_code)
    """
)

# dept/desig priority 1: last daily_attendance record per eb_id (read-only —
# daily_attendance is never written by this module).
_LAST_DAILY_ATT_SQL = text(
    """
    SELECT da.eb_id,
           da.worked_department_id,
           da.worked_designation_id
    FROM daily_attendance da
    INNER JOIN (
        SELECT eb_id, MAX(attendance_date) AS last_date
        FROM daily_attendance
        WHERE eb_id IN :eb_ids
          AND worked_department_id IS NOT NULL
        GROUP BY eb_id
    ) lx ON lx.eb_id = da.eb_id AND lx.last_date = da.attendance_date
    WHERE da.eb_id IN :eb_ids
      AND da.worked_department_id IS NOT NULL
    """
)

# dept/desig fallback: official details.
_OFFICIAL_SQL = text(
    """
    SELECT eb_id, sub_dept_id AS dept_id, designation_id AS desig_id
    FROM hrms_ed_official_details
    WHERE eb_id IN :eb_ids
      AND active = 1
    """
)

_RESOLVE_UPDATE_SQL = text(
    """
    UPDATE bio_attendance_table
       SET eb_id    = :eb_id,
           dept_id  = :dept_id,
           desig_id = :desig_id
     WHERE emp_code = :emp_code
       AND eb_id IS NULL
    """
)

_DELETE_BASIC_DAY_SQL = text(
    "DELETE FROM bio_attendance_basic WHERE tran_date = :tran_date"
)
_DELETE_PROCESS_DAY_SQL = text(
    "DELETE FROM bio_attendance_process WHERE attendance_date = :tran_date"
)

_INSERT_SPELL_ROW_SQL = text(
    """
    INSERT INTO bio_attendance_process
        (eb_id, bio_id, dept_id, desig_id,
         attendance_date, spell_name, attendance_type,
         attendance_source, check_in, check_out,
         Time_duration, Working_hours, Ot_hours,
         spell_start_time, spell_end_time, spell_hours, processed)
    VALUES
        (:eb_id, :bio_id, :dept_id, :desig_id,
         :tran_date, :spell_name, :attendance_type,
         'BIO', :check_in, :check_out,
         :time_duration, :working_hours, :ot_hours,
         :spell_start, :spell_end, :spell_hours, 1)
    """
)

_INSERT_BASIC_SQL = text(
    """
    INSERT INTO bio_attendance_basic
        (eb_id, emp_code, bio_id, dept_id, desig_id, tran_date, shift,
         sched_in_time, sched_out_time, actual_in, actual_out,
         work_dur_minutes, ot_minutes, break_minutes, total_dur_minutes,
         late_by_minutes, early_going_minutes, status, punch_records)
    VALUES
        (:eb_id, :emp_code, :bio_id, :dept_id, :desig_id, :tran_date, :shift,
         :sched_in_time, :sched_out_time, :actual_in, :actual_out,
         :work_dur_minutes, :ot_minutes, :break_minutes, :total_dur_minutes,
         :late_by_minutes, :early_going_minutes, :status, :punch_records)
    """
)

# Window fetch: every linked punch across [tran_date-14 .. tran_date+1] so the
# resolver can group a night shift (evening IN + next-morning OUT) and anchor
# session segmentation for continuous workers. Only the session that STARTS on
# tran_date is written; the extra days are context.
_FETCH_WINDOW_SQL = text(
    """
    SELECT b.eb_id, b.emp_code, b.bio_att_log_id, b.dept_id, b.desig_id,
           b.log_date, b.device_direction, b.device_id
    FROM bio_attendance_table b
    WHERE b.eb_id     IS NOT NULL
      AND b.dept_id   IS NOT NULL
      AND b.desig_id  IS NOT NULL
      AND b.device_id IS NOT NULL
      AND DATE(b.log_date) BETWEEN DATE_SUB(:tran_date, INTERVAL 14 DAY)
                               AND DATE_ADD(:tran_date, INTERVAL 1 DAY)
    ORDER BY b.eb_id, b.log_date
    """
)

_SPELL_TIMES: dict[str, tuple[str, str]] = {
    "A":  ("06:00:00", "14:00:00"),
    "A1": ("06:00:00", "14:00:00"),
    "A2": ("09:00:00", "13:00:00"),
    "GS": ("10:00:00", "18:00:00"),
    "B":  ("14:00:00", "22:00:00"),
    "B1": ("14:00:00", "22:00:00"),
    "B2": ("17:00:00", "22:00:00"),
    "C":  ("22:00:00", "06:00:00"),
    "C1": ("22:00:00", "06:00:00"),
    "C2": ("00:00:00", "06:00:00"),
    "O":  ("00:00:00", "00:00:00"),
}


def prepare_punches(db: Session) -> dict:
    """Table-global pre-steps on bio_attendance_table (date-independent):

      1. emp_code       <- linked employee's official emp_code (by bio_id)
      2. bio_att_log_id <- generated 500000+ sequence where blank (NULL or 0)
      3. device_id      <- 22 when device_direction='in', else 14

    Returns the three row counts.
    """
    empcode_res = db.execute(_EMPCODE_BACKFILL_SQL)
    emp_code_updated = int(empcode_res.rowcount or 0)
    db.commit()

    logid_max_row = db.execute(
        text("SELECT MAX(bio_att_log_id) AS max_id FROM bio_attendance_table")
    ).mappings().first()
    cur_max = (
        int(logid_max_row["max_id"])
        if logid_max_row and logid_max_row["max_id"] is not None
        else 0
    )
    logid_base = 500000 if cur_max < 500000 else cur_max + 1
    db.execute(_LOGID_INIT_SQL, {"base": logid_base})
    logid_res = db.execute(_LOGID_FILL_SQL)
    log_id_filled = int(logid_res.rowcount or 0)
    db.commit()

    device_res = db.execute(_DEVICE_ID_SQL)
    device_id_updated = int(device_res.rowcount or 0)
    db.commit()

    return {
        "emp_code_updated": emp_code_updated,
        "bio_att_log_id_filled": log_id_filled,
        "device_id_updated": device_id_updated,
    }


def _resolve_eb_ids(db: Session) -> dict:
    """Resolve eb_id / dept_id / desig_id for punches where eb_id IS NULL:
    emp_code -> eb_id via the link master, dept/desig from the last
    daily_attendance record (fallback: hrms_ed_official_details). Commits.
    """
    result: dict = {
        "resolved": 0,
        "updated": 0,
        "from_daily_attendance": 0,
        "fallback_official": 0,
        "no_source": 0,
    }

    unresolved_rows = db.execute(_UNRESOLVED_SQL).fetchall()
    if not unresolved_rows:
        return result

    emp_eb_map: dict[str, int] = {
        str(r.emp_code): int(r.eb_id)
        for r in unresolved_rows
        if r.eb_id is not None
    }
    if not emp_eb_map:
        return result

    unique_eb_ids = list(set(emp_eb_map.values()))

    da_map: dict[int, tuple] = {}
    try:
        da_rows = db.execute(
            _LAST_DAILY_ATT_SQL, {"eb_ids": tuple(unique_eb_ids)}
        ).fetchall()
        da_map = {
            int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
            for r in da_rows
            if r.worked_department_id is not None
        }
    except Exception:
        # daily_attendance absent/odd schema on this tenant — official fallback
        # below covers everyone.
        db.rollback()

    missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
    official_map: dict[int, tuple] = {}
    if missing_eb_ids:
        off_rows = db.execute(
            _OFFICIAL_SQL, {"eb_ids": tuple(missing_eb_ids)}
        ).fetchall()
        official_map = {int(r.eb_id): (r.dept_id, r.desig_id) for r in off_rows}

    updated = 0
    fallback_official = 0
    no_source = 0
    for emp_code, eb_id in emp_eb_map.items():
        if eb_id in da_map:
            dept_id, desig_id = da_map[eb_id]
        elif eb_id in official_map:
            dept_id, desig_id = official_map[eb_id]
            fallback_official += 1
        else:
            no_source += 1
            continue

        res = db.execute(
            _RESOLVE_UPDATE_SQL,
            {
                "eb_id": eb_id,
                "dept_id": dept_id,
                "desig_id": desig_id,
                "emp_code": emp_code,
            },
        )
        updated += int(res.rowcount or 0)

    db.commit()
    result.update({
        "resolved": len(emp_eb_map),
        "updated": updated,
        "from_daily_attendance": len(da_map),
        "fallback_official": fallback_official,
        "no_source": no_source,
    })
    return result


def _mysql_dow(d: date) -> int:
    """Weekday in MySQL DAYOFWEEK() convention: 1=Sun..7=Sat."""
    return d.isoweekday() % 7 + 1


def _get_off_day_map(db: Session) -> dict[int, int]:
    """Per-employee weekly-off weekday (1=Sun..7=Sat) from
    hrms_ed_official_details.off_day. Tenants without that column (schema
    drift, e.g. dev3) get an empty map — no weekly offs."""
    try:
        rows = db.execute(text(
            """
            SELECT eb_id, off_day FROM hrms_ed_official_details
            WHERE COALESCE(active, 1) = 1 AND off_day IS NOT NULL
            """
        )).fetchall()
        return {int(r.eb_id): int(r.off_day) for r in rows}
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return {}


def bprocess_core(db: Session, tran_date: str) -> dict:
    """Resolve punches, then rebuild bio_attendance_basic (+ one spell row in
    bio_attendance_process) for tran_date using the pure essl_resolver. The
    weekly-off flag is PER EMPLOYEE (hrms_ed_official_details.off_day vs the
    date's weekday). Idempotent: deletes the date's rows in both tables first.
    """
    resolve_result = _resolve_eb_ids(db)
    dow = _mysql_dow(date.fromisoformat(tran_date))
    off_map = _get_off_day_map(db)

    db.execute(_DELETE_BASIC_DAY_SQL, {"tran_date": tran_date})
    db.execute(_DELETE_PROCESS_DAY_SQL, {"tran_date": tran_date})

    rows = db.execute(_FETCH_WINDOW_SQL, {"tran_date": tran_date}).fetchall()

    punches_by_emp: dict[int, list[Punch]] = {}
    identity: dict[int, dict] = {}
    for r in rows:
        m = r._mapping
        eb = m["eb_id"]
        punches_by_emp.setdefault(eb, []).append(
            Punch(
                log_date=m["log_date"],
                raw_dir=(m["device_direction"] or ""),
                device_id=m["device_id"],
            )
        )
        identity.setdefault(eb, {
            "emp_code": m["emp_code"],
            "bio_id": m["bio_att_log_id"],
            "dept_id": m["dept_id"],
            "desig_id": m["desig_id"],
        })

    target = date.fromisoformat(tran_date) if isinstance(tran_date, str) else tran_date

    inserted = 0
    for eb, plist in punches_by_emp.items():
        is_off_day = off_map.get(int(eb)) == dow
        day = None
        for sess in segment_sessions(plist):
            if sess[0].log_date.date() != target:
                continue
            day = resolve_session(sess, is_off_day=is_off_day)
            if day is not None:
                break
        if day is None:
            continue

        ident = identity.get(eb, {})
        bio_id = ident.get("bio_id")
        dept_id = ident.get("dept_id")
        desig_id = ident.get("desig_id")
        spell_start, spell_end = _SPELL_TIMES.get(day.shift, ("00:00:00", "00:00:00"))
        working_hours = round(day.work_minutes / 60.0, 2)
        ot_hours = round(day.ot_minutes / 60.0, 2)
        time_duration = round(working_hours + ot_hours, 2)

        db.execute(
            _INSERT_SPELL_ROW_SQL,
            {
                "eb_id":           int(eb),
                "bio_id":          int(bio_id) if bio_id is not None else None,
                "dept_id":         int(dept_id) if dept_id is not None else None,
                "desig_id":        int(desig_id) if desig_id is not None else None,
                "tran_date":       tran_date,
                "spell_name":      day.shift,
                "attendance_type": "O" if is_off_day else "R",
                "check_in":        day.actual_in,
                "check_out":       day.actual_out,
                "time_duration":   time_duration,
                "working_hours":   working_hours,
                "ot_hours":        ot_hours,
                "spell_start":     spell_start,
                "spell_end":       spell_end,
                "spell_hours":     SPELL_HOURS,
            },
        )

        db.execute(
            _INSERT_BASIC_SQL,
            {
                "eb_id":               int(eb),
                "emp_code":            ident.get("emp_code"),
                "bio_id":              int(bio_id) if bio_id is not None else None,
                "dept_id":             int(dept_id) if dept_id is not None else None,
                "desig_id":            int(desig_id) if desig_id is not None else None,
                "tran_date":           tran_date,
                "shift":               day.shift,
                "sched_in_time":       spell_start,
                "sched_out_time":      spell_end,
                "actual_in":           day.actual_in,
                "actual_out":          day.actual_out,
                "work_dur_minutes":    day.work_minutes,
                "ot_minutes":          day.ot_minutes,
                "break_minutes":       day.break_minutes,
                "total_dur_minutes":   day.total_minutes,
                "late_by_minutes":     day.late_minutes,
                "early_going_minutes": day.early_going_minutes,
                "status":              day.status,
                "punch_records":       day.punch_records,
            },
        )
        inserted += 1

    db.commit()
    return {
        "resolve": resolve_result,
        "inserted": inserted,
    }


def _minutes_bucket(mins: int) -> int:
    """Bucket used ONLY to pick the spell label (>420 -> 8, [180, 420] -> 4,
    (0, 180) -> 0, <=0 -> -1). Stored hours are the ACTUAL hours, not these."""
    if mins <= 0:
        return -1
    if mins > 420:
        return 8
    if mins >= 180:
        return 4
    return 0


def _spell_label(shift: str, kind: str, bucket: int, intime_h: float) -> str:
    """Spell label for the projected R/O rows. `kind` is 'R' or 'O'."""
    if shift == "A":
        if kind == "R":
            if bucket == 8:
                return "A"
            if bucket == 4:
                return "A1" if intime_h < 9 else "A2"
            return "A"
        if bucket == 8:
            return "B"
        if bucket == 4:
            return "B1"
        return "O"
    if shift == "B":
        if kind == "R":
            if bucket == 8:
                return "B"
            if bucket == 4:
                return "B1" if intime_h < 17 else "B2"
            return "B"
        if bucket == 8:
            return "C"
        if bucket == 4:
            return "C1"
        return "O"
    if shift == "C":
        if kind == "R":
            if bucket == 8:
                return "C"
            if bucket == 4:
                if 0 <= intime_h <= 4:
                    return "C2"
                return "C1"
            return "C"
        if bucket == 8 and intime_h > 21:
            return "A"
        if bucket == 4:
            if intime_h < 19:
                return "B2"
            if intime_h > 21:
                return "A1"
        return "O"
    return shift if kind == "R" else "O"


def b_atten_core(db: Session, tran_date: str) -> dict:
    """Project rows from bio_attendance_basic into bio_attendance_process for
    tran_date: one R row for working hours, one O row for overtime, each
    storing the ACTUAL hours (minutes/60, 2dp); the 8/4 bucket is used only to
    pick the spell label. The weekly-off flag is per employee (official
    details off_day). Idempotent: wipes the date's process rows first.
    """
    dow = _mysql_dow(date.fromisoformat(tran_date))
    off_map = _get_off_day_map(db)

    db.execute(_DELETE_PROCESS_DAY_SQL, {"tran_date": tran_date})

    basic_rows = db.execute(
        text(
            """
            SELECT eb_id, bio_id, dept_id, desig_id, shift,
                   actual_in, actual_out,
                   work_dur_minutes, ot_minutes, total_dur_minutes
            FROM bio_attendance_basic
            WHERE tran_date = :tran_date
            """
        ),
        {"tran_date": tran_date},
    ).fetchall()

    inserted = 0
    for r in basic_rows:
        m = r._mapping
        is_off_day = off_map.get(int(m["eb_id"])) == dow
        shift = m["shift"]
        actual_in = m["actual_in"]
        intime_h = (
            actual_in.hour + actual_in.minute / 60.0 + actual_in.second / 3600.0
            if isinstance(actual_in, datetime) else 0.0
        )
        work_min = int(m["work_dur_minutes"] or 0)
        ot_min = int(m["ot_minutes"] or 0)

        common_base = {
            "eb_id":     int(m["eb_id"]),
            "bio_id":    int(m["bio_id"]) if m["bio_id"] is not None else None,
            "dept_id":   int(m["dept_id"]) if m["dept_id"] is not None else None,
            "desig_id":  int(m["desig_id"]) if m["desig_id"] is not None else None,
            "tran_date": tran_date,
            "check_in":   actual_in,
            "check_out":  m["actual_out"],
            "spell_hours": SPELL_HOURS,
        }

        if work_min > 0:
            work_hrs = round(work_min / 60.0, 2)
            spell = _spell_label(shift, "R", _minutes_bucket(work_min), intime_h)
            spell_start, spell_end = _SPELL_TIMES.get(spell, ("00:00:00", "00:00:00"))
            db.execute(
                _INSERT_SPELL_ROW_SQL,
                {
                    **common_base,
                    "spell_name":      spell,
                    "attendance_type": "O" if is_off_day else "R",
                    "time_duration":   work_hrs,
                    "working_hours":   work_hrs,
                    "ot_hours":        0,
                    "spell_start":     spell_start,
                    "spell_end":       spell_end,
                },
            )
            inserted += 1

        if ot_min > 0:
            ot_hrs = round(ot_min / 60.0, 2)
            spell = _spell_label(shift, "O", _minutes_bucket(ot_min), intime_h)
            spell_start, spell_end = _SPELL_TIMES.get(spell, ("00:00:00", "00:00:00"))
            db.execute(
                _INSERT_SPELL_ROW_SQL,
                {
                    **common_base,
                    "spell_name":      spell,
                    "attendance_type": "O",
                    "time_duration":   ot_hrs,
                    "working_hours":   ot_hrs,
                    "ot_hours":        0,
                    "spell_start":     spell_start,
                    "spell_end":       spell_end,
                },
            )
            inserted += 1

    db.commit()
    return {
        "basic_rows": len(basic_rows),
        "inserted": inserted,
    }


def run_process_for_date(db: Session, tran_date: str, *, run_prepare: bool = True) -> dict:
    """Full bio-data chain for one date: prepare (optional, table-global) →
    bprocess (bio_attendance_basic) → b_atten (bio_attendance_process).
    daily_attendance is never touched.
    """
    ensure_bio_tables(db)
    prep = prepare_punches(db) if run_prepare else {}
    bprocess = bprocess_core(db, tran_date)
    b_atten = b_atten_core(db, tran_date)
    return {
        "tran_date": tran_date,
        "prepare": prep,
        "resolve": bprocess["resolve"],
        "basic_inserted": bprocess["inserted"],
        "process_inserted": b_atten["inserted"],
    }


# ─── Bio data process (tab 2) endpoints ─────────────────────────────


def _validated_date(name: str, raw) -> date:
    if not raw:
        raise HTTPException(status_code=400, detail=f"{name} is required")
    try:
        return date.fromisoformat(str(raw))
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid {name} {raw!r}, expected YYYY-MM-DD",
        )


# ponytail: 92-day cap keeps an accidental year-long range from tying up the
# request; raise it if a real backfill ever needs more.
_MAX_RANGE_DAYS = 92


def _parse_period(body: dict, qp) -> tuple[date, date]:
    """Parse from_date/to_date (to_date defaults to from_date; legacy
    tran_date accepted as from_date)."""
    from_raw = body.get("from_date") or qp.get("from_date") or body.get("tran_date") or qp.get("tran_date")
    from_date = _validated_date("from_date", from_raw)
    to_raw = body.get("to_date") or qp.get("to_date")
    to_date = _validated_date("to_date", to_raw) if to_raw else from_date
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date is before from_date")
    if (to_date - from_date).days + 1 > _MAX_RANGE_DAYS:
        raise HTTPException(
            status_code=400,
            detail=f"Period too long — max {_MAX_RANGE_DAYS} days per run",
        )
    return from_date, to_date


# The run is many sequential statements (minutes over a WAN link) — it MUST
# NOT run on the event loop, and two runs must not interleave their
# delete+insert passes on the same dates.
_process_run_lock = threading.Lock()


def _process_run_sync(db: Session, from_date: date, to_date: date) -> dict:
    """Blocking body of bio_att_process_run — called via run_in_threadpool."""
    ensure_bio_tables(db)
    punch_sync = sync_punch_logs(db)
    prepare = prepare_punches(db)

    processed: list[dict] = []
    failed: list[str] = []
    basic_total = 0
    process_total = 0
    d = from_date - timedelta(days=1)
    while d <= to_date:
        tran_date = d.isoformat()
        try:
            bprocess = bprocess_core(db, tran_date)
            b_atten = b_atten_core(db, tran_date)
            basic_total += bprocess["inserted"]
            process_total += b_atten["inserted"]
            processed.append({
                "tran_date": tran_date,
                "basic_inserted": bprocess["inserted"],
                "process_inserted": b_atten["inserted"],
            })
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            failed.append(tran_date)
        d += timedelta(days=1)

    return {
        "status": "ok",
        "result": {
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "punch_sync": punch_sync,
            "prepare": prepare,
            "dates_processed": len(processed),
            "basic_inserted": basic_total,
            "process_inserted": process_total,
            "failed": failed,
            "days": processed,
        },
    }


@router.post("/bio_att_process_run")
async def bio_att_process_run(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Run the bio data process for a period (from_date..to_date, to_date
    optional): pull new device punches from punch_logs, prepare them once, then
    build bio_attendance_basic + bio_attendance_process for every date in the
    range. The day BEFORE from_date is processed too, so a night shift whose
    OUT punch lands on from_date gets closed.
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        from_date, to_date = _parse_period(body, request.query_params)

        if not _process_run_lock.acquire(blocking=False):
            raise HTTPException(
                status_code=409,
                detail="A bio process run is already in progress — wait for it to finish",
            )
        try:
            return await run_in_threadpool(_process_run_sync, db, from_date, to_date)
        finally:
            _process_run_lock.release()
    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


_BASIC_LIST_SQL = text(
    """
    SELECT b.id, b.eb_id, b.emp_code, b.dept_id, b.desig_id, b.tran_date,
           b.shift, b.sched_in_time, b.sched_out_time, b.actual_in, b.actual_out,
           b.work_dur_minutes, b.ot_minutes, b.break_minutes, b.total_dur_minutes,
           b.late_by_minutes, b.early_going_minutes, b.status, b.punch_records,
           CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) AS employee_name
    FROM bio_attendance_basic b
    LEFT JOIN hrms_ed_personal_details p ON p.eb_id = b.eb_id
    WHERE b.tran_date BETWEEN :from_date AND :to_date
      AND (:search IS NULL
           OR b.emp_code LIKE :search
           OR CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) LIKE :search)
    ORDER BY b.tran_date, b.emp_code, b.id
    """
)

_PROCESS_LIST_SQL = text(
    """
    SELECT d.bio_att_process_id, d.eb_id, d.dept_id, d.desig_id,
           d.attendance_date, d.spell_name, d.attendance_type,
           d.check_in, d.check_out, d.Time_duration, d.Working_hours,
           d.Ot_hours, d.spell_start_time, d.spell_end_time,
           o.emp_code,
           CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) AS employee_name
    FROM bio_attendance_process d
    LEFT JOIN hrms_ed_personal_details p ON p.eb_id = d.eb_id
    LEFT JOIN hrms_ed_official_details o ON o.eb_id = d.eb_id AND COALESCE(o.active, 1) = 1
    WHERE d.attendance_date BETWEEN :from_date AND :to_date
      AND (:search IS NULL
           OR o.emp_code LIKE :search
           OR CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) LIKE :search)
    ORDER BY d.attendance_date, o.emp_code, d.bio_att_process_id
    """
)


def _paginated_day_list(db: Session, stmt, request: Request) -> dict:
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    ensure_bio_tables(db)
    from_date, to_date = _parse_period({}, request.query_params)
    search = request.query_params.get("search")
    page = int(request.query_params.get("page", 1))
    limit = int(request.query_params.get("limit", 10))

    rows = db.execute(
        stmt,
        {
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "search": f"%{search}%" if search else None,
        },
    ).fetchall()
    all_data = [dict(r._mapping) for r in rows]
    start = (page - 1) * limit
    return {
        "data": all_data[start : start + limit],
        "total": len(all_data),
        "page": page,
        "limit": limit,
    }


@router.get("/bio_att_basic_list")
async def bio_att_basic_list(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated bio_attendance_basic rows for a period (from_date..to_date)."""
    try:
        return _paginated_day_list(db, _BASIC_LIST_SQL, request)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bio_att_process_list")
async def bio_att_process_list(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated bio_attendance_process rows for a period (from_date..to_date)."""
    try:
        return _paginated_day_list(db, _PROCESS_LIST_SQL, request)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Day-wise report (tab 3) ────────────────────────────────────────

# Presence classification per day (hours = bio_attendance_basic total, i.e.
# work + OT):  > 5h -> P,  2..5h -> 1/2P,  else A.  On the EMPLOYEE'S weekly
# off (hrms_ed_official_details.off_day, 1=Sun..7=Sat): WOP when present
# (>= 2h) else WO — downgraded to WOA (sandwich rule) when the day before AND
# the day after the off day are both absent.
_P_MIN_HOURS = 5.0
_HALF_MIN_HOURS = 2.0


def _day_status(hours: float, is_off: bool) -> str:
    if is_off:
        return "WOP" if hours >= _HALF_MIN_HOURS else "WO"
    if hours > _P_MIN_HOURS:
        return "P"
    if hours >= _HALF_MIN_HOURS:
        return "1/2P"
    return "A"


def _bio_linked_employees(
    db: Session, branch_ids: list[int], search: str | None
):
    """Bio-linked employees (the bio-attendance population), branch/search
    scoped. Shared by the report builders."""
    params: dict = {"search": f"%{search}%" if search else None}
    branch_filter_sql = ""
    if branch_ids:
        placeholders = ",".join(f":b{i}" for i in range(len(branch_ids)))
        branch_filter_sql = f"AND o.branch_id IN ({placeholders})"
        for i, bid in enumerate(branch_ids):
            params[f"b{i}"] = bid

    return db.execute(
        text(f"""
            SELECT DISTINCT l.master_id AS eb_id, o.emp_code,
                   CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) AS employee_name
            FROM tbl_master_bio_link_mst l
            JOIN hrms_ed_official_details o ON o.eb_id = l.master_id AND COALESCE(o.active, 1) = 1
            LEFT JOIN hrms_ed_personal_details p ON p.eb_id = l.master_id
            WHERE l.match_type = 'E'
              {branch_filter_sql}
              AND (:search IS NULL
                   OR o.emp_code LIKE :search
                   OR CONCAT(p.first_name, ' ', COALESCE(p.middle_name, ''), ' ', COALESCE(p.last_name, '')) LIKE :search)
            ORDER BY o.emp_code
        """),
        params,
    ).fetchall()


def _get_company_name(db: Session, co_id) -> str | None:
    try:
        row = db.execute(
            text("SELECT co_name FROM co_mst WHERE co_id = :co_id"),
            {"co_id": int(co_id)},
        ).fetchone()
        return row.co_name if row else None
    except Exception:
        return None


def _build_daywise_matrix(
    db: Session,
    branch_ids: list[int],
    from_date: date,
    to_date: date,
    search: str | None,
) -> tuple[list[dict], list[dict]]:
    """Build the day-wise matrix: (days meta, one dict per employee with
    per-day statuses and totals). Shared by the JSON and Excel endpoints."""
    employees = _bio_linked_employees(db, branch_ids, search)

    # Worked hours per (eb_id, date) from the basic table. Padded ±1 day so
    # the WOA sandwich rule can see the neighbours of a boundary off day.
    hours_map: dict[tuple[int, str], float] = {}
    for r in db.execute(
        text("""
            SELECT eb_id, tran_date, total_dur_minutes
            FROM bio_attendance_basic
            WHERE tran_date BETWEEN :from_date AND :to_date
        """),
        {
            "from_date": (from_date - timedelta(days=1)).isoformat(),
            "to_date": (to_date + timedelta(days=1)).isoformat(),
        },
    ).fetchall():
        hours_map[(int(r.eb_id), str(r.tran_date))] = (
            int(r.total_dur_minutes or 0) / 60.0
        )

    # Per-employee weekly-off weekday (official details, 1=Sun..7=Sat).
    off_map = _get_off_day_map(db)

    days: list[dict] = []
    d = from_date
    while d <= to_date:
        days.append({
            "date": d.isoformat(),
            "day": d.day,
            "dow": _mysql_dow(d),
        })
        d += timedelta(days=1)

    data = []
    for emp in employees:
        eb_id = int(emp.eb_id)
        emp_off = off_map.get(eb_id)
        statuses: list[str] = []
        for day in days:
            hrs = hours_map.get((eb_id, day["date"]), 0.0)
            statuses.append(_day_status(hrs, emp_off == day["dow"]))

        # Sandwich rule: a WO whose calendar neighbours (day before AND
        # after, even outside the period thanks to the ±1-day padding) are
        # both absent (< 2h) becomes WOA.
        for i, day in enumerate(days):
            if statuses[i] != "WO":
                continue
            day_d = date.fromisoformat(day["date"])
            prev_hrs = hours_map.get((eb_id, (day_d - timedelta(days=1)).isoformat()), 0.0)
            next_hrs = hours_map.get((eb_id, (day_d + timedelta(days=1)).isoformat()), 0.0)
            if prev_hrs < _HALF_MIN_HOURS and next_hrs < _HALF_MIN_HOURS:
                statuses[i] = "WOA"

        totals = {"P": 0, "1/2P": 0, "WOP": 0, "WOA": 0, "A": 0}
        for status in statuses:
            if status in totals:
                totals[status] += 1
        data.append({
            "eb_id": eb_id,
            "emp_code": emp.emp_code,
            "employee_name": emp.employee_name,
            "days": statuses,
            "total_present": totals["P"],
            "total_half": totals["1/2P"],
            "total_wop": totals["WOP"],
            "total_woa": totals["WOA"],
            "total_absent": totals["A"],
        })

    return days, data


@router.get("/bio_att_daywise_report")
async def bio_att_daywise_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Day-wise attendance matrix for a period: one row per bio-linked
    employee, one status cell per day (P / 1/2P / A / WO / WOP / WOA) plus
    totals.
    """
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        ensure_bio_tables(db)
        from_date, to_date = _parse_period({}, request.query_params)
        search = request.query_params.get("search")
        branch_ids = _parse_branch_ids(request.query_params.get("branch_id"))

        days, data = _build_daywise_matrix(db, branch_ids, from_date, to_date, search)
        return {"days": days, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bio_att_daywise_report_excel")
async def bio_att_daywise_report_excel(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Day-wise attendance register as a styled .xlsx: company name +
    "Attendance Register for the Period ..." title block (no borders), then a
    bordered table with day/month column headers.
    """
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        ensure_bio_tables(db)
        from_date, to_date = _parse_period({}, request.query_params)
        search = request.query_params.get("search")
        branch_ids = _parse_branch_ids(request.query_params.get("branch_id"))

        days, data = _build_daywise_matrix(db, branch_ids, from_date, to_date, search)

        co_name = _get_company_name(db, co_id)

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Register"

        headers = (
            ["Emp Code", "Name"]
            + [f"{d['day']:02d}/{int(d['date'][5:7]):02d}" for d in days]
            + ["P", "1/2P", "WOP", "WOA", "A"]
        )
        n_cols = len(headers)

        # Title block — no borders.
        center = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=1, value=co_name or "").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.cell(row=1, column=1).alignment = center
        ws.cell(
            row=2, column=1,
            value=f"Attendance Register for the Period {from_date.isoformat()} To {to_date.isoformat()}",
        ).font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.cell(row=2, column=1).alignment = center

        # Table — header styled like the on-screen grid, thin borders on every
        # row and column of the table.
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="3EA6DA")
        header_font = Font(bold=True, color="FFFFFF")

        header_row = 4
        for c, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=c, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center

        for r_idx, emp in enumerate(data, start=header_row + 1):
            values = (
                [emp["emp_code"], emp["employee_name"]]
                + emp["days"]
                + [emp["total_present"], emp["total_half"], emp["total_wop"],
                   emp["total_woa"], emp["total_absent"]]
            )
            for c_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if c_idx > 2:
                    cell.alignment = center

        from openpyxl.utils import get_column_letter

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 28
        for c in range(3, n_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 7

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"AttendanceRegister_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
        return StreamingResponse(
            iter([bio.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Monthly In-Out report (tab 4) ──────────────────────────────────
# Mirrors the legacy "Monthly InOut Report" sheet: one row per employee, one
# column per day holding "IN\nOUT" ("00:00" when there is no out punch), WO on
# the employee's weekly off, A when absent. IN/OUT come straight from the RAW
# punch log (first/last punch of the calendar day in bio_attendance_table) so
# the report works even before the process chain has run. Totals: Total Days
# (days in period minus absences), Late Count (late days per the processed
# basic table, when available), Appr. Days (= Total Days).


def _day_label(d: date) -> str:
    """Column label like "01-Jun", annotated "(Sun)" / "(2nd Sat)" / "(4th Sat)"."""
    label = d.strftime("%d-%b")
    if d.isoweekday() == 7:
        return f"{label} (Sun)"
    if d.isoweekday() == 6:
        nth = (d.day - 1) // 7 + 1
        if nth == 2:
            return f"{label} (2nd Sat)"
        if nth == 4:
            return f"{label} (4th Sat)"
    return label


def _fmt_time(value) -> str:
    if value is None:
        return "00:00"
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    return str(value)


def _build_inout_matrix(
    db: Session,
    branch_ids: list[int],
    from_date: date,
    to_date: date,
    search: str | None,
) -> tuple[list[dict], list[dict]]:
    """(days meta, employee rows) for the Monthly In-Out report."""
    employees = _bio_linked_employees(db, branch_ids, search)

    # First/last RAW punch per (employee, calendar day), matched to the link
    # master by resolved eb_id OR device-side bio id.
    punch_map: dict[tuple[int, str], tuple] = {}
    for r in db.execute(
        text("""
            SELECT m.master_id AS eb_id, DATE(b.log_date) AS d,
                   MIN(b.log_date) AS first_p, MAX(b.log_date) AS last_p
            FROM bio_attendance_table b
            JOIN tbl_master_bio_link_mst m
                ON m.match_type = 'E'
               AND (b.eb_id = m.master_id OR b.bio_id = m.bio_dev_id)
            WHERE b.log_date IS NOT NULL
              AND DATE(b.log_date) BETWEEN :from_date AND :to_date
            GROUP BY m.master_id, DATE(b.log_date)
        """),
        {"from_date": from_date.isoformat(), "to_date": to_date.isoformat()},
    ).fetchall():
        punch_map[(int(r.eb_id), str(r.d))] = (r.first_p, r.last_p)

    # Late-day counts from the processed basic table (0 when not processed).
    late_map: dict[int, int] = {}
    for r in db.execute(
        text("""
            SELECT eb_id, COUNT(*) AS c
            FROM bio_attendance_basic
            WHERE tran_date BETWEEN :from_date AND :to_date
              AND late_by_minutes > 0
            GROUP BY eb_id
        """),
        {"from_date": from_date.isoformat(), "to_date": to_date.isoformat()},
    ).fetchall():
        late_map[int(r.eb_id)] = int(r.c)

    off_map = _get_off_day_map(db)

    days: list[dict] = []
    d = from_date
    while d <= to_date:
        days.append({"date": d.isoformat(), "label": _day_label(d), "dow": _mysql_dow(d)})
        d += timedelta(days=1)

    data = []
    for emp in employees:
        eb_id = int(emp.eb_id)
        emp_off = off_map.get(eb_id)
        cells: list[str] = []
        absent = 0
        for day in days:
            punch = punch_map.get((eb_id, day["date"]))
            if punch is not None:
                first_p, last_p = punch
                # OUT = last punch of the day; a lone punch has no out (00:00).
                out = last_p if last_p != first_p else None
                cells.append(f"{_fmt_time(first_p)}\n{_fmt_time(out)}")
            elif emp_off == day["dow"]:
                cells.append("WO")
            else:
                cells.append("A")
                absent += 1
        total_days = len(days) - absent
        data.append({
            "eb_id": eb_id,
            "emp_code": emp.emp_code,
            "employee_name": emp.employee_name,
            "days": cells,
            "total_days": total_days,
            "late_count": late_map.get(eb_id, 0),
            "appr_days": total_days,
        })

    return days, data


@router.get("/bio_att_inout_report")
async def bio_att_inout_report(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Monthly In-Out matrix for a period: per-day IN/OUT punch times
    (WO / A markers) plus Total Days, Late Count and Appr. Days."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        ensure_bio_tables(db)
        # Raw-log report: pull any punches that arrived since the last run so
        # today's in/out always shows without waiting for a process run.
        sync_punch_logs(db)
        from_date, to_date = _parse_period({}, request.query_params)
        search = request.query_params.get("search")
        branch_ids = _parse_branch_ids(request.query_params.get("branch_id"))

        days, data = _build_inout_matrix(db, branch_ids, from_date, to_date, search)
        return {"days": days, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bio_att_inout_report_excel")
async def bio_att_inout_report_excel(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Monthly In-Out report as a styled .xlsx: company + "Monthly InOut
    Report" + period title block (no borders), then a bordered table with
    IN/OUT stacked in each day cell."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        ensure_bio_tables(db)
        from_date, to_date = _parse_period({}, request.query_params)
        search = request.query_params.get("search")
        branch_ids = _parse_branch_ids(request.query_params.get("branch_id"))

        days, data = _build_inout_matrix(db, branch_ids, from_date, to_date, search)
        co_name = _get_company_name(db, co_id)

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly InOut"

        headers = (
            ["Emp Code", "EMP Name"]
            + [d["label"] for d in days]
            + ["Total Days", "Late Count", "Appr. Days"]
        )
        n_cols = len(headers)

        center = Alignment(horizontal="center", vertical="center")
        wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Title block — no borders (like the legacy sheet).
        ws.cell(row=1, column=1, value=co_name or "").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.cell(row=1, column=1).alignment = center
        ws.cell(row=2, column=1, value="Monthly InOut Report").font = Font(bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.cell(row=2, column=1).alignment = center
        ws.cell(
            row=3, column=1,
            value=f"{from_date.strftime('%d-%B-%Y')} to {to_date.strftime('%d-%B-%Y')}",
        ).font = Font(bold=True)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
        ws.cell(row=3, column=1).alignment = center

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="3EA6DA")
        header_font = Font(bold=True, color="FFFFFF")

        header_row = 5
        for c, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=c, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = wrap_center

        for r_idx, emp in enumerate(data, start=header_row + 1):
            values = (
                [emp["emp_code"], emp["employee_name"]]
                + emp["days"]
                + [emp["total_days"], emp["late_count"], emp["appr_days"]]
            )
            for c_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if c_idx > 2:
                    cell.alignment = wrap_center
            ws.row_dimensions[r_idx].height = 26

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 28
        for c in range(3, n_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 11

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"MonthlyInOut_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
        return StreamingResponse(
            iter([bio.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
