"""
Tests for the Jute Purchase Tally download endpoint.
Tests for src/juteProcurement/reports.py @ GET /api/juteReports/tally-download

The endpoint now returns a multi-sheet xlsx (Purchase + Check List). Tests
validate the wire format and the fixed-block layout per MR.
"""

from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from unittest.mock import MagicMock

from src.main import app
from src.config.db import get_tenant_db
from src.authorization.utils import get_current_user_with_refresh

client = TestClient(app)


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# Expected Purchase-sheet headers in order (73 columns). Duplicates allowed.
EXPECTED_HEADERS = [
    "Vch No.",                        # 1
    "Vch Type",                       # 2
    "Date",                           # 3
    "Supplier Inv No",                # 4
    "Supplier Inv Date",              # 5
    "Receipt Note No",                # 6
    "Receipt Note Date",              # 7
    "Order No",                       # 8
    "Order Date",                     # 9
    "Party Name",                     # 10
    "Registration Type",              # 11
    "GSTIN No",                       # 12
    "Country",                        # 13
    "State",                          # 14
    "Pincode",                        # 15
    "Address 1",                      # 16
    "Address 2",                      # 17
    "Address 3",                      # 18
    "Purchase Ledger",                # 19
    "Purchase Ledger Description",    # 20
    "Item Name",                      # 21
    "Item Description",               # 22
    "UNITS.",                         # 23
    "Tracking No",                    # 24
    "Order No",                       # 25
    "Order Due Date",                 # 26
    "Godown",                         # 27
    "Batch",                          # 28
    "Qty",                            # 29
    "Rate",                           # 30
    "Amt",                            # 31
    "Amount1",                        # 32
    "Discount Ledger",                # 33
    "Discount Ledger Description",    # 34
    "Amount1",                        # 35
    "Additional Ledger",              # 36
    "Additional Ledger Description",  # 37
    "Amount",                         # 38
    "INPUT IGST",                     # 39
    "INPUT CGST",                     # 40
    "INPUT SGST",                     # 41
    "Roundoff amt",                   # 42
    "Total",                          # 43
    "Ref: No.",                       # 44
    "Due on",                         # 45
    "e Way Bill No.",                 # 46
    "e Way Bill Date",                # 47
    "Sub Type",                       # 48
    "Doc Type",                       # 49
    "Status Of eway Bill",            # 50
    "Mode",                           # 51
    "Distance (In KM)",               # 52
    "Transporter Name",               # 53
    "Vehicle No.",                    # 54
    "Doc/Loading/RR/AirWay No.",      # 55
    "Date",                           # 56
    "Transporter ID",                 # 57
    "Consignor:",                     # 58
    "Address 1",                      # 59
    "Address 2",                      # 60
    "Pin Code",                       # 61
    "Place",                          # 62
    "State",                          # 63
    "GSTIN/UIN",                      # 64
    "Consignee",                      # 65
    "Address 1",                      # 66
    "Address 2",                      # 67
    "To Place (Destination )",        # 68
    "Pin",                            # 69
    "State",                          # 70
    "GSTIN/UIN",                      # 71
    "Narration",                      # 72
    "TALLYIMPORTSTATUS",              # 73
]


def _mock_row(mapping):
    row = MagicMock()
    row._mapping = mapping
    return row


def _mr_count_row(jute_mr_id, item_count):
    return _mock_row({"jute_mr_id": jute_mr_id, "item_count": item_count})


def _make_tally_row(**overrides):
    """Build a mock SQL row (one per MR line item) with all expected columns."""
    base = {
        "Vch No.": "ABC/FAC/JMR/25-26/00001",
        "Vch Type": "Purchase",
        "Date": "15-01-2026",
        "Supplier Inv No": "INV/2026/001",
        "Supplier Inv Date": "14-01-2026",
        "Receipt Note No": "ABC/FAC/JMR/25-26/00001",
        "Receipt Note Date": "15-01-2026",
        "Order No": "ABC/FAC/JPO/25-26/00001",
        "Order Date": "10-01-2026",
        "Party Name": "Acme Jute Traders",
        "Registration Type": "",
        "GSTIN No": "",
        "Country": "",
        "State": "",
        "Pincode": "",
        "Address 1": "",
        "Address 2": "",
        "Address 3": "",
        "Purchase Ledger": "Purchase of Raw Jute",
        "Purchase Ledger Description": "",
        "Item Name": "Raw Jute TD-5",
        "Item Description": "Raw Jute: 10 Bales",
        "UNITS.": "",
        "Tracking No": "",
        "Order No_1": "",
        "Order Due Date": "",
        "Godown": "Main Factory",
        "Batch": "ABC/FAC/JMR/25-26/00001",
        "Qty": 1500.0,
        "Rate": 45.25,
        "Amt": 67875.0,
        "Amount1": "",
        "Discount Ledger": "",
        "Discount Ledger Description": "",
        "Amount1_1": "",
        "Additional Ledger": "",
        "Additional Ledger Description": "",
        "Amount": "",
        "INPUT IGST": "",
        "INPUT CGST": "",
        "INPUT SGST": "",
        "Roundoff amt": "",
        "Total": "",
        "Ref: No.": "INV/2026/001/ABC/FAC/JMR/25-26/00001",
        "Due on": 180,
        "Amount (dup2)": "",
        "e Way Bill No.": "",
        "e Way Bill Date": "",
        "Sub Type": "",
        "Doc Type": "",
        "Status Of eway Bill": "",
        "Mode": "",
        "Distance (In KM)": "",
        "Transporter Name": "",
        "Vehicle No.": "",
        "Doc/Loading/RR/AirWay No.": "",
        "Date_1": "",
        "Transporter ID": "",
        "Consignor:": "",
        "Address 1_1": "",
        "Address 2_1": "",
        "Pin Code": "",
        "Place": "",
        "State_1": "",
        "GSTIN/UIN": "",
        "Consignee": "",
        "Address 1_2": "",
        "Address 2_2": "",
        "To Place (Destination )": "",
        "Pin": "",
        "State_2": "",
        "GSTIN/UIN_1": "",
        "Narration": (
            "Being 1500 Kg. Raw Jute Purchase From Acme Jute Traders "
            "Against Inv: INV/2026/001 Dt :15-01-2026 CH: CH/001,"
            "Dt: 14-01-2026,Ref.NO. ABC/FAC/JMR/25-26/00001 Amt Rs. 67875 /-"
        ),
        "TALLYIMPORTSTATUS": "",
        "Cumulative Amount": 67875.0,
        "tds Amount": 0,
        "noofitems": 1,
        "MR Amount": 67875.0,
        "TDS Deducted": 500,  # triggers TDS row
    }
    base.update(overrides)
    return _mock_row(base)


def _make_check_row(**overrides):
    base = {
        "company_id": 1,
        "mrdate": None,
        "mr_print_no": "ABC/FAC/JMR/25-26/00001",
        "invoice_no": "INV/2026/001",
        "invoice_date": None,
        "supp_name": "ACME JUTE TRADERS",
        "supptally": "Acme Jute Traders",
        "jute_quality": "Raw Jute TD-5",
        "qualitytally": "Raw Jute TD-5",
        "godowsntally": "Main Factory",
        "claimtally": "Claim on Gross Purchase",
        "purtally": "Purchase of Raw Jute",
    }
    base.update(overrides)
    return _mock_row(base)


class TestJuteTallyDownload:
    """Tests for GET /api/juteReports/tally-download."""

    @pytest.fixture(autouse=True)
    def setup_overrides(self):
        self._mock_session = MagicMock()
        app.dependency_overrides[get_current_user_with_refresh] = lambda: {
            "user_id": 1
        }
        app.dependency_overrides[get_tenant_db] = lambda: self._mock_session
        yield
        app.dependency_overrides.clear()

    # --- Parameter validation ---

    def test_missing_co_id_returns_400(self):
        response = client.get(
            "/api/juteReports/tally-download"
            "?date_from=2026-01-01&date_to=2026-01-31"
        )
        assert response.status_code == 400
        assert "co_id" in response.json().get("detail", "").lower()

    def test_missing_date_from_returns_400(self):
        response = client.get(
            "/api/juteReports/tally-download?co_id=1&date_to=2026-01-31"
        )
        assert response.status_code == 400
        assert "date_from" in response.json().get("detail", "").lower()

    def test_missing_date_to_returns_400(self):
        response = client.get(
            "/api/juteReports/tally-download?co_id=1&date_from=2026-01-01"
        )
        assert response.status_code == 400
        assert "date_to" in response.json().get("detail", "").lower()

    def test_bad_date_format_returns_400(self):
        response = client.get(
            "/api/juteReports/tally-download"
            "?co_id=1&date_from=01-01-2026&date_to=2026-01-31"
        )
        assert response.status_code == 400
        assert "date_from" in response.json().get("detail", "").lower()

    def test_invalid_co_id_returns_400(self):
        response = client.get(
            "/api/juteReports/tally-download"
            "?co_id=abc&date_from=2026-01-01&date_to=2026-01-31"
        )
        assert response.status_code == 400

    # --- Happy path: xlsx structure ---

    def test_happy_path_returns_xlsx_with_correct_structure(self):
        """
        2 items for one MR, TDS=500. Expect a 7-row Purchase sheet
        (1 header + 6 block rows) with item rows, TDS row, and padding.
        """
        # Sequence of execute() calls in the handler:
        #   1) items-per-MR query -> .fetchall() -> [(jute_mr_id, N), ...]
        #   2) main tally query   -> .fetchall() -> tally rows
        #   3) check list query   -> .fetchall() -> check rows
        items_per_mr_result = MagicMock()
        items_per_mr_result.fetchall.return_value = [_mr_count_row(1, 2)]

        tally_result = MagicMock()
        tally_result.fetchall.return_value = [
            _make_tally_row(
                **{
                    "Item Name": "Raw Jute TD-5",
                    "Qty": 1500.0,
                    "Amt": 67875.0,
                    "TDS Deducted": 500,
                }
            ),
            _make_tally_row(
                **{
                    "Item Name": "Raw Jute TD-6",
                    "Qty": 1000.0,
                    "Amt": 45000.0,
                    "TDS Deducted": 500,  # invariant within an MR group
                    "Narration": "",       # narration only on first item row
                }
            ),
        ]

        check_result = MagicMock()
        check_result.fetchall.return_value = [
            _make_check_row(jute_quality="Raw Jute TD-5"),
            _make_check_row(jute_quality="Raw Jute TD-6"),
        ]

        self._mock_session.execute.side_effect = [
            items_per_mr_result,
            tally_result,
            check_result,
        ]

        response = client.get(
            "/api/juteReports/tally-download"
            "?co_id=1&branch_id=1"
            "&date_from=2026-01-01&date_to=2026-01-31"
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
        assert ".xlsx" in response.headers.get("content-disposition", "")

        wb = openpyxl.load_workbook(BytesIO(response.content))
        assert wb.sheetnames == ["Purchase", "Check List"]

        ws = wb["Purchase"]
        # 73 headers at row 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, 74)]
        assert headers == EXPECTED_HEADERS

        # Exactly 7 total rows (1 header + 6 block rows for one MR w/ 2 items).
        assert ws.max_row == 7

        # Row 2 — first item row: shared + item + first-only
        assert ws.cell(row=2, column=2).value == "Purchase"
        assert ws.cell(row=2, column=19).value == "Purchase of Raw Jute"
        assert ws.cell(row=2, column=21).value == "Raw Jute TD-5"
        assert ws.cell(row=2, column=29).value == 1500.0
        narration = ws.cell(row=2, column=72).value
        assert narration is not None and narration != ""

        # Receipt Note No (col 6) and Date (col 7) mirror Vch No. (col 1)
        # and Date (col 3) on every block row — shared across the MR group.
        for r in range(2, 8):
            assert ws.cell(row=r, column=6).value == ws.cell(row=r, column=1).value
            assert ws.cell(row=r, column=7).value == ws.cell(row=r, column=3).value
        assert ws.cell(row=2, column=6).value == "ABC/FAC/JMR/25-26/00001"
        assert ws.cell(row=2, column=7).value == "15-01-2026"

        # Row 3 — second item row: shared + per-item, NOT first-only
        assert ws.cell(row=3, column=2).value == "Purchase"
        assert ws.cell(row=3, column=19).value in (None, "")
        assert ws.cell(row=3, column=21).value == "Raw Jute TD-6"
        assert ws.cell(row=3, column=72).value in (None, "")

        # Row 4 — TDS row: Additional Ledger + negative amount in col 38
        assert (
            ws.cell(row=4, column=36).value
            == "TDS ON PURCHASE OF GOODS (194Q)"
        )
        assert ws.cell(row=4, column=38).value == -500
        # TDS row has shared cols but no item data
        assert ws.cell(row=4, column=21).value in (None, "")

        # Rows 5-7 — padding: only shared cols populated
        for pad_row in (5, 6, 7):
            assert ws.cell(row=pad_row, column=1).value == "ABC/FAC/JMR/25-26/00001"
            assert ws.cell(row=pad_row, column=2).value == "Purchase"
            assert ws.cell(row=pad_row, column=21).value in (None, "")
            assert ws.cell(row=pad_row, column=36).value in (None, "")
            assert ws.cell(row=pad_row, column=38).value in (None, "")
            assert ws.cell(row=pad_row, column=72).value in (None, "")

        # --- Check List sheet ---
        ws2 = wb["Check List"]
        cl_headers = [ws2.cell(row=1, column=c).value for c in range(1, 13)]
        assert cl_headers == [
            "company_id",
            "mrdate",
            "mr_print_no",
            "invoice_no",
            "invoice_date",
            "supp_name",
            "supptally",
            "jute_quality",
            "qualitytally",
            "godowsntally",
            "claimtally",
            "purtally",
        ]
        assert ws2.max_row == 3  # 1 header + 2 data rows
        assert ws2.cell(row=2, column=1).value == 1  # co_id echoed as int
        assert ws2.cell(row=2, column=6).value == "ACME JUTE TRADERS"
        assert ws2.cell(row=2, column=11).value == "Claim on Gross Purchase"
        assert ws2.cell(row=2, column=12).value == "Purchase of Raw Jute"

    def test_over_cap_returns_400(self):
        """5000 MRs × 2 items each -> 5000×6 = 30000 rows -> 400."""
        items_per_mr_result = MagicMock()
        items_per_mr_result.fetchall.return_value = [
            _mr_count_row(i, 2) for i in range(1, 5001)
        ]

        self._mock_session.execute.side_effect = [items_per_mr_result]

        response = client.get(
            "/api/juteReports/tally-download"
            "?co_id=1&date_from=2025-04-01&date_to=2026-03-31"
        )

        assert response.status_code == 400
        detail = response.json().get("detail", "")
        assert "max 10000 rows" in detail.lower() or "10000" in detail
        assert "30000" in detail

    def test_empty_range_returns_xlsx_with_headers_only(self):
        """No SQL rows -> xlsx with headers only on both sheets."""
        items_per_mr_result = MagicMock()
        items_per_mr_result.fetchall.return_value = []

        tally_result = MagicMock()
        tally_result.fetchall.return_value = []

        check_result = MagicMock()
        check_result.fetchall.return_value = []

        self._mock_session.execute.side_effect = [
            items_per_mr_result,
            tally_result,
            check_result,
        ]

        response = client.get(
            "/api/juteReports/tally-download"
            "?co_id=1&date_from=2026-02-01&date_to=2026-02-02"
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(XLSX_MEDIA_TYPE)

        wb = openpyxl.load_workbook(BytesIO(response.content))
        assert wb.sheetnames == ["Purchase", "Check List"]

        ws = wb["Purchase"]
        # Headers only — row 1 populated, no data rows.
        headers = [ws.cell(row=1, column=c).value for c in range(1, 74)]
        assert headers == EXPECTED_HEADERS
        assert ws.max_row == 1

        ws2 = wb["Check List"]
        assert ws2.max_row == 1
        assert ws2.cell(row=1, column=1).value == "company_id"

    def test_tds_zero_emits_padding_not_tds_row(self):
        """When TDS Deducted is 0, the TDS slot becomes a padding row."""
        items_per_mr_result = MagicMock()
        items_per_mr_result.fetchall.return_value = [_mr_count_row(1, 1)]

        tally_result = MagicMock()
        tally_result.fetchall.return_value = [
            _make_tally_row(**{"TDS Deducted": 0}),
        ]

        check_result = MagicMock()
        check_result.fetchall.return_value = [_make_check_row()]

        self._mock_session.execute.side_effect = [
            items_per_mr_result,
            tally_result,
            check_result,
        ]

        response = client.get(
            "/api/juteReports/tally-download"
            "?co_id=1&date_from=2026-01-01&date_to=2026-01-31"
        )

        assert response.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb["Purchase"]
        # 1 header + 6 block rows = 7
        assert ws.max_row == 7
        # No TDS row anywhere in the block (col 36 / col 38 empty on every
        # non-header row).
        for r in range(2, 8):
            assert ws.cell(row=r, column=36).value in (None, "")
            assert ws.cell(row=r, column=38).value in (None, "")
