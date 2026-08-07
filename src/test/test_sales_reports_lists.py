"""
Tests for the Sales Order list and Sales Invoice list report endpoints.
Covers JSON + xlsx variants in src/sales/reports.py.
"""

from io import BytesIO
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import get_tenant_db
from src.main import app

client = TestClient(app)

_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _mock_row(mapping: dict):
    row = MagicMock()
    row._mapping = mapping
    return row


def _so_row(**overrides):
    base = {
        "sales_order_id": 101,
        "sales_no": 101,
        "sales_no_formatted": "ABC/HQ/SO/25-26/101",
        "sales_order_date": "2026-04-15",
        "branch_name": "HQ",
        "party_name": "TEST PARTY",
        "quotation_no": "Q-001",
        "invoice_type": 1,
        "status_name": "Approved",
        "status_id": 3,
        "approval_level": 0,
        "gross_amount": 50000.0,
        "net_amount": 49500.0,
    }
    base.update(overrides)
    return base


def _inv_row(**overrides):
    base = {
        "invoice_id": 501,
        "invoice_no": 501,
        "invoice_no_formatted": "ABC/HQ/SI/25-26/501",
        "invoice_date": "2026-04-20",
        "branch_name": "HQ",
        "party_name": "TEST PARTY",
        "invoice_type": 2,
        "status_name": "Open",
        "status_id": 1,
        "approval_level": 0,
        "challan_no": "C-9",
        "challan_date": "2026-04-19",
        "due_date": "2026-05-20",
        "invoice_amount": 11800.0,
        "tax_amount": 1800.0,
        "tax_payable": 1800.0,
        "round_off": 0.0,
    }
    base.update(overrides)
    return base


class _ListReportTestBase:
    @pytest.fixture(autouse=True)
    def setup_overrides(self):
        self._mock_session = MagicMock()
        app.dependency_overrides[get_current_user_with_refresh] = (
            lambda: {"user_id": 1}
        )
        app.dependency_overrides[get_tenant_db] = lambda: self._mock_session
        yield
        app.dependency_overrides.clear()


class TestSalesOrderListReport(_ListReportTestBase):
    """GET /api/salesReports/sales-order-list (JSON)."""

    URL = "/api/salesReports/sales-order-list"

    def test_missing_co_id_returns_400(self):
        r = client.get(f"{self.URL}?date_from=2026-04-01&date_to=2026-05-06")
        assert r.status_code == 400
        assert "co_id" in r.json().get("detail", "").lower()

    def test_missing_date_from_returns_400(self):
        r = client.get(f"{self.URL}?co_id=1&date_to=2026-05-06")
        assert r.status_code == 400
        assert "date_from" in r.json().get("detail", "").lower()

    def test_bad_date_format_returns_400(self):
        r = client.get(
            f"{self.URL}?co_id=1&date_from=01-04-2026&date_to=2026-05-06"
        )
        assert r.status_code == 400
        assert "date_from" in r.json().get("detail", "").lower()

    def test_invalid_branch_id_returns_400(self):
        r = client.get(
            f"{self.URL}?co_id=1&branch_id=abc"
            "&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 400
        assert "branch_id" in r.json().get("detail", "").lower()

    def test_happy_path_returns_data(self):
        rows = [_mock_row(_so_row()),
                _mock_row(_so_row(sales_order_id=102, sales_no=102))]
        execute_result = MagicMock()
        execute_result.fetchall.return_value = rows
        self._mock_session.execute.return_value = execute_result

        r = client.get(
            f"{self.URL}?co_id=1&branch_id=1"
            "&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 200
        body = r.json()
        assert "data" in body
        assert len(body["data"]) == 2
        first = body["data"][0]
        assert first["sales_no_formatted"] == "ABC/HQ/SO/25-26/101"
        assert first["gross_amount"] == 50000.0
        assert first["net_amount"] == 49500.0
        assert first["status_name"] == "Approved"

    def test_branch_id_is_passed_to_query_as_int(self):
        execute_result = MagicMock()
        execute_result.fetchall.return_value = []
        self._mock_session.execute.return_value = execute_result

        r = client.get(
            f"{self.URL}?co_id=1&branch_id=7"
            "&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 200
        _, params = self._mock_session.execute.call_args[0]
        assert params["co_id"] == 1
        assert params["branch_id"] == 7
        assert params["date_from"] == "2026-04-01"
        assert params["date_to"] == "2026-05-06"

    def test_omit_branch_passes_none(self):
        execute_result = MagicMock()
        execute_result.fetchall.return_value = []
        self._mock_session.execute.return_value = execute_result

        r = client.get(
            f"{self.URL}?co_id=1"
            "&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 200
        _, params = self._mock_session.execute.call_args[0]
        assert params["branch_id"] is None


class TestSalesOrderListDownload(_ListReportTestBase):
    """GET /api/salesReports/sales-order-list-download (xlsx)."""

    URL = "/api/salesReports/sales-order-list-download"

    def test_happy_path_returns_xlsx(self):
        rows = [_mock_row(_so_row())]
        execute_result = MagicMock()
        execute_result.fetchall.return_value = rows
        self._mock_session.execute.return_value = execute_result

        r = client.get(
            f"{self.URL}?co_id=1&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(_XLSX_MIME)
        assert ".xlsx" in r.headers.get("content-disposition", "")

        wb = load_workbook(BytesIO(r.content))
        assert wb.sheetnames == ["SalesOrders"]
        ws = wb["SalesOrders"]
        assert ws.max_row == 2
        assert ws.max_column == 11
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert headers[0] == "Sales No"
        assert "SO Number" in headers
        assert "Gross Amount" in headers


class TestInvoiceListReport(_ListReportTestBase):
    """GET /api/salesReports/invoice-list (JSON)."""

    URL = "/api/salesReports/invoice-list"

    def test_missing_co_id_returns_400(self):
        r = client.get(f"{self.URL}?date_from=2026-04-01&date_to=2026-05-06")
        assert r.status_code == 400

    def test_missing_date_to_returns_400(self):
        r = client.get(f"{self.URL}?co_id=1&date_from=2026-04-01")
        assert r.status_code == 400
        assert "date_to" in r.json().get("detail", "").lower()

    def test_happy_path_returns_data(self):
        rows = [_mock_row(_inv_row())]
        execute_result = MagicMock()
        execute_result.fetchall.return_value = rows
        self._mock_session.execute.return_value = execute_result

        r = client.get(
            f"{self.URL}?co_id=1&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 200
        body = r.json()
        assert len(body["data"]) == 1
        first = body["data"][0]
        assert first["invoice_no_formatted"] == "ABC/HQ/SI/25-26/501"
        assert first["invoice_amount"] == 11800.0
        assert first["tax_amount"] == 1800.0
        assert first["round_off"] == 0.0


class TestInvoiceListDownload(_ListReportTestBase):
    """GET /api/salesReports/invoice-list-download (xlsx)."""

    URL = "/api/salesReports/invoice-list-download"

    def test_happy_path_returns_xlsx(self):
        rows = [_mock_row(_inv_row())]
        execute_result = MagicMock()
        execute_result.fetchall.return_value = rows
        self._mock_session.execute.return_value = execute_result

        r = client.get(
            f"{self.URL}?co_id=1&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(_XLSX_MIME)

        wb = load_workbook(BytesIO(r.content))
        assert wb.sheetnames == ["Invoices"]
        ws = wb["Invoices"]
        assert ws.max_row == 2
        assert ws.max_column == 15
        headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        assert headers[0] == "Invoice No"
        assert "Invoice Number" in headers
        assert "Tax Payable" in headers
        assert "Round Off" in headers

    def test_empty_range_returns_xlsx_with_header_only(self):
        execute_result = MagicMock()
        execute_result.fetchall.return_value = []
        self._mock_session.execute.return_value = execute_result

        r = client.get(
            f"{self.URL}?co_id=1&date_from=2026-04-01&date_to=2026-05-06"
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        ws = wb["Invoices"]
        assert ws.max_row == 1
