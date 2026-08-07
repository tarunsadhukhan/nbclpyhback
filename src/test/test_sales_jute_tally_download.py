"""
Tests for the Sales Raw Jute Tally xlsx download endpoint.
Tests for src/sales/reports.py @ GET /api/salesReports/jute-tally-download
"""

from io import BytesIO
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import get_tenant_db
from src.main import app

client = TestClient(app)


# Canonical 31-column Tally header — MUST match reports.py exactly. Kept
# duplicated here so the test is sensitive to any drift in column order,
# misspelling, or punctuation.
_EXPECTED_HEADERS = [
    "Voucher Type Name",
    "Voucher Date",
    "Voucher Number",
    "Ledger Name",
    "Ledger Amount",
    "Ledger Amount Dr/Cr",
    "Despatch Doc no.",
    "Despath though",          # (sic) misspelling
    "Destination",
    "other references",
    "Item Name",
    "Godown",
    "Batch/Lotno.",            # trailing dot
    "Billed Quantity",
    "Item Rate",
    "Item Rate per",
    "Item Amount",
    "Description ?",           # space + ?
    "Change Mode",
    "Bill Amount",
    "Bill Amount - Dr/Cr",
    "Bill Name",
    "Bill Type of Ref",
    "Due or credit days",
    "Narration",
    "Address1",
    "State1",
    "Country1",
    "Address2",
    "State2",
    "Country2",
]

_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _mock_row(mapping: dict):
    row = MagicMock()
    row._mapping = mapping
    return row


def _empty_tally_mapping(**overrides):
    """Full 31-column + ordering-helper mapping with empty defaults."""
    m = {col: "" for col in _EXPECTED_HEADERS}
    m.update({
        "invoice_date": "2026-01-15",
        "invoice_no_string": "ABC/FAC/SI/25-26/101",
        "invoice_line_item_id": 0,
        "rem": "hd",
        "rem_rank": 1,
        "slno": 0,
    })
    m.update(overrides)
    return m


class TestSalesJuteTallyDownload:
    """Tests for GET /api/salesReports/jute-tally-download."""

    @pytest.fixture(autouse=True)
    def setup_overrides(self):
        """Override FastAPI dependencies for all endpoint tests."""
        self._mock_session = MagicMock()
        app.dependency_overrides[get_current_user_with_refresh] = (
            lambda: {"user_id": 1}
        )
        app.dependency_overrides[get_tenant_db] = lambda: self._mock_session
        yield
        app.dependency_overrides.clear()

    # ----- validation -----

    def test_missing_co_id_returns_400(self):
        response = client.get(
            "/api/salesReports/jute-tally-download"
            "?date_from=2026-01-01&date_to=2026-01-31"
        )
        assert response.status_code == 400
        assert "co_id" in response.json().get("detail", "").lower()

    def test_missing_date_from_returns_400(self):
        response = client.get(
            "/api/salesReports/jute-tally-download?co_id=1&date_to=2026-01-31"
        )
        assert response.status_code == 400
        assert "date_from" in response.json().get("detail", "").lower()

    def test_bad_date_format_returns_400(self):
        response = client.get(
            "/api/salesReports/jute-tally-download"
            "?co_id=1&date_from=01-01-2026&date_to=2026-01-31"
        )
        assert response.status_code == 400
        assert "date_from" in response.json().get("detail", "").lower()

    # ----- happy path -----

    def test_happy_path_four_rows_returns_xlsx(self):
        """
        Mock four rows for the same invoice:
          - hd (header debit)
          - ln #1 (first line credit — keeps Ledger fields)
          - ln #2 (second line credit — Ledger fields MUST be blanked)
          - oth (header claim debit, negative)

        Assert xlsx response structure matches the reference file.
        """
        hd = _mock_row(_empty_tally_mapping(**{
            "Voucher Type Name": "Sales",
            "Voucher Date": "15-Jan-2026",
            "Voucher Number": "ABC/FAC/SI/25-26/101",
            "Ledger Name": "Acme Jute Traders",
            "Ledger Amount": 100000.0,
            "Ledger Amount Dr/Cr": "Dr",
            "Despatch Doc no.": " Ch.No: CH/001",
            "Despath though": "Truck No: WB12AB1234",
            "Destination": "Kolkata Depot",
            "other references": "ABC/FAC/JMR/25-26/00001",
            "Item Name": " ",
            "Godown": " ",
            "Change Mode": "Item Invoice",
            "Bill Amount": 100000.0,
            "Bill Amount - Dr/Cr": "Dr",
            "Bill Name": "ABC/FAC/SI/25-26/101 / ABC/FAC/JMR/25-26/00001",
            "Bill Type of Ref": "New Ref",
            "Due or credit days": 180,
            "Narration": "Being ... Kg. Raw Jute Sold ...",
            "Address1": "12 Main Road",
            "State1": "West Bengal",
            "Country1": "India",
            "Address2": "12 Main Road",
            "State2": "West Bengal",
            "Country2": "India",
            "rem": "hd",
            "rem_rank": 1,
            "invoice_line_item_id": 0,
        }))

        ln1 = _mock_row(_empty_tally_mapping(**{
            "Ledger Name": "Sale of Raw Jute",
            "Ledger Amount": 100000.0,
            "Ledger Amount Dr/Cr": "Cr",
            "Item Name": "TD6 (J)",
            "Godown": "Main Factory",
            "Batch/Lotno.": "ABC/FAC/JMR/25-26/00001",
            "Billed Quantity": 50000.0,
            "Item Rate": 2.0,
            "Item Rate per": "kg",
            "Item Amount": 100000.0,
            "Description ?": "Raw Jute: 100 Bales",
            "rem": "ln",
            "rem_rank": 2,
            "invoice_line_item_id": 5001,
            "slno": 1,
        }))

        ln2 = _mock_row(_empty_tally_mapping(**{
            # SQL emits Ledger fields on every ln row — post-processor must blank.
            "Ledger Name": "Sale of Raw Jute",
            "Ledger Amount": 100000.0,
            "Ledger Amount Dr/Cr": "Cr",
            "Item Name": "TD7 (J)",
            "Godown": "Main Factory",
            "Batch/Lotno.": "ABC/FAC/JMR/25-26/00001",
            "Billed Quantity": 25000.0,
            "Item Rate": 2.5,
            "Item Rate per": "kg",
            "Item Amount": 62500.0,
            "Description ?": "Raw Jute: 50 Bales",
            "rem": "ln",
            "rem_rank": 2,
            "invoice_line_item_id": 5002,
            "slno": 2,
        }))

        oth = _mock_row(_empty_tally_mapping(**{
            "Ledger Name": "Claim on Gross Sales",
            "Ledger Amount": -100,
            "Ledger Amount Dr/Cr": "",
            "Item Name": " ",
            "Godown": " ",
            "rem": "oth",
            "rem_rank": 3,
            "invoice_line_item_id": 0,
        }))

        execute_result = MagicMock()
        execute_result.fetchall.return_value = [hd, ln1, ln2, oth]
        self._mock_session.execute.return_value = execute_result

        response = client.get(
            "/api/salesReports/jute-tally-download"
            "?co_id=1&branch_id=1&date_from=2026-01-01&date_to=2026-01-31"
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(_XLSX_MIME)
        cd = response.headers.get("content-disposition", "")
        assert ".xlsx" in cd
        assert "attachment" in cd.lower()

        wb = load_workbook(BytesIO(response.content))
        assert wb.sheetnames == ["Sales"]
        ws = wb["Sales"]
        # header + 4 data rows
        assert ws.max_row == 5
        assert ws.max_column == 31

        # Header row exactly matches spec (including misspellings).
        headers = [ws.cell(row=1, column=c).value for c in range(1, 32)]
        assert headers == _EXPECTED_HEADERS

        col = {h: i + 1 for i, h in enumerate(_EXPECTED_HEADERS)}

        def cell(row, name):
            return ws.cell(row=row, column=col[name]).value

        # Row 2 — hd
        assert cell(2, "Voucher Type Name") == "Sales"
        assert cell(2, "Ledger Amount Dr/Cr") == "Dr"
        bill_name = cell(2, "Bill Name")
        assert bill_name is not None and " / " in bill_name

        # Row 3 — first ln (keeps Ledger fields)
        assert cell(3, "Ledger Name") == "Sale of Raw Jute"
        assert cell(3, "Ledger Amount Dr/Cr") == "Cr"
        assert cell(3, "Item Name") == "TD6 (J)"
        assert cell(3, "Item Rate per") == "kg"

        # Row 4 — second ln (Ledger fields MUST be blank/None)
        # openpyxl treats an empty string written to a cell as None on read-back.
        assert cell(4, "Ledger Name") in (None, "")
        assert cell(4, "Ledger Amount") in (None, "")
        assert cell(4, "Ledger Amount Dr/Cr") in (None, "")
        assert cell(4, "Item Name") == "TD7 (J)"

        # Row 5 — oth (claim)
        assert cell(5, "Ledger Name") == "Claim on Gross Sales"
        assert cell(5, "Ledger Amount") == -100
        assert cell(5, "Ledger Amount Dr/Cr") in (None, "")

    def test_empty_range_returns_xlsx_with_header_only(self):
        execute_result = MagicMock()
        execute_result.fetchall.return_value = []
        self._mock_session.execute.return_value = execute_result

        response = client.get(
            "/api/salesReports/jute-tally-download"
            "?co_id=1&date_from=2026-01-01&date_to=2026-01-31"
        )
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(_XLSX_MIME)

        wb = load_workbook(BytesIO(response.content))
        assert wb.sheetnames == ["Sales"]
        ws = wb["Sales"]
        # Only the header row.
        assert ws.max_row == 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, 32)]
        assert headers == _EXPECTED_HEADERS

    def test_over_cap_returns_400(self):
        """Result sets over 10,000 rows must be rejected with 400."""
        rows = [
            _mock_row(_empty_tally_mapping(
                invoice_no_string=f"INV/{i}",
                rem="hd",
                rem_rank=1,
            ))
            for i in range(10001)
        ]

        execute_result = MagicMock()
        execute_result.fetchall.return_value = rows
        self._mock_session.execute.return_value = execute_result

        response = client.get(
            "/api/salesReports/jute-tally-download"
            "?co_id=1&date_from=2026-01-01&date_to=2026-01-31"
        )

        assert response.status_code == 400
        detail = response.json().get("detail", "").lower()
        assert "too large" in detail or "max" in detail
