"""Import NBCL's legacy WORK MUSTER workbook into the HRMS employee tables.

Loads Sheet1 of "WORK MUSTER.xlsx" (one row per worker) into:
    hrms_ed_personal_details / _official_details / _contact_details /
    _address_details / _bank_details / _pf / _esi

Masters (dept_mst, sub_dept_mst, category_mst, designation_mst) are expected to
exist already; this script only adds the fallback "UNASSIGNED" rows and the
handful of occupations the OCCUPATION LIST workbook was missing.

Dry-run by default — pass --commit to actually write.

    python scripts/import_worker_muster.py --file "e:/test/nbcl/data/WORK MUSTER.xlsx"
    python scripts/import_worker_muster.py --file "..." --commit
"""
import argparse
import os
import re
from collections import Counter
from datetime import datetime

import openpyxl
import pymysql
from dotenv import load_dotenv

# ─── Mapping constants ──────────────────────────────────────────────

STATUS_JOINED = 35
STATUS_RESIGNED = 39
COUNTRY_INDIA = 1
UNASSIGNED = "UNASSIGNED"

GENDER = {"M": "Male", "F": "Female"}
MARITAL = {"Y": 1, "N": 0}  # 0=Single 1=Married (see PersonalStep.tsx)
PAYMENT_MODE = {"CASH": 0, "CHEQUE": 1, "NEFT": 2, "UPI": 3}  # see BankStep.tsx
RELIGION = {"HINDU": "Hindu", "HINDUSIM": "Hindu", "MUSLIM": "Muslim"}
# Legacy sheet's category spelling → category_mst.cata_desc
CATEGORY_ALIAS = {"02 - SPACEL": "02 - SPECIAL"}
# Occupations used by the muster that the OCCUPATION LIST workbook never had.
# (department name as it appears in the muster, designation name)
MISSING_DESIGNATIONS = [
    ("SPINNING", "Apprentice/trainee"),
    ("WINDING", "CLERK"),
    ("WINDING", "GD5-WEAVING"),
    ("GODOWN - 5 CHAINA LOOM", "GD5-SWEEPER"),
    ("GODOWN - 5 CHAINA LOOM", "GD5-HELPER"),
    ("ESI", "Assistant/ Supervisor"),
]


# ─── Cell helpers ───────────────────────────────────────────────────

def s(v, maxlen=None):
    """Excel cell → clean string or None. 'N/A' and blanks are nulls."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    t = str(v).replace("\xa0", " ").strip()
    t = re.sub(r"\s+", " ", t)
    if not t or t.upper() in ("N/A", "NA", "-"):
        return None
    return t[:maxlen] if maxlen else t


def dt(v):
    """Excel cell → date or None. 1900-01-01 is the legacy null sentinel."""
    if not isinstance(v, datetime):
        return None
    return None if v.year <= 1900 else v.date()


def num(v):
    """Excel cell → positive int or None (0 is the legacy 'unset' zip)."""
    t = s(v)
    try:
        n = int(float(t))
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def pin(v):
    """Excel cell → 6-digit Indian PIN, or 0. The legacy sheet has phone
    numbers and truncated pins in the zip columns."""
    n = num(v)
    return n if n and 100000 <= n <= 999999 else 0


def norm(v):
    return re.sub(r"\s+", " ", str(v or "").replace("\xa0", " ")).strip().upper()


def split_name(full):
    """'A B C D' → ('A', 'B C', 'D'). Single word → (word, None, None)."""
    parts = (full or "").split()
    if not parts:
        return ("UNKNOWN", None, None)
    first = parts[0][:50]
    last = parts[-1][:50] if len(parts) > 1 else None
    middle = " ".join(parts[1:-1])[:50] or None
    return (first, middle, last)


# ─── Master lookups ─────────────────────────────────────────────────

def ensure_masters(cur, branch_id, dry):
    """Create the UNASSIGNED fallbacks + missing designations. Idempotent."""
    created = Counter()

    def one(sql, params):
        cur.execute(sql, params)
        row = cur.fetchone()
        return row[0] if row else None

    # dept + sub-dept + designation + category fallbacks
    dept_id = one("SELECT dept_id FROM dept_mst WHERE dept_desc=%s AND branch_id=%s",
                  (UNASSIGNED, branch_id))
    if not dept_id:
        cur.execute(
            "INSERT INTO dept_mst (branch_id, created_by, dept_desc, dept_code, order_id,"
            " created_date, updated_by, updated_date_time, Worker_staff)"
            " VALUES (%s, 1, %s, '999', 999, NOW(), 1, NOW(), 1)", (branch_id, UNASSIGNED))
        dept_id = cur.lastrowid
        created["dept_mst"] += 1

    sub_dept_id = one("SELECT sub_dept_id FROM sub_dept_mst WHERE dept_id=%s", (dept_id,))
    if not sub_dept_id:
        cur.execute(
            "INSERT INTO sub_dept_mst (updated_by, sub_dept_code, sub_dept_desc, dept_id,"
            " updated_date_time, order_no) VALUES (1, '999', %s, %s, NOW(), 999)",
            (UNASSIGNED, dept_id))
        sub_dept_id = cur.lastrowid
        created["sub_dept_mst"] += 1

    desig_id = one("SELECT designation_id FROM designation_mst WHERE dept_id=%s AND desig=%s",
                   (dept_id, UNASSIGNED))
    if not desig_id:
        cur.execute(
            "INSERT INTO designation_mst (branch_id, dept_id, desig, time_piece, piece_rate_type,"
            " active, updated_by, updated_date_time) VALUES (%s, %s, %s, 'T', 'T', 1, 1, NOW())",
            (branch_id, dept_id, UNASSIGNED))
        desig_id = cur.lastrowid
        created["designation_mst"] += 1

    grade_id = one("SELECT grade_id FROM grade_table WHERE grade_type=0 ORDER BY grade_id LIMIT 1", ())
    cata_id = one("SELECT cata_id FROM category_mst WHERE cata_desc=%s AND branch_id=%s",
                  (UNASSIGNED, branch_id))
    if not cata_id:
        cur.execute(
            "INSERT INTO category_mst (cata_code, cata_desc, branch_id, updated_by,"
            " updated_date_time, grade_id) VALUES ('99', %s, %s, 1, NOW(), %s)",
            (UNASSIGNED, branch_id, grade_id))
        cata_id = cur.lastrowid
        created["category_mst"] += 1

    # occupations the OCCUPATION LIST workbook was missing
    for dept_name, desig in MISSING_DESIGNATIONS:
        d_id = one("SELECT dept_id FROM dept_mst WHERE UPPER(dept_desc)=%s AND branch_id=%s",
                   (norm(dept_name), branch_id))
        if not d_id:
            raise SystemExit(f"department {dept_name!r} not found in dept_mst")
        if not one("SELECT designation_id FROM designation_mst WHERE dept_id=%s AND desig=%s",
                   (d_id, desig)):
            cur.execute(
                "INSERT INTO designation_mst (branch_id, dept_id, desig, time_piece,"
                " piece_rate_type, active, updated_by, updated_date_time)"
                " VALUES (%s, %s, %s, 'T', 'T', 1, 1, NOW())", (branch_id, d_id, desig))
            created["designation_mst"] += 1

    if dry:
        # roll the fallback rows back — they are re-created on the real run
        cur.connection.rollback()
    return {"sub_dept_id": sub_dept_id, "designation_id": desig_id, "cata_id": cata_id,
            "created": created}


def build_lookups(cur, branch_id, occupation_file):
    """Return the maps used to resolve every muster row to master IDs."""
    cur.execute(
        "SELECT d.designation_id, d.dept_id, dm.dept_desc, d.desig"
        " FROM designation_mst d JOIN dept_mst dm ON dm.dept_id = d.dept_id"
        " WHERE d.branch_id = %s ORDER BY d.designation_id", (branch_id,))
    desigs = cur.fetchall()

    # designation_mst was loaded from OCCUPATION LIST.xlsx in file order and has
    # no OcpCode column — recover OcpCode → designation_id positionally, and
    # verify every name/dept lines up before trusting it.
    wb = openpyxl.load_workbook(occupation_file, read_only=True, data_only=True)
    it = wb["WORKER"].iter_rows(values_only=True)
    next(it)
    occ = [r for r in it if r[0]]
    if len(occ) > len(desigs):
        raise SystemExit("OCCUPATION LIST has more rows than designation_mst")
    ocp2desig = {}
    for i, o in enumerate(occ):
        if norm(o[1]) != norm(desigs[i][3]):
            raise SystemExit(f"designation_mst drifted from OCCUPATION LIST at row {i}: "
                             f"{o[1]!r} vs {desigs[i][3]!r}")
        ocp2desig[int(o[0])] = desigs[i][0]
    wb.close()

    # (dept name, designation name) → designation_id, for the rows whose OcpCode
    # is not in the occupation list
    by_name = {(norm(d[2]), norm(d[3])): d[0] for d in desigs}
    desig_dept = {d[0]: d[1] for d in desigs}

    cur.execute("SELECT sub_dept_id, dept_id FROM sub_dept_mst")
    dept2sub = dict((dept, sub) for sub, dept in cur.fetchall())
    cur.execute("SELECT dept_id, dept_desc FROM dept_mst WHERE branch_id = %s", (branch_id,))
    name2dept = {}
    for dept_id, desc in cur.fetchall():
        name2dept.setdefault(norm(desc), dept_id)

    cur.execute("SELECT cata_id, cata_desc FROM category_mst WHERE branch_id = %s"
                " AND grade_id = (SELECT grade_id FROM grade_table WHERE grade_type=0"
                " ORDER BY grade_id LIMIT 1)", (branch_id,))
    categories = {}
    for cata_id, desc in cur.fetchall():
        categories.setdefault(norm(desc), cata_id)

    cur.execute("SELECT state_id, state FROM state_mst WHERE country_id = %s", (COUNTRY_INDIA,))
    states = {norm(n): i for i, n in cur.fetchall()}

    return {"ocp2desig": ocp2desig, "by_name": by_name, "desig_dept": desig_dept,
            "dept2sub": dept2sub, "name2dept": name2dept, "categories": categories,
            "states": states}


# ─── Row → table payloads ───────────────────────────────────────────

def address_rows(r, eb_id, states, prefix, addr_type, is_corr):
    line1 = s(r[f"{prefix}Add1"], 150) or s(r[f"{prefix}Add2"], 150)
    if not line1:
        return None
    rest = [s(r[f"{prefix}Add2"]), s(r[f"{prefix}Add3"])]
    if s(r[f"{prefix}Add1"]) is None:
        rest = [s(r[f"{prefix}Add3"])]
    line2 = ", ".join(x for x in rest if x)[:150] or None
    return (eb_id, addr_type, COUNTRY_INDIA, states.get(norm(r[f"{prefix}State"])),
            s(r[f"{prefix}City"], 150), line1, line2, pin(r[f"{prefix}Zip"]), 1, is_corr)


def build(rows, lk, fb, branch_id, first_eb_id, skip_codes):
    """Turn muster rows into per-table insert tuples. Returns (payloads, report)."""
    out = {k: [] for k in ("personal", "official", "contact", "address", "bank", "pf", "esi")}
    rep = Counter()
    eb_id = first_eb_id

    for r in rows:
        emp_code = s(r["C.NO."], 20)
        if not emp_code:
            rep["skipped_no_emp_code"] += 1
            continue
        if emp_code in skip_codes:
            rep["skipped_already_imported"] += 1
            continue
        skip_codes.add(emp_code)

        active = 1 if norm(r["Active"]) == "Y" else 0
        status_id = STATUS_JOINED if active else STATUS_RESIGNED
        first, middle, last = split_name(s(r["EName"]))
        mobile = s(r["CMobile"], 15) or s(r["PMobile"], 15)

        # ── designation drives the department: designation_mst is dept-scoped,
        # and the UI's sub-dept → designation cascade only works if they agree.
        desig_id = lk["ocp2desig"].get(num(r["OcpCode"]))
        dept_name = norm(r["DeptName"])
        if desig_id is None and dept_name:
            desig_id = lk["by_name"].get((dept_name, norm(r["Designation"])))
        if desig_id is not None:
            dept_id = lk["desig_dept"][desig_id]
            if dept_name and lk["name2dept"].get(dept_name) not in (None, dept_id):
                rep["dept_overridden_by_occupation"] += 1
            sub_dept_id = lk["dept2sub"].get(dept_id, fb["sub_dept_id"])
        else:
            desig_id = fb["designation_id"]
            sub_dept_id = lk["dept2sub"].get(lk["name2dept"].get(dept_name), fb["sub_dept_id"])
            rep["designation_unassigned"] += 1
        if sub_dept_id == fb["sub_dept_id"]:
            rep["sub_dept_unassigned"] += 1

        cata_desc = norm(CATEGORY_ALIAS.get(s(r["EGroup"]), s(r["EGroup"])))
        cata_id = lk["categories"].get(cata_desc, fb["cata_id"])
        if cata_id == fb["cata_id"]:
            rep["category_unassigned"] += 1

        out["personal"].append((
            eb_id, first, middle, last, GENDER.get(norm(r["Sex"])), dt(r["DoB"]), mobile,
            MARITAL.get(norm(r["MStatus"])), COUNTRY_INDIA,
            RELIGION.get(norm(r["Religion"]), s(r["Religion"], 100)),
            s(r["Father"], 100), s(r["PassPortNo"], 20), s(r["DrivingLicense"], 50),
            s(r["PANNo"], 15), s(r["AadharNo"], 20), s(r["VoterID"], 20),
            branch_id, active, status_id))

        out["official"].append((
            eb_id, sub_dept_id, cata_id, desig_id, branch_id, dt(r["DoJ"]), 0, emp_code))

        if mobile:
            alt = s(r["PMobile"], 15)
            out["contact"].append((eb_id, mobile, alt if alt and alt != mobile else None))

        perm = address_rows(r, eb_id, lk["states"], "P", 1, 0)
        curr = address_rows(r, eb_id, lk["states"], "C", 2, 1)
        if perm and curr and perm[3:8] == curr[3:8]:  # state..pin equal, ignore the type/corr flags
            # identical current & permanent — keep one row, flagged as correspondence
            out["address"].append(perm[:8] + (1, 1))  # active=1, is_correspondent=1
        else:
            out["address"].extend(a for a in (perm, curr) if a)

        acc, ifsc, pmode = s(r["AccNo"], 20), s(r["IFSC"], 15), PAYMENT_MODE.get(norm(r["PMode"]))
        if acc or ifsc or pmode is not None:
            out["bank"].append((eb_id, ifsc or "", acc or "", s(r["BankName"], 100) or "",
                                s(r["Branch"], 300) or "", pmode, s(r["BeneficiaryName"], 100)))
            if not acc:
                rep["bank_row_without_account"] += 1

        uan, pension = s(r["UAN"], 50), s(r["PensionNo"], 10)
        if uan or pension:
            # the muster carries no PF number — only UAN and the pension record
            out["pf"].append((eb_id, uan or "", pension, dt(r["PensionDate"])))

        esi_no = s(r["ESINo"], 50)
        if esi_no:
            out["esi"].append((eb_id, esi_no, dt(r["ESIDate"])))

        rep["employees"] += 1
        rep["active" if active else "inactive"] += 1
        eb_id += 1

    return out, rep


# (sql, trailing constant values). Every value must be a %s placeholder or
# pymysql refuses to batch executemany into one multi-row INSERT and instead
# issues one round trip per row — which the server kills at this volume.
INSERTS = {
    "personal": (
        "INSERT INTO hrms_ed_personal_details (eb_id, first_name, middle_name, last_name,"
        " gender, date_of_birth, mobile_no, marital_status, country_id, relegion_name,"
        " father_spouse_name, passport_no, driving_licence_no, pan_no, aadhar_no,"
        " voter_card_no, branch_id, active, status_id, updated_by)"
        " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (1,)),
    "official": (
        "INSERT INTO hrms_ed_official_details (eb_id, sub_dept_id, catagory_id, designation_id,"
        " branch_id, date_of_join, minimum_working_commitment, emp_code, active, updated_by)"
        " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (1, 1)),
    "contact": (
        "INSERT INTO hrms_ed_contact_details (eb_id, mobile_no, emergency_no, active, updated_by)"
        " VALUES (%s,%s,%s,%s,%s)", (1, 1)),
    "address": (
        "INSERT INTO hrms_ed_address_details (eb_id, address_type, country_id, state_id,"
        " city_name, address_line_1, address_line_2, pin_code, active, is_correspondent_address,"
        " updated_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (1,)),
    "bank": (
        "INSERT INTO hrms_ed_bank_details (eb_id, ifsc_code, bank_acc_no, bank_name,"
        " bank_branch_name, payment_mode, beneficiary_name, active, is_verified, updated_by)"
        " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (1, 0, 1)),
    "pf": (
        "INSERT INTO hrms_ed_pf (eb_id, pf_uan_no, pension_no, pension_date, pf_no,"
        " pf_previous_no, active, updated_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", ("", "", 1, 1)),
    "esi": (
        "INSERT INTO hrms_ed_esi (eb_id, esi_no, esi_date, active, updated_by)"
        " VALUES (%s,%s,%s,%s,%s)", (1, 1)),
}
CHUNK = 1000


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="WORK MUSTER.xlsx")
    ap.add_argument("--occupations", default=r"e:/test/nbcl/data/OCCUPATION LIST.xlsx")
    ap.add_argument("--db", default="nbcl")
    ap.add_argument("--branch", type=int, default=87)
    ap.add_argument("--commit", action="store_true", help="write (default: dry run)")
    args = ap.parse_args()

    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    conn = pymysql.connect(host=os.getenv("DATABASE_HOST"), user=os.getenv("DATABASE_USER"),
                           password=os.getenv("DATABASE_PASSWORD"),
                           port=int(os.getenv("DATABASE_PORT")), database=args.db,
                           autocommit=False, charset="utf8mb4")
    cur = conn.cursor()

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    it = wb["Sheet1"].iter_rows(values_only=True)
    hdr = list(next(it))
    rows = [dict(zip(hdr, r)) for r in it]
    rows = [r for r in rows if r.get("C.NO.")]
    print(f"muster rows: {len(rows)}")

    fb = ensure_masters(cur, args.branch, dry=not args.commit)
    if fb["created"]:
        print(f"masters {'created' if args.commit else 'to create'}: {dict(fb['created'])}")
    lk = build_lookups(cur, args.branch, args.occupations)
    print(f"occupation codes mapped: {len(lk['ocp2desig'])}")

    cur.execute("SELECT emp_code FROM hrms_ed_official_details WHERE branch_id = %s", (args.branch,))
    skip_codes = {c for (c,) in cur.fetchall()}
    cur.execute("SELECT COALESCE(MAX(eb_id), 0) + 1 FROM hrms_ed_personal_details")
    first_eb_id = max(cur.fetchone()[0], 1)

    payloads, rep = build(rows, lk, fb, args.branch, first_eb_id, skip_codes)
    print("\n--- report ---")
    for k, v in sorted(rep.items()):
        print(f"  {k:32} {v}")
    print("\n--- rows to insert ---")
    for k, (sql, const) in INSERTS.items():  # dict order is insert order (FK parents first)
        want = sql.count("%s")
        for t in payloads[k]:
            assert len(t) + len(const) == want, \
                f"{k}: row has {len(t)}+{len(const)} values, SQL wants {want}: {t}"
        print(f"  {k:32} {len(payloads[k])}")

    if not args.commit:
        conn.rollback()
        print("\nDRY RUN — nothing written. Re-run with --commit.")
        return

    for k, (sql, const) in INSERTS.items():
        rows_k = [t + const for t in payloads[k]]
        for i in range(0, len(rows_k), CHUNK):
            cur.executemany(sql, rows_k[i:i + CHUNK])
        print(f"  inserted {k}: {len(rows_k)}")
    conn.commit()
    print(f"\nCOMMITTED to `{args.db}` branch {args.branch}, eb_id {first_eb_id}"
          f"..{first_eb_id + rep['employees'] - 1}")


if __name__ == "__main__":
    main()
